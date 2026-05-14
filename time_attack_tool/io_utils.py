from __future__ import annotations

import csv
import io
from pathlib import Path
from typing import Iterable
import urllib.parse
import urllib.request

from openpyxl import load_workbook

from time_attack_tool.models import (
    NormalizedRecord,
    format_lap_time_ms,
    parse_lap_time_to_ms,
)


def _clean(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _pick(row: dict[str, object], *keys: str, default: str | None = None) -> str | None:
    for key in keys:
        if key in row:
            value = _clean(row.get(key))
            if value is not None:
                return value
    return default


def _is_url(value: str) -> bool:
    parsed = urllib.parse.urlparse(value)
    return parsed.scheme in {"http", "https"}


def _extract_google_sheet_id(parsed: urllib.parse.ParseResult) -> str | None:
    parts = [part for part in parsed.path.split("/") if part]
    for index, part in enumerate(parts):
        if part == "d" and index + 1 < len(parts):
            return parts[index + 1]
    return None


def _extract_gid(parsed: urllib.parse.ParseResult) -> str | None:
    query_gid = urllib.parse.parse_qs(parsed.query).get("gid")
    if query_gid and query_gid[0]:
        return query_gid[0]

    fragment_params = urllib.parse.parse_qs(parsed.fragment)
    fragment_gid = fragment_params.get("gid")
    if fragment_gid and fragment_gid[0]:
        return fragment_gid[0]

    if "gid=" in parsed.fragment:
        fragment_query = urllib.parse.parse_qs(parsed.fragment.replace("#", "", 1))
        fragment_gid = fragment_query.get("gid")
        if fragment_gid and fragment_gid[0]:
            return fragment_gid[0]

    return None


def normalize_table_source(
    source: str | Path,
    sheet_name: str | None = None,
    sheet_gid: str | None = None,
) -> str:
    raw = str(source)
    if not _is_url(raw):
        return raw

    parsed = urllib.parse.urlparse(raw)
    if "docs.google.com" not in parsed.netloc or "/spreadsheets/" not in parsed.path:
        return raw

    query = urllib.parse.parse_qs(parsed.query)
    is_csv_export = query.get("output") == ["csv"] or query.get("format") == ["csv"] or query.get("tqx") == ["out:csv"]
    if is_csv_export:
        return raw

    sheet_id = _extract_google_sheet_id(parsed)
    if not sheet_id:
        return raw

    params: dict[str, str] = {"tqx": "out:csv"}
    if sheet_name:
        params["sheet"] = sheet_name
    else:
        gid = sheet_gid or _extract_gid(parsed)
        if gid:
            params["gid"] = gid

    return f"https://docs.google.com/spreadsheets/d/{sheet_id}/gviz/tq?{urllib.parse.urlencode(params)}"


def _read_csv_text(csv_text: str) -> list[dict[str, object]]:
    return list(csv.DictReader(io.StringIO(csv_text)))


def _read_xlsx_bytes(blob: bytes, sheet_name: str | None = None) -> list[dict[str, object]]:
    workbook = load_workbook(io.BytesIO(blob), read_only=True, data_only=True)
    if sheet_name:
        worksheet = workbook[sheet_name]
    else:
        worksheet = workbook[workbook.sheetnames[0]]

    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    items: list[dict[str, object]] = []
    for values in rows[1:]:
        item = {header[index]: values[index] for index in range(len(header)) if header[index]}
        items.append(item)
    return items


def read_table(
    path: str | Path,
    sheet_name: str | None = None,
    sheet_gid: str | None = None,
) -> list[dict[str, object]]:
    source_text = normalize_table_source(path, sheet_name=sheet_name, sheet_gid=sheet_gid)
    if _is_url(source_text):
        with urllib.request.urlopen(source_text) as response:
            payload = response.read()
            content_type = response.headers.get("Content-Type", "")
        if source_text.lower().endswith((".xlsx", ".xlsm")) or "spreadsheetml" in content_type:
            return _read_xlsx_bytes(payload, sheet_name=sheet_name)
        return _read_csv_text(payload.decode("utf-8-sig"))

    source = Path(source_text)
    suffix = source.suffix.lower()
    if suffix == ".csv":
        with source.open("r", encoding="utf-8-sig", newline="") as fh:
            return list(csv.DictReader(fh))

    if suffix in {".xlsx", ".xlsm"}:
        return _read_xlsx_bytes(source.read_bytes(), sheet_name=sheet_name)

    raise ValueError(f"Unsupported input format: {source.suffix}")


def load_approved_records(
    path: str | Path,
    sheet_name: str | None = None,
    sheet_gid: str | None = None,
) -> list[NormalizedRecord]:
    rows = read_table(path, sheet_name=sheet_name, sheet_gid=sheet_gid)
    records: list[NormalizedRecord] = []
    for index, row in enumerate(rows, start=2):
        record_id = _pick(row, "record_id")
        if not record_id:
            raise ValueError(f"approved records row {index} is missing record_id")

        racer_id = _pick(row, "racer_id", "player_id")
        racer_name = _pick(row, "racer_display_name", "player_display_name", "racer_name")
        vehicle_id = _pick(row, "vehicle_id")
        vehicle_name = _pick(row, "vehicle_display_name", "vehicle_name")
        variant_id = _pick(row, "track_variant_id", "track_id")
        variant_name = _pick(row, "track_variant_name", "track_display_name", "track_name")
        route_id = _pick(row, "track_route_id")
        route_name = _pick(row, "track_route_name")

        if not racer_id or not racer_name:
            raise ValueError(f"approved records row {index} is missing racer identity")
        if not vehicle_id or not vehicle_name:
            raise ValueError(f"approved records row {index} is missing vehicle identity")
        if not variant_id or not variant_name:
            raise ValueError(f"approved records row {index} is missing track variant")

        if not route_id:
            route_id = f"{variant_id}:default"
        if not route_name:
            route_name = "Default"

        lap_time_value = _pick(row, "lap_time_ms", "lap_time", "lap_time_text")
        lap_time_ms = parse_lap_time_to_ms(lap_time_value)
        lap_time_text = _pick(row, "lap_time_text") or format_lap_time_ms(lap_time_ms)

        record = NormalizedRecord(
            record_id=record_id,
            submission_id=_pick(row, "submission_id"),
            source_type=_pick(row, "source_type", default="approved_records") or "approved_records",
            review_status=_pick(row, "review_status", default="approved") or "approved",
            record_tier=_pick(row, "record_tier", "baseline_tier", default="qualified") or "qualified",
            racer_id=racer_id,
            racer_display_name=racer_name,
            vehicle_id=vehicle_id,
            vehicle_display_name=vehicle_name,
            track_family_id=_pick(row, "track_family_id"),
            track_family_name=_pick(row, "track_family_name"),
            track_variant_id=variant_id,
            track_variant_name=variant_name,
            track_route_id=route_id,
            track_route_name=route_name,
            lap_time_ms=lap_time_ms,
            lap_time_text=lap_time_text,
            platform=_pick(row, "platform"),
            fps_value=_pick(row, "fps_value"),
            recorded_at=_pick(row, "recorded_at"),
            approved_at=_pick(row, "approved_at"),
            event_id=_pick(row, "event_id"),
            notes=_pick(row, "notes", "note"),
        )
        records.append(record)
    return records


def build_review_summary(
    path: str | Path | None,
    sheet_name: str | None = None,
    sheet_gid: str | None = None,
) -> dict[str, object] | None:
    if path is None:
        return None

    rows = read_table(path, sheet_name=sheet_name, sheet_gid=sheet_gid)
    counts: dict[str, int] = {}
    pending_items: list[dict[str, str | None]] = []

    for row in rows:
        review_status = (_pick(row, "review_status", default="submitted") or "submitted").strip()
        counts[review_status] = counts.get(review_status, 0) + 1
        if review_status in {"submitted", "needs_info"}:
            pending_items.append(
                {
                    "submission_id": _pick(row, "submission_id"),
                    "racer_name_input": _pick(row, "racer_name_input", "racer_display_name"),
                    "track_input": _pick(row, "track_input", "track_variant_name", "track_display_name"),
                    "vehicle_input": _pick(row, "vehicle_input", "vehicle_display_name"),
                    "lap_time_text": _pick(row, "lap_time_text", "lap_time"),
                    "review_status": review_status,
                }
            )

    return {
        "total_submissions": len(rows),
        "by_status": counts,
        "pending_items": pending_items,
    }


def records_to_dicts(records: Iterable[NormalizedRecord]) -> list[dict[str, object]]:
    return [record.to_dict() for record in records]
