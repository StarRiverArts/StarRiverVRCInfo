from __future__ import annotations

from pathlib import Path
from typing import Optional
import logging

from scraper.scraper import (
    fetch_worlds,
    _load_headers,
    record_row,
    enrich_visits,
    vrchat_login,
    vrchat_verify_2fa,
    vrchat_check_session,
)

try:  # optional dependency
    from openpyxl import load_workbook, Workbook  # type: ignore
except Exception:  # pragma: no cover - optional
    load_workbook = None  # type: ignore
    Workbook = None  # type: ignore

try:
    from world_info.constants import (
        BASE,
        RAW_FILE,
        USER_FILE,
        STAR_RIVER_FILE,
        TAIWAN_FILE,
    )
except ModuleNotFoundError:  # pragma: no cover - package path
    from constants import (
        BASE,
        RAW_FILE,
        USER_FILE,
        STAR_RIVER_FILE,
        TAIWAN_FILE,
    )

logger = logging.getLogger(__name__)


def load_auth_headers(cookie: Optional[str], user: Optional[str], pw: Optional[str]) -> dict:
    """Return headers for authenticated requests."""
    return _load_headers(cookie, user, pw)


def _enrich_if_needed(worlds: list[dict], headers: dict) -> list[dict]:
    """Run enrich_visits only when some worlds are missing visit data."""
    null_count = sum(1 for w in worlds if w.get("visits") is None)
    if null_count:
        logger.info(
            "%d world(s) missing visits – fetching individually to fill data…",
            null_count,
        )
        worlds = enrich_visits(worlds, headers or None)
        still_null = sum(1 for w in worlds if w.get("visits") is None)
        if still_null:
            logger.warning(
                "%d world(s) still have no visits after individual fetch "
                "(API may require a valid Cookie).",
                still_null,
            )
    return worlds


def search_keyword(keyword: str, headers: dict, limit: int = 50) -> list[dict]:
    """Fetch worlds by keyword, enriching missing visit counts."""
    worlds = fetch_worlds(keyword=keyword, limit=limit, headers=headers)
    logger.info("Fetched %d worlds for keyword '%s'", len(worlds), keyword)
    return _enrich_if_needed(worlds, headers)


def search_user(user_id: str, headers: dict, limit: int = 50) -> list[dict]:
    """Fetch worlds created by a user, enriching missing visit counts."""
    worlds = fetch_worlds(user_id=user_id, limit=limit, headers=headers)
    logger.info("Fetched %d worlds for user '%s'", len(worlds), user_id)
    return _enrich_if_needed(worlds, headers)


def search_fixed(keywords: str, headers: dict, blacklist: set[str]) -> list[dict]:
    """Fetch worlds for a comma-separated keyword list, enriching missing visits."""
    kw_list = [k.strip() for k in keywords.split(",") if k.strip()]
    all_worlds: list[dict] = []
    for kw in kw_list:
        if kw in blacklist:
            continue
        worlds = fetch_worlds(keyword=kw, limit=50, headers=headers)
        all_worlds.extend(worlds)
    logger.info("Fetched %d worlds for keywords %s", len(all_worlds), kw_list)
    return _enrich_if_needed(all_worlds, headers)


def save_worlds(worlds: list[dict], file: Path) -> None:
    """Overwrite *file* with a fresh header + one row per world.

    Always overwrites so the file reflects the current scrape result without
    accumulating duplicate rows across multiple runs.
    """
    if Workbook is None:
        logger.error("openpyxl not available; cannot save %s", file)
        return
    col_headers = [
        "爬取日期",
        "世界名稱",
        "世界ID",
        "發布日期",
        "最後更新",
        "瀏覽人次",
        "大小",
        "收藏次數",
        "熱度",
        "人氣",
        "實驗室到發布",
        "瀏覽蒐藏比",
        "距離上次更新",
        "已發布",
        "人次發布比",
    ]
    wb = Workbook()
    ws = wb.active
    ws.append(col_headers)
    for w in worlds:
        ws.append(record_row(w))
    wb.save(file)
    logger.info("Saved %d worlds to %s", len(worlds), file)
