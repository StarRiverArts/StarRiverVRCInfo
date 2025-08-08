from __future__ import annotations

import datetime as dt
from pathlib import Path
from typing import List
import logging

try:
    from openpyxl import Workbook, load_workbook  # type: ignore
except Exception:  # pragma: no cover - optional dependency
    Workbook = None  # type: ignore
    load_workbook = None  # type: ignore

from scraper.scraper import _parse_date

BASE = Path(__file__).resolve().parent
ANALYTICS_DIR = BASE.parent / "analytics"

COLUMNS = ["date", "total_worlds", "new_worlds_today"]

def update_daily_stats(source_name: str, worlds: List[dict],
                       file_path: Path | str | None = None) -> None:
    """Update daily world statistics for ``source_name``.

    Parameters
    ----------
    source_name:
        Identifier for the world source, used when ``file_path`` is not
        provided.
    worlds:
        List of world dictionaries to summarise.
    file_path:
        Optional path to the stats workbook.  If omitted, the file will be
        created inside ``ANALYTICS_DIR`` using ``source_name``.
    """
    if Workbook is None or load_workbook is None:
        logging.warning("openpyxl not available; skipping statistics update")
        return

    ANALYTICS_DIR.mkdir(exist_ok=True)
    if file_path is None:
        file_path = ANALYTICS_DIR / f"daily_stats_{source_name}.xlsx"
    else:
        file_path = Path(file_path)
        if not file_path.is_absolute():
            file_path = ANALYTICS_DIR / file_path
        file_path.parent.mkdir(parents=True, exist_ok=True)

    stats = _calculate_stats(worlds, tzinfo=dt.timezone.utc)
    wb, ws = _load_or_create_workbook(file_path)
    _write_row(ws, stats)
    wb.save(file_path)

def _calculate_stats(worlds: List[dict],
                     tzinfo: dt.tzinfo = dt.timezone.utc) -> dict:
    today = dt.datetime.now(tzinfo)
    today_str = today.strftime("%Y/%m/%d")
    today_date = today.date()
    new_today = 0
    for w in worlds:
        pub = _parse_date(w.get("publicationDate"))
        if pub and pub.astimezone(tzinfo).date() == today_date:
            new_today += 1
    return {
        "date": today_str,
        "total_worlds": len(worlds),
        "new_worlds_today": new_today,
    }

def _load_or_create_workbook(file_path: Path):
    if file_path.exists():
        wb = load_workbook(file_path)
        ws = wb.active
        header = [cell.value for cell in ws[1]]
        if header != COLUMNS:
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            wb = Workbook()
            ws = wb.active
            ws.append(COLUMNS)
            for row in rows:
                row = list(row)
                row += [""] * (len(COLUMNS) - len(row))
                ws.append(row)
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(COLUMNS)
    return wb, ws

def _write_row(ws, stats: dict) -> None:
    date_str = stats["date"]
    row_idx = None
    for idx in range(2, ws.max_row + 1):
        if ws.cell(row=idx, column=1).value == date_str:
            row_idx = idx
            break
    if row_idx is None:
        row_idx = ws.max_row + 1
    for col, key in enumerate(COLUMNS, start=1):
        ws.cell(row=row_idx, column=col, value=stats.get(key, ""))
