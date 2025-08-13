from __future__ import annotations

import json
import logging
from pathlib import Path
from typing import List, Dict

from analytics import update_daily_stats
from scraper.scraper import (
    search_worlds,
    get_user_worlds,
    update_history,
    record_row,
)
try:
    from world_info.constants import BASE, METRIC_COLS
except ModuleNotFoundError:  # pragma: no cover - package path
    from constants import BASE, METRIC_COLS

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

try:
    from openpyxl import Workbook, load_workbook  # type: ignore
except Exception:  # pragma: no cover - optional dependency
    Workbook = None  # type: ignore
    load_workbook = None  # type: ignore

CONFIG_FILE = BASE / "config" / "search_modes.json"
ANALYTICS_DIR = BASE.parent / "analytics"

def _load_taiwan_blacklist() -> set[str]:
    """Return a set of world IDs from the Taiwan blacklist if present."""
    txt_file = BASE / "blacklist_taiwan.txt"
    xlsx_file = BASE / "blacklist_taiwan.xlsx"
    if txt_file.exists():
        with open(txt_file, "r", encoding="utf-8") as f:
            return {line.strip() for line in f if line.strip()}
    if xlsx_file.exists() and load_workbook is not None:
        wb = load_workbook(xlsx_file)
        ws = wb.active
        ids = {
            str(row[0].value).strip()
            for row in ws.iter_rows(min_row=1, max_col=1)
            if row[0].value
        }
        wb.close()
        return ids
    return set()


def _save_worlds(worlds: List[dict], file_path: Path) -> None:
    """Append ``worlds`` to the Excel sheet at ``file_path``."""
    if Workbook is None or load_workbook is None:
        return
    if file_path.exists():
        wb = load_workbook(file_path)
        ws = wb.active
    else:
        file_path.parent.mkdir(parents=True, exist_ok=True)
        wb = Workbook()
        ws = wb.active
        ws.append(["爬取日期"] + METRIC_COLS)
    for w in worlds:
        ws.append(record_row(w))
    wb.save(file_path)


def _run_mode(name: str, cfg: Dict[str, object]) -> int:
    logger.info("Starting mode %s", name)
    mode_type = cfg.get("type")
    worlds: List[dict] = []

    def _fetch(func, *args, **kwargs):
        for attempt in range(1, 4):
            try:
                return func(*args, **kwargs)
            except Exception as exc:  # pragma: no cover - network errors
                logger.error(
                    "%s attempt %d failed: %s", func.__name__, attempt, exc
                )
        return []

    if mode_type == "keyword":
        for kw in cfg.get("keywords", []):
            worlds.extend(_fetch(search_worlds, str(kw), limit=50))
    elif mode_type == "user":
        user_id = str(cfg.get("user_id", ""))
        if user_id:
            worlds = _fetch(get_user_worlds, user_id, limit=50)
        else:
            logger.error("Mode %s missing user_id", name)
    else:
        logger.error("Mode %s has unknown type %s", name, mode_type)

    if name.lower() == "taiwan":
        blacklist = _load_taiwan_blacklist()
        if blacklist:
            worlds = [
                w
                for w in worlds
                if (w.get("id") or w.get("worldId")) not in blacklist
            ]

    if worlds:
        update_history(worlds)
        sheet_file = ANALYTICS_DIR / f"{name}WorldSheet.xlsx"
        _save_worlds(worlds, sheet_file)
        stats_path = cfg.get("stats")
        if stats_path:
            stats_path = Path(str(stats_path))
            if not stats_path.is_absolute():
                stats_path = BASE.parent / stats_path
        update_daily_stats(name, worlds, stats_path)
    else:
        logger.error("No worlds fetched for mode %s", name)

    logger.info("Finished mode %s with %d worlds", name, len(worlds))
    return len(worlds)


def main() -> None:
    if not CONFIG_FILE.exists():
        raise SystemExit(f"Config file not found: {CONFIG_FILE}")
    with open(CONFIG_FILE, "r", encoding="utf-8") as f:
        modes = json.load(f)
    total = 0
    for name, cfg in modes.items():
        if isinstance(cfg, dict):
            total += _run_mode(name, cfg)
    print(f"總共抓取世界數量：{total}")


if __name__ == "__main__":
    main()
