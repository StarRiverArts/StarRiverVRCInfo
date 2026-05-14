"""EditableTableFrame – a Treeview-based table with in-cell editing and row deletion.

Usage
-----
    frame = EditableTableFrame(parent, columns=["爬取日期", "世界名稱", ...],
                               on_save=my_save_callback)
    frame.pack(fill=tk.BOTH, expand=True)
    frame.load_rows(list_of_tuples)        # populate
    rows = frame.get_rows()                # read back
    # on_save(rows) is called when the user clicks 儲存
"""
from __future__ import annotations

import tkinter as tk
from tkinter import ttk, messagebox
from pathlib import Path
from typing import Callable, List, Optional, Sequence

try:
    from openpyxl import Workbook, load_workbook  # type: ignore
except Exception:  # pragma: no cover
    Workbook = None  # type: ignore
    load_workbook = None  # type: ignore


class EditableTableFrame(ttk.Frame):
    """A self-contained editable table widget.

    Supports:
    - Double-click on any cell to open an inline Entry for editing.
    - Selecting one or more rows + clicking 刪除 (or pressing Delete) to remove them.
    - A 儲存 button that calls *on_save(rows)* with the current table data.
    - Optional direct Excel load/save via load_xlsx / save_xlsx.
    """

    def __init__(
        self,
        parent,
        columns: Sequence[str],
        on_save: Optional[Callable[[List[tuple]], None]] = None,
        col_width: int = 90,
        **kwargs,
    ) -> None:
        super().__init__(parent, **kwargs)
        self._columns = list(columns)
        self._on_save = on_save
        self._col_width = col_width
        self._entry_popup: Optional[tk.Entry] = None
        self._popup_row: Optional[str] = None
        self._popup_col: int = 0
        self._popup_var: Optional[tk.StringVar] = None
        self._build()

    # ── Layout ────────────────────────────────────────────────────────

    def _build(self) -> None:
        # Toolbar
        toolbar = ttk.Frame(self)
        toolbar.pack(fill=tk.X, padx=4, pady=(4, 2))

        self._var_status = tk.StringVar(value="0 筆")
        ttk.Label(toolbar, textvariable=self._var_status, foreground="gray").pack(
            side="left", padx=4
        )
        ttk.Button(toolbar, text="🗑 刪除選取行", command=self._delete_selected).pack(
            side="right", padx=2
        )
        ttk.Button(toolbar, text="💾 儲存", command=self._save).pack(
            side="right", padx=2
        )
        ttk.Button(toolbar, text="↩ 取消編輯", command=self._cancel_edit).pack(
            side="right", padx=2
        )

        # Treeview + scrollbars
        tree_frame = ttk.Frame(self)
        tree_frame.pack(fill=tk.BOTH, expand=True)

        col_ids = [str(i) for i in range(len(self._columns))]
        self.tree = ttk.Treeview(tree_frame, columns=col_ids, show="headings",
                                  selectmode="extended")
        for i, col in enumerate(self._columns):
            self.tree.heading(str(i), text=col)
            self.tree.column(str(i), width=self._col_width, anchor="center",
                             stretch=False)

        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        tree_frame.rowconfigure(0, weight=1)
        tree_frame.columnconfigure(0, weight=1)

        # Bindings
        self.tree.bind("<Double-1>", self._on_double_click)
        self.tree.bind("<Delete>", lambda _e: self._delete_selected())

    # ── Public API ────────────────────────────────────────────────────

    def load_rows(self, rows: Sequence) -> None:
        """Replace table contents with *rows* (each row is a sequence of values)."""
        self._cancel_edit()
        for item in self.tree.get_children():
            self.tree.delete(item)
        for row in rows:
            self.tree.insert("", tk.END, values=list(row))
        self._update_status()

    def get_rows(self) -> List[tuple]:
        """Return current visible rows as a list of tuples."""
        return [self.tree.item(item, "values") for item in self.tree.get_children()]

    def load_xlsx(self, path: Path) -> None:
        """Load data rows from an Excel file (header in row 1)."""
        if load_workbook is None:
            messagebox.showerror("錯誤", "openpyxl 未安裝")
            return
        if not path.exists():
            self.load_rows([])
            return
        wb = load_workbook(path)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        self.load_rows(rows)

    def save_xlsx(self, path: Path) -> None:
        """Write current rows to *path*, preserving the column header."""
        if Workbook is None:
            messagebox.showerror("錯誤", "openpyxl 未安裝")
            return
        wb = Workbook()
        ws = wb.active
        ws.append(self._columns)
        for row in self.get_rows():
            ws.append(list(row))
        wb.save(path)

    # ── Internal helpers ──────────────────────────────────────────────

    def _update_status(self, extra: str = "") -> None:
        count = len(self.tree.get_children())
        self._var_status.set(f"{count} 筆{extra}")

    def _cancel_edit(self) -> None:
        if self._entry_popup:
            self._entry_popup.destroy()
            self._entry_popup = None

    def _on_double_click(self, event: tk.Event) -> None:
        region = self.tree.identify("region", event.x, event.y)
        if region != "cell":
            return

        row_id = self.tree.identify_row(event.y)
        col_id = self.tree.identify_column(event.x)
        if not row_id or not col_id:
            return

        # col_id is "#1", "#2", … → 0-based index
        col_idx = int(col_id.lstrip("#")) - 1
        bbox = self.tree.bbox(row_id, col_id)
        if not bbox:
            return

        values = list(self.tree.item(row_id, "values"))
        cur_val = values[col_idx] if col_idx < len(values) else ""

        self._cancel_edit()

        var = tk.StringVar(value=str(cur_val) if cur_val is not None else "")
        entry = tk.Entry(self.tree, textvariable=var, justify="center",
                         font=("TkDefaultFont", 9))
        entry.place(x=bbox[0], y=bbox[1], width=bbox[2], height=bbox[3])
        entry.focus_set()
        entry.select_range(0, tk.END)

        self._entry_popup = entry
        self._popup_row = row_id
        self._popup_col = col_idx
        self._popup_var = var

        entry.bind("<Return>", self._commit_edit)
        entry.bind("<Escape>", lambda _e: self._cancel_edit())
        entry.bind("<FocusOut>", self._commit_edit)

    def _commit_edit(self, _event=None) -> None:
        if self._entry_popup is None:
            return
        new_val = self._popup_var.get()
        vals = list(self.tree.item(self._popup_row, "values"))
        if self._popup_col < len(vals):
            vals[self._popup_col] = new_val
        else:
            vals.extend([""] * (self._popup_col - len(vals) + 1))
            vals[self._popup_col] = new_val
        self.tree.item(self._popup_row, values=vals)
        self._cancel_edit()

    def _delete_selected(self) -> None:
        selected = self.tree.selection()
        if not selected:
            return
        if not messagebox.askyesno(
            "確認刪除",
            f"確定要刪除選取的 {len(selected)} 筆資料嗎？\n（儲存後才會寫入 Excel）",
        ):
            return
        for item in selected:
            self.tree.delete(item)
        self._update_status(f"（已移除 {len(selected)} 筆）")

    def _save(self) -> None:
        self._commit_edit()
        rows = self.get_rows()
        if self._on_save:
            self._on_save(rows)
        self._update_status("（已儲存）")
