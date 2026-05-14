"""Tkinter GUI for fetching and filtering VRChat world data.

The interface provides several tabs:
- Entrance: input authentication (cookie or basic auth), a search keyword
  and creator user ID.
- Data: fetch worlds by keyword and display the raw JSON.
- Filter: filter the keyword results by tag and sort order.
- World List: show the filtered worlds in a simple list.
- User Worlds: fetch and display worlds created by a specific user.

The tool relies on functions in ``scraper/scraper.py``. Results are saved under
that folder for reuse by other scripts. Creator worlds are fetched through the
worlds API by creator user ID, which is a better fit for the future web app
than scraping website HTML.
"""
from __future__ import annotations

import json
import logging
import datetime as dt
from pathlib import Path
import traceback
import re
import tkinter as tk
from tkinter import ttk, messagebox
# Support running both as a module (``python -m world_info.ui``) and as a
# stand-alone script. When executed directly, ``__package__`` is ``None`` and
# relative imports would fail, causing the console window to close immediately
# on error. Adjust ``sys.path`` first so that subsequent imports work in both
# contexts and any errors remain visible.
if __package__ is None or __package__ == "":
    import sys

    sys.path.append(str(Path(__file__).resolve().parent.parent))

    from world_info.scraper.scraper import (
        load_history,
        update_history,
        record_row,
        _parse_date,
        vrchat_login,
        vrchat_verify_2fa,
        vrchat_check_session,
        HISTORY_TABLE,
        HEADERS_FILE,
    )
    from world_info.analytics import update_daily_stats
    from world_info.constants import (
        BASE,
        RAW_FILE,
        USER_FILE,
        STAR_RIVER_FILE,
        TAIWAN_FILE,
        METRIC_COLS,
        LEGEND_TEXT,
    )
    from world_info.tabs import (
        EntryTab,
        SearchResultsTab,
        TaiwanTab,
        UserTab,
        HistoryTab,
        SettingsTab,
        AboutTab,
        LogTab,
    )
    from world_info.actions import (
        load_auth_headers,
        search_keyword,
        search_user,
        search_fixed,
        save_worlds,
    )
else:
    from .scraper.scraper import (
        load_history,
        update_history,
        record_row,
        _parse_date,
        vrchat_login,
        vrchat_verify_2fa,
        vrchat_check_session,
        HISTORY_TABLE,
        HEADERS_FILE,
    )
    from .analytics import update_daily_stats
    from .constants import (
        BASE,
        RAW_FILE,
        USER_FILE,
        STAR_RIVER_FILE,
        TAIWAN_FILE,
        METRIC_COLS,
        LEGEND_TEXT,
    )
    from .tabs import (
        EntryTab,
        SearchResultsTab,
        TaiwanTab,
        UserTab,
        HistoryTab,
        SettingsTab,
        AboutTab,
        LogTab,
    )
    from .actions import (
        load_auth_headers,
        search_keyword,
        search_user,
        search_fixed,
        save_worlds,
    )

try:
    from openpyxl import load_workbook, Workbook  # type: ignore
except Exception:  # pragma: no cover - optional
    load_workbook = None  # type: ignore
    Workbook = None  # type: ignore

logger = logging.getLogger(__name__)


def _warn_null_visits(worlds: list, source: str) -> None:
    """Warn (log + dialog) when worlds are missing visits data from the API."""
    null_count = sum(1 for w in worlds if w.get("visits") is None)
    if null_count:
        logger.warning(
            "[%s] %d 個世界的瀏覽人次未從 API 取得（顯示為空白）。"
            " 請確認 Cookie 是否有效，或重新登入 VRChat。",
            source,
            null_count,
        )
        try:
            messagebox.showwarning(
                "⚠️ 拜訪人次未取得",
                f"有 {null_count} 個世界的拜訪人次無法從 API 取得。\n\n"
                "原因：VRChat API 需要登入才能回傳拜訪數據。\n"
                "請到「入口」頁面登入 VRChat，或貼上有效的 Cookie 後再爬取。",
            )
        except Exception:
            pass  # no Tk root in headless / test context


# configuration and extra spreadsheets
SETTINGS_FILE = BASE / "scraper" / "settings.json"

class WorldInfoUI(tk.Tk):
    def __init__(self) -> None:
        super().__init__()
        self.title("World Info")
        self.geometry("800x600")

        self.headers = {}
        self.data: list[dict] = []
        self.user_data: list[dict] = []
        self.filtered: list[dict] = []
        self.history: dict[str, list[dict]] = load_history()

        self.settings = self._load_settings()

        self.nb = ttk.Notebook(self)
        self.nb.pack(fill=tk.BOTH, expand=True)

        # Tab order: 入口 → 個人世界 → 台灣世界 → 搜尋結果 → 歷史 → 設定 → 關於 → 日誌
        self.tab_entry   = EntryTab(self.nb, self)
        self.tab_user    = UserTab(self.nb, self)
        self.tab_taiwan  = TaiwanTab(self.nb, self)
        self.tab_search  = SearchResultsTab(self.nb, self)
        self.tab_history = HistoryTab(self.nb, self)
        self.tab_settings = SettingsTab(self.nb, self)
        self.tab_about   = AboutTab(self.nb, self)
        self.tab_log     = LogTab(self.nb, self)

        # Legacy aliases so older methods that reference tab_list / tab_data still work
        self.tab_list   = self.tab_search
        self.tab_data   = self.tab_search
        self.tab_filter = self.tab_search

        handler = self.tab_log.handler
        logging.getLogger().addHandler(handler)
        logging.getLogger().setLevel(logging.INFO)

        self._load_local_tables()
        self._apply_settings()






    # ── VRChat 登入 ────────────────────────────────────────────────────

    def _set_login_status(self, text: str, color: str = "gray") -> None:
        self.var_login_status.set(text)
        self.tab_entry.app.lbl_login_status.configure(foreground=color)

    def _persist_cookie(self, cookie: str) -> None:
        """Save cookie to settings and headers.json so it survives restarts."""
        self.settings["cookie"] = cookie
        self.var_cookie.set(cookie)
        self.var_set_cookie.set(cookie)
        with open(SETTINGS_FILE, "w", encoding="utf-8") as f:
            json.dump(self.settings, f, ensure_ascii=False, indent=2)
        with open(HEADERS_FILE, "w", encoding="utf-8") as f:
            json.dump({"Cookie": cookie}, f, ensure_ascii=False, indent=2)
        self.headers = load_auth_headers(cookie, None, None)
        logger.info("Cookie 已儲存至 headers.json")

    def _on_login(self) -> None:
        username = self.var_user.get().strip()
        password = self.var_pass.get().strip()
        if not username or not password:
            messagebox.showerror("登入失敗", "請輸入帳號和密碼")
            return

        self._set_login_status("⏳ 登入中…", "orange")
        self.update_idletasks()

        result = vrchat_login(username, password)

        if result.get("requires_2fa"):
            methods = result.get("methods", ["totp"])
            self._pending_auth_cookie = result.get("auth_cookie", "")
            self.var_twofa_method.set(methods[0])
            self._twofa_frame.grid()          # 顯示 2FA 區塊
            self._set_login_status(
                f"⚠️ 需要雙因素驗證（{', '.join(methods)}）", "orange"
            )
            logger.info("VRChat 登入需要 2FA，方式: %s", methods)
            return

        if not result.get("ok"):
            err = result.get("error", "未知錯誤")
            self._set_login_status(f"❌ 登入失敗：{err}", "red")
            messagebox.showerror("登入失敗", err)
            return

        self._twofa_frame.grid_remove()
        cookie = result["cookie"]
        user = result.get("user", {})
        display_name = user.get("displayName", username)
        self._persist_cookie(cookie)
        self._set_login_status(f"✅ 已登入：{display_name}", "green")
        logger.info("VRChat 登入成功：%s", display_name)

    def _on_verify_2fa(self) -> None:
        code = self.var_twofa_code.get().strip().replace(" ", "")
        method = self.var_twofa_method.get().strip()
        auth_cookie = getattr(self, "_pending_auth_cookie", "")

        if not code:
            messagebox.showerror("驗證失敗", "請輸入驗證碼")
            return

        self._set_login_status("⏳ 驗證中…", "orange")
        self.update_idletasks()

        result = vrchat_verify_2fa(code, method, auth_cookie)
        if not result.get("ok"):
            err = result.get("error", "未知錯誤")
            self._set_login_status(f"❌ 2FA 驗證失敗：{err}", "red")
            messagebox.showerror("2FA 驗證失敗", err)
            return

        self._twofa_frame.grid_remove()
        cookie = result["cookie"]
        self._persist_cookie(cookie)
        self._set_login_status("✅ 2FA 驗證成功，已登入", "green")
        logger.info("VRChat 2FA 驗證成功")

    def _async_verify_session(self) -> None:
        """Called once after startup to verify the saved cookie is still valid."""
        cookie = self.settings.get("cookie", "")
        if not cookie:
            return
        result = vrchat_check_session(cookie)
        if result.get("ok"):
            name = result["user"].get("displayName", "")
            self._set_login_status(
                f"✅ Session 有效：{name}" if name else "✅ Session 有效", "green"
            )
            logger.info("啟動時 session 驗證成功：%s", name)
        else:
            err = result.get("error", "未知")
            self._set_login_status(f"⚠️ Session 已失效：{err}", "red")
            logger.warning("啟動時 session 驗證失敗：%s", err)

    def _handle_auth_error(self, err: Exception) -> bool:
        """Show a helpful prompt when a scraping call fails with a 401/auth error.

        Returns True so callers can use it as a guard:
            except RuntimeError as e:
                if self._handle_auth_error(e): return
        """
        msg = str(err)
        if "401" in msg or "Unauthorized" in msg.lower() or "auth" in msg.lower():
            self._set_login_status("⚠️ 認證已失效，請重新登入", "red")
            messagebox.showwarning(
                "認證失效",
                "VRChat 回傳 401 Unauthorized。\n"
                "請至「入口」頁重新登入，或貼上有效的 Cookie。",
            )
            return True
        return False

    def _on_logout(self) -> None:
        self.settings.pop("cookie", None)
        self.var_cookie.set("")
        self.var_set_cookie.set("")
        self.headers = {}
        if HEADERS_FILE.exists():
            HEADERS_FILE.write_text("{}", encoding="utf-8")
        with open(SETTINGS_FILE, "w", encoding="utf-8") as f:
            json.dump(self.settings, f, ensure_ascii=False, indent=2)
        self._set_login_status("⬜ 已登出", "gray")
        self._twofa_frame.grid_remove()
        logger.info("已清除登入資訊")

    def _apply_manual_cookie(self) -> None:
        cookie = self.var_cookie.get().strip()
        if not cookie:
            messagebox.showerror("錯誤", "Cookie 欄位為空")
            return
        self._persist_cookie(cookie)
        self._set_login_status("✅ Cookie 已套用（手動）", "green")
        logger.info("手動套用 Cookie")

    def _load_settings(self) -> dict:
        if SETTINGS_FILE.exists():
            try:
                with open(SETTINGS_FILE, "r", encoding="utf-8") as f:
                    return json.load(f)
            except Exception:
                return {}
        return {}

    def _save_settings(self) -> None:
        self.settings["cookie"] = self.var_set_cookie.get()
        self.settings["personal_keywords"] = self.var_personal_kw.get()
        self.settings["taiwan_keywords"] = self.var_taiwan_kw.get()
        self.settings["blacklist"] = self.var_blacklist.get()
        self.settings["player_id"] = self.var_playerid_set.get()
        with open(SETTINGS_FILE, "w", encoding="utf-8") as f:
            json.dump(self.settings, f, ensure_ascii=False, indent=2)
        self._apply_settings()

    def _apply_settings(self) -> None:
        cookie = self.settings.get("cookie", "")
        self.var_cookie.set(cookie)
        self.var_set_cookie.set(cookie)
        self.var_personal_kw.set(self.settings.get("personal_keywords", ""))
        self.var_taiwan_kw.set(self.settings.get("taiwan_keywords", ""))
        self.var_blacklist.set(self.settings.get("blacklist", ""))
        player_id = self.settings.get("player_id", "")
        self.var_userid.set(player_id)
        self.var_playerid_set.set(player_id)
        # Reflect saved login state in the status label; async-check in background
        if cookie:
            self._set_login_status("⏳ 驗證 session 中…", "orange")
            self.after(200, self._async_verify_session)
        else:
            self._set_login_status("⬜ 尚未登入", "gray")

    def _load_local_tables(self) -> None:
        """Load existing personal Excel file and populate the user world list."""
        if load_workbook is None:
            return

        # clear previous content so the method can be reused for reloading
        for item in self.user_tree.get_children():
            self.user_tree.delete(item)
        self.user_data.clear()

        file_path = STAR_RIVER_FILE
        if file_path.exists():
            wb = load_workbook(file_path)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                if len(row) >= 15:
                    row = row[:15]
                elif len(row) >= 14:
                    # backward compatibility with old files without fetch date
                    row = ("",) + row[:14]
                else:
                    # skip rows that don't have enough columns
                    continue
                self.user_tree.insert("", tk.END, values=row)
                (
                    fetched,
                    name,
                    wid,
                    pub,
                    upd,
                    visits,
                    size,
                    fav,
                    heat,
                    pop,
                    labs_to_pub,
                    vf,
                    since_upd,
                    since_pub,
                    vpp,
                ) = row
                self.user_data.append(
                    {
                        "爬取日期": fetched,
                        "世界名稱": name,
                        "世界ID": wid,
                        "發布日期": pub,
                        "最後更新": upd,
                        "瀏覽人次": visits,
                        "大小": size,
                        "收藏次數": fav,
                        "熱度": heat,
                        "人氣": pop,
                        "實驗室到發布": labs_to_pub,
                        "瀏覽蒐藏比": vf,
                        "距離上次更新": since_upd,
                        "已發布": since_pub,
                        "人次發布比": vpp,
                    }
                )
                ts = _parse_date(fetched)
                if ts:
                    rec = {
                        "timestamp": int(ts.timestamp()),
                        "name": name,
                        "visits": visits,
                        "favorites": fav,
                        "heat": heat,
                        "popularity": pop,
                        "updated_at": upd,
                        "publicationDate": pub,
                        "labsPublicationDate": "",
                    }
                    recs = self.history.setdefault(wid, [])
                    if not any(r.get("timestamp") == rec["timestamp"] for r in recs):
                        recs.append(rec)
                        recs.sort(key=lambda r: r.get("timestamp", 0))
            self._create_world_tabs()
            self._update_dashboard()
            self._refresh_history_table()

        # populate world list tab from fixed keyword spreadsheet
        for item in self.tree.get_children():
            self.tree.delete(item)
        if TAIWAN_FILE.exists():
            wb = load_workbook(TAIWAN_FILE)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                if len(row) >= 15:
                    self.tree.insert("", tk.END, values=row[:15])


    def _search_fixed(self, keywords: str, out_file: Path, source_name: str | None = None,
                      update_search_tab: bool = True) -> bool:
        """Search fixed keywords and save to *out_file*.

        Returns True when worlds were fetched, False on failure/empty.
        *update_search_tab* controls whether results are also pushed to the
        搜尋結果 tab (set False for Taiwan search which has its own tab).
        """
        blacklist = {k.strip() for k in self.settings.get("blacklist", "").split(",") if k.strip()}
        logger.info("Searching keywords %s", keywords)
        try:
            all_worlds = search_fixed(keywords, self.headers, blacklist)
        except RuntimeError as e:  # pragma: no cover - runtime only
            logger.error("Keyword search failed: %s", e)
            if not self._handle_auth_error(e):
                messagebox.showerror("Error", str(e))
            return False
        except Exception as e:  # pragma: no cover - runtime only
            logger.error("Keyword search failed: %s", e)
            messagebox.showerror("Error", str(e))
            return False
        if not all_worlds:
            logger.warning("No worlds found for %s", keywords)
            return False
        _warn_null_visits(all_worlds, keywords)
        update_history(all_worlds)
        self.history = load_history()
        self._refresh_history_table()
        save_worlds(all_worlds, out_file)
        if source_name:
            update_daily_stats(source_name, all_worlds)
        if update_search_tab:
            self.data = all_worlds
            self._update_tag_options()
            self._apply_filter()
        return True

    def _search_personal(self) -> None:
        """Fetch worlds for the configured player ID and overwrite the local table."""
        self._load_auth_headers()
        user_id = self.settings.get("player_id", "").strip()
        if not user_id:
            messagebox.showerror("Error", "Player ID required")
            logger.error("Player ID required for personal search")
            return
        try:
            worlds = search_user(user_id, self.headers)
        except RuntimeError as e:  # pragma: no cover - runtime only
            logger.error("Personal search failed: %s", e)
            if not self._handle_auth_error(e):
                messagebox.showerror("Error", str(e))
            return
        except Exception as e:  # pragma: no cover - runtime only
            logger.error("Personal search failed: %s", e)
            messagebox.showerror("Error", str(e))
            return

        logger.info("Fetched %d worlds for %s", len(worlds), user_id)
        _warn_null_visits(worlds, user_id)

        # save_worlds now overwrites the file so no manual dedup is needed
        save_worlds(worlds, STAR_RIVER_FILE)
        update_history(worlds)
        self.history = load_history()
        self._refresh_history_table()
        source_name = re.sub(r"[^A-Za-z0-9_-]+", "_", user_id)
        update_daily_stats(source_name, worlds)

        # Reload the table, then navigate to the editor sub-tab so the user
        # can immediately review and delete any bad records.
        self._load_local_tables()
        if hasattr(self, "personal_editor"):
            self.personal_editor.load_xlsx(STAR_RIVER_FILE)
            n = len(self.personal_editor.get_rows())
            if hasattr(self, "var_personal_editor_status"):
                self.var_personal_editor_status.set(f"已載入 {n} 筆（{STAR_RIVER_FILE.name}）")
        self.nb.select(self.tab_user.frame)
        if hasattr(self, "user_nb") and hasattr(self, "tab_editor"):
            self.user_nb.select(self.tab_editor)

    def _search_taiwan(self) -> None:
        self._load_auth_headers()
        ok = self._search_fixed(
            self.settings.get("taiwan_keywords", ""),
            TAIWAN_FILE,
            "taiwan",
            update_search_tab=False,
        )
        if ok:
            # Reload the editable Taiwan table and navigate to it
            self.tab_taiwan._reload()
            self.nb.select(self.tab_taiwan.frame)
    # ------------------------------------------------------------------
    # Actions
    def _load_auth_headers(self) -> None:
        cookie = self.var_cookie.get() or None
        user = self.var_user.get() or None
        pw = self.var_pass.get() or None
        self.headers = load_auth_headers(cookie, user, pw)

    def _on_search(self) -> None:
        self._load_auth_headers()
        keyword = self.var_keyword.get().strip()
        if not keyword:
            messagebox.showerror("Error", "Keyword required")
            return
        try:
            logger.info("Searching keyword %s", keyword)
            self.data = search_keyword(keyword, self.headers)
            _warn_null_visits(self.data, keyword)
            with open(RAW_FILE, "w", encoding="utf-8") as f:
                json.dump(self.data, f, ensure_ascii=False, indent=2)
            update_history(self.data)
            self.history = load_history()
            self._refresh_history_table()
            self.text_data.delete("1.0", tk.END)
            self.text_data.insert(tk.END, json.dumps(self.data, ensure_ascii=False, indent=2))
            self._update_tag_options()
            self.nb.select(self.tab_data.frame)
            logger.info("Keyword search complete: %s", keyword)
        except RuntimeError as e:  # pragma: no cover - runtime only
            logger.error("HTTP error during keyword search: %s", e)
            if not self._handle_auth_error(e):
                messagebox.showerror("HTTP Error", str(e))
        except Exception as e:  # pragma: no cover - runtime only
            logger.error("Keyword search failed: %s", e)
            messagebox.showerror("Error", str(e))

    def _on_user(self) -> None:
        self._load_auth_headers()
        user_id = self.var_userid.get().strip()
        if not user_id:
            messagebox.showerror("Error", "User ID required")
            logger.error("User ID required for search")
            return
        try:
            logger.info("Searching user %s", user_id)
            self.user_data = search_user(user_id, self.headers)
            _warn_null_visits(self.user_data, user_id)
            fetch_date = dt.datetime.now(dt.timezone.utc).strftime("%Y/%m/%d")
            for w in self.user_data:
                w["爬取日期"] = fetch_date
            with open(USER_FILE, "w", encoding="utf-8") as f:
                json.dump(self.user_data, f, ensure_ascii=False, indent=2)
            update_history(self.user_data)
            self.history = load_history()
            self._refresh_history_table()

            save_worlds(self.user_data, STAR_RIVER_FILE)
            self.settings["player_id"] = user_id
            self.var_playerid_set.set(user_id)
            self._save_settings()

            # Reload from Excel so user_data uses Chinese keys (for dashboard),
            # then navigate directly to the editor sub-tab for review/delete.
            self._load_local_tables()
            if hasattr(self, "personal_editor"):
                self.personal_editor.load_xlsx(STAR_RIVER_FILE)
            self.nb.select(self.tab_user.frame)
            if hasattr(self, "user_nb") and hasattr(self, "tab_editor"):
                self.user_nb.select(self.tab_editor)
            logger.info("User search complete: %s", user_id)
        except RuntimeError as e:  # pragma: no cover - runtime only
            logger.error("HTTP error during user search: %s", e)
            if not self._handle_auth_error(e):
                messagebox.showerror("HTTP Error", str(e))
        except Exception as e:  # pragma: no cover - runtime only
            logger.error("User search failed: %s", e)
            messagebox.showerror("Error", str(e))

    def _update_tag_options(self) -> None:
        tags = set()
        for w in self.data:
            for t in w.get("tags", []):
                tags.add(t)
        self.box_tag["values"] = ["all"] + sorted(tags)
        self.var_tag.set("all")

    def _apply_filter(self) -> None:
        worlds = list(self.data)
        tag = self.var_tag.get()
        if tag != "all":
            worlds = [w for w in worlds if tag in w.get("tags", [])]
        if self.var_sort.get() == "latest":
            worlds.sort(key=lambda w: w.get("publicationDate", ""), reverse=True)
        else:
            worlds.sort(key=lambda w: w.get("visits", 0), reverse=True)
        self.filtered = worlds
        for item in self.tree.get_children():
            self.tree.delete(item)
        for w in self.filtered:
            row = record_row(w)
            self.tree.insert("", tk.END, values=row)
        self.nb.select(self.tab_list.frame)
        
    def _refresh_history_table(self) -> None:
        if not hasattr(self, "hist_tree"):
            return
        for item in self.hist_tree.get_children():
            self.hist_tree.delete(item)
        for wid, recs in self.history.items():
            if not recs:
                continue
            latest = recs[-1]
            row = record_row({**latest, "id": wid}, latest.get("timestamp"))
            self.hist_tree.insert("", tk.END, values=row)

    def _on_select_user_world(self, event=None) -> None:
        item = self.user_tree.focus()
        if not item:
            return
        values = self.user_tree.item(item, "values")
        if len(values) < 3:
            return
        world_id = values[2]
        self.current_world_id = world_id
        self._draw_user_chart(world_id)

    def _draw_user_chart(self, world_id: str | None) -> None:
        self.user_canvas.delete("all")
        if not world_id:
            return
        data = self.history.get(world_id, [])
        if not data:
            return
        width = int(self.user_canvas.winfo_width() or 600)
        height = int(self.user_canvas.winfo_height() or 200)
        pad = 40

        times = [d["timestamp"] for d in data]
        first = data[0]
        created = _parse_date(first.get("created_at"))
        labs = _parse_date(first.get("labsPublicationDate"))
        pub = _parse_date(first.get("publicationDate"))
        update_times = sorted(
            {int(_parse_date(d.get("updated_at")).timestamp()) for d in data if _parse_date(d.get("updated_at"))}
        )

        t_points = times + update_times
        for t in (created, labs, pub):
            if t:
                t_points.append(int(t.timestamp()))
        min_t = min(t_points)
        max_t = max(t_points)
        if max_t == min_t:
            max_t += 1

        scale_x = width - 2 * pad
        scale_y = height - 2 * pad

        def x_at(ts: int) -> float:
            return pad + (ts - min_t) / (max_t - min_t) * scale_x

        def xy(idx: int, val: float, max_val: float):
            x = x_at(times[idx])
            y = height - pad - min(val, max_val) / max_val * scale_y
            return x, y

        colors = {
            "visits": "blue",
            "favorites": "green",
            "heat": "red",
            "popularity": "purple",
        }
        max_vis = max((d.get("visits", 0) or 0) for d in data)
        max_fav = max((d.get("favorites", 0) or 0) for d in data)
        vf_limit = max(max_vis, max_fav, 1)
        limits: dict[str, float] = {
            "visits": vf_limit,
            "favorites": vf_limit,
        }
        for key in ("heat", "popularity"):
            max_val = max((d.get(key, 0) or 0) for d in data)
            limits[key] = max_val * 1.1 if max_val > 0 else 1
        for key, color in colors.items():
            pts = [xy(i, d.get(key, 0), limits[key]) for i, d in enumerate(data)]
            for a, b in zip(pts, pts[1:]):
                self.user_canvas.create_line(a[0], a[1], b[0], b[1], fill=color)

        # event lines
        if labs:
            x = x_at(int(labs.timestamp()))
            self.user_canvas.create_line(x, pad, x, height - pad, fill="orange", dash=(4, 2))
            self.user_canvas.create_text(x + 2, pad, text=f"實驗 {labs:%m/%d}", anchor="nw", font=("TkDefaultFont", 8), fill="orange")
        if pub:
            x = x_at(int(pub.timestamp()))
            self.user_canvas.create_line(x, pad, x, height - pad, fill="black", dash=(4, 2))
            self.user_canvas.create_text(x + 2, pad, text=f"發布 {pub:%m/%d}", anchor="nw", font=("TkDefaultFont", 8), fill="black")
        for t in update_times:
            x = x_at(t)
            self.user_canvas.create_line(x, pad, x, height - pad, fill="gray", dash=(2, 2))
            date = dt.datetime.fromtimestamp(t, dt.timezone.utc)
            self.user_canvas.create_text(x + 2, pad, text=f"更新 {date:%m/%d}", anchor="nw", font=("TkDefaultFont", 8), fill="gray")

        # axes with ticks
        self.user_canvas.create_line(pad, height - pad, width - pad, height - pad)
        self.user_canvas.create_line(pad, pad, pad, height - pad)
        for i in range(5):  # x-axis ticks
            ts = min_t + (max_t - min_t) * i / 4
            x = x_at(ts)
            self.user_canvas.create_line(x, height - pad, x, height - pad + 5)
            label = dt.datetime.fromtimestamp(int(ts), dt.timezone.utc).strftime("%m/%d")
            self.user_canvas.create_text(x, height - pad + 15, text=label, anchor="n", font=("TkDefaultFont", 8))
        for i in range(5):  # y-axis ticks
            val = vf_limit * i / 4
            y = height - pad - val / vf_limit * scale_y
            self.user_canvas.create_line(pad - 5, y, pad, y)
            self.user_canvas.create_text(pad - 8, y, text=str(int(val)), anchor="e", font=("TkDefaultFont", 8))

        # title with world name
        name = data[0].get("name", world_id)
        self.user_canvas.create_text(width / 2, pad / 2, text=name, font=("TkDefaultFont", 12, "bold"))

    def _load_history_rows(self, world_id: str) -> list[dict]:
        """Return history rows for a world ID."""
        return list(self.history.get(world_id, []))

    def _draw_world_chart(self, canvas: tk.Canvas, world: dict) -> None:
        world_id = world.get("id") or world.get("worldId") or world.get("世界ID")
        data = self.history.get(world_id, [])
        canvas.delete("all")
        if not data:
            return

        width = int(canvas.winfo_width() or 600)
        height = int(canvas.winfo_height() or 200)
        pad = 40

        times = [d["timestamp"] for d in data]
        created = _parse_date(world.get("created_at") or world.get("上傳日期"))
        labs = _parse_date(world.get("labsPublicationDate"))
        pub = _parse_date(world.get("publicationDate"))
        update_times = sorted(
            {int(_parse_date(d.get("updated_at")).timestamp()) for d in data if _parse_date(d.get("updated_at"))}
        )

        t_points = times + update_times
        for t in (created, labs, pub):
            if t:
                t_points.append(int(t.timestamp()))
        min_t = min(t_points)
        max_t = max(t_points)
        if max_t == min_t:
            max_t += 1

        scale_x = width - 2 * pad
        scale_y = height - 2 * pad

        def x_at(ts: int) -> float:
            return pad + (ts - min_t) / (max_t - min_t) * scale_x

        def y_val(val: float, limit: float) -> float:
            return height - pad - min(val, limit) / limit * scale_y

        colors = {
            "visits": "blue",
            "favorites": "green",
            "heat": "red",
            "popularity": "purple",
        }
        max_vis = max((rec.get("visits", 0) or 0) for rec in data)
        max_fav = max((rec.get("favorites", 0) or 0) for rec in data)
        vf_limit = max(max_vis, max_fav, 1)
        limits: dict[str, float] = {
            "visits": vf_limit,
            "favorites": vf_limit,
        }
        for key in ("heat", "popularity"):
            max_val = max((rec.get(key, 0) or 0) for rec in data)
            limits[key] = max_val * 1.1 if max_val > 0 else 1

        for key, color in colors.items():
            pts = []
            for rec in data:
                ts = rec["timestamp"]
                val = rec.get(key, 0) or 0
                pts.append((x_at(ts), y_val(val, limits[key])))
            for a, b in zip(pts, pts[1:]):
                canvas.create_line(a[0], a[1], b[0], b[1], fill=color)

        # event lines
        if labs:
            x = x_at(int(labs.timestamp()))
            canvas.create_line(x, pad, x, height - pad, fill="orange", dash=(4, 2))
            canvas.create_text(x + 2, pad, text=f"實驗 {labs:%m/%d}", anchor="nw", font=("TkDefaultFont", 8), fill="orange")
        if pub:
            x = x_at(int(pub.timestamp()))
            canvas.create_line(x, pad, x, height - pad, fill="black", dash=(4, 2))
            canvas.create_text(x + 2, pad, text=f"發布 {pub:%m/%d}", anchor="nw", font=("TkDefaultFont", 8), fill="black")
        for t in update_times:
            x = x_at(t)
            canvas.create_line(x, pad, x, height - pad, fill="gray", dash=(2, 2))
            date = dt.datetime.fromtimestamp(t, dt.timezone.utc)
            canvas.create_text(x + 2, pad, text=f"更新 {date:%m/%d}", anchor="nw", font=("TkDefaultFont", 8), fill="gray")

        # axes with ticks and title
        canvas.create_line(pad, height - pad, width - pad, height - pad)
        canvas.create_line(pad, pad, pad, height - pad)
        canvas.create_line(width - pad, pad, width - pad, height - pad)
        for i in range(5):  # x-axis ticks
            ts = min_t + (max_t - min_t) * i / 4
            x = x_at(ts)
            canvas.create_line(x, height - pad, x, height - pad + 5)
            label = dt.datetime.fromtimestamp(int(ts), dt.timezone.utc).strftime("%m/%d")
            canvas.create_text(x, height - pad + 15, text=label, anchor="n", font=("TkDefaultFont", 8))
        for i in range(5):  # y-axis ticks based on visits/favorites
            val = vf_limit * i / 4
            y = height - pad - val / vf_limit * scale_y
            canvas.create_line(pad - 5, y, pad, y)
            canvas.create_text(pad - 8, y, text=str(int(val)), anchor="e", font=("TkDefaultFont", 8))

        name = world.get("name") or world.get("世界名稱") or world_id
        canvas.create_text(width / 2, pad / 2, text=name, font=("TkDefaultFont", 12, "bold"))

    def _create_world_tabs(self) -> None:
        """Create sub-tabs for each fetched user world with history."""
        if not hasattr(self, "detail_nb"):
            return
        # remove old tabs except the first (list tab)
        for tab_id in self.detail_nb.tabs()[1:]:
            self.detail_nb.forget(tab_id)


        unique: dict[str, dict] = {}
        for w in self.user_data:
            wid = w.get("世界ID") or w.get("worldId") or w.get("id")
            if wid:
                unique[wid] = w  # keep last occurrence

        for w in unique.values():
            frame = ttk.Frame(self.detail_nb)


            # dashboard table with a single metrics row
            dash = ttk.LabelFrame(frame, text="儀表板")
            dash.pack(fill=tk.X, padx=4, pady=2)
            dash_tree = ttk.Treeview(dash, columns=list(range(len(METRIC_COLS))), show="headings", height=2)
            for idx, col in enumerate(METRIC_COLS):
                dash_tree.heading(str(idx), text=col)
                dash_tree.column(str(idx), width=80, anchor="center")
            row = [w.get(col, "") for col in METRIC_COLS]
            dash_tree.insert("", tk.END, values=row)
            dash_tree.pack(fill=tk.X, expand=True)

            # section 1: latest fetched info
            sec1 = ttk.LabelFrame(frame, text="本次資料")
            sec1.pack(fill=tk.X, padx=4, pady=2)
            info_tree = ttk.Treeview(sec1, columns=("k", "v"), show="headings", height=8)
            info_tree.heading("k", text="欄位")
            info_tree.heading("v", text="值")
            for key, val in w.items():
                info_tree.insert("", tk.END, values=(key, val))
            info_tree.pack(fill=tk.BOTH, expand=True)

            # section 2: history table from history JSON
            sec2 = ttk.LabelFrame(frame, text="歷史紀錄")
            sec2.pack(fill=tk.BOTH, expand=True, padx=4, pady=2)
            hist_tree = ttk.Treeview(sec2, show="headings")
            cols = [
                "timestamp",
                "visits",
                "favorites",
                "heat",
                "popularity",
                "updated_at",
                "created_at",
                "labs",
                "pub",
                "days_to_pub",
                "days_since_upd",
                "visits_per_day",
                "fav_per_day",
            ]
            hist_tree["columns"] = cols
            headers = [
                "時間",
                "人次",
                "收藏",
                "熱度",
                "熱門度",
                "更新",
                "上傳",
                "實驗室",
                "公開",
                "上傳到公開",
                "距離更新",
                "人次/天",
                "收藏/天",
            ]
            for c, h in zip(cols, headers):
                hist_tree.heading(c, text=h)
                hist_tree.column(c, width=80, anchor="center")
            rows = self._load_history_rows(
                w.get("id") or w.get("worldId") or w.get("世界ID")
            )
            for r in rows:
                ts = r["timestamp"]
                ts_dt = dt.datetime.fromtimestamp(ts, dt.timezone.utc)
                upd = _parse_date(r.get("updated_at"))
                created = _parse_date(r.get("created_at"))
                labs = _parse_date(r.get("labsPublicationDate"))
                pub = _parse_date(r.get("publicationDate"))
                ts_str = ts_dt.strftime("%Y/%m/%d")
                upd_str = upd.strftime("%Y/%m/%d") if upd else ""
                created_str = created.strftime("%Y/%m/%d") if created else ""
                labs_str = labs.strftime("%Y/%m/%d") if labs else ""
                pub_str = pub.strftime("%Y/%m/%d") if pub else ""
                days_to_pub = ""
                if pub and created:
                    days_to_pub = (pub - created).days
                elif pub and labs:
                    days_to_pub = (pub - labs).days
                days_since = (ts_dt - upd).days if upd else ""
                since_pub = (ts_dt - pub).days if pub else 0
                vpd = round((r.get("visits", 0) or 0) / since_pub, 2) if since_pub > 0 else ""
                fpd = round((r.get("favorites", 0) or 0) / since_pub, 2) if since_pub > 0 else ""
                hist_tree.insert(
                    "",
                    tk.END,
                    values=(
                        ts_str,
                        r.get("visits"),
                        r.get("favorites"),
                        r.get("heat"),
                        r.get("popularity"),
                        upd_str,
                        created_str,
                        labs_str,
                        pub_str,
                        days_to_pub,
                        days_since,
                        vpd,
                        fpd,
                    ),
                )
            hist_tree.pack(fill=tk.BOTH, expand=True)

            # section 3: chart
            sec3 = ttk.LabelFrame(frame, text="折線圖")
            sec3.pack(fill=tk.BOTH, expand=True, padx=4, pady=2)
            canvas = tk.Canvas(sec3, bg="white")
            canvas.pack(fill=tk.BOTH, expand=True)
            canvas.bind("<Configure>", lambda e, c=canvas, ww=w: self._draw_world_chart(c, ww))
            ttk.Label(sec3, text=LEGEND_TEXT).pack()

            name = w.get("name") or w.get("世界名稱") or w.get("id")
            self.detail_nb.add(frame, text=str(name)[:15])

    def _update_dashboard(self) -> None:
        """Refresh the dashboard table and charts."""
        if not hasattr(self, "dash_tree"):
            return
        for item in self.dash_tree.get_children():
            self.dash_tree.delete(item)

        unique: dict[str, dict] = {}
        for w in self.user_data:
            wid = w.get("世界ID") or w.get("worldId") or w.get("id")
            if wid:
                unique[wid] = w
        for w in unique.values():
            row = [w.get("爬取日期", "")] + [w.get(col, "") for col in METRIC_COLS]
            self.dash_tree.insert("", tk.END, values=row)

        for frame, _, _ in getattr(self, "chart_frames", []):
            frame.destroy()
        self.chart_frames = []
        for w in unique.values():
            frm = ttk.Frame(self.chart_container)
            canvas = tk.Canvas(frm, bg="white")
            canvas.pack(fill=tk.BOTH, expand=True)
            canvas.bind("<Configure>", lambda e, c=canvas, ww=w: self._draw_world_chart(c, ww))
            ttk.Label(frm, text=LEGEND_TEXT).pack()
            self.chart_frames.append((frm, canvas, w))
        self._arrange_dashboard_charts()

    def _arrange_dashboard_charts(self, event=None) -> None:
        if not hasattr(self, "chart_frames"):
            return
        width = self.chart_canvas.winfo_width() if event is None else event.width
        cols = max(1, width // 260)
        for idx, (frm, _c, _w) in enumerate(self.chart_frames):
            frm.grid(row=idx // cols, column=idx % cols, padx=4, pady=4, sticky="nsew")
        for c in range(cols):
            self.chart_container.columnconfigure(c, weight=1)

    def _sort_tree(self, tree: ttk.Treeview, col: str, reverse: bool = False) -> None:
        """Sort a ``ttk.Treeview`` by the given column."""
        data = [(tree.set(k, col), k) for k in tree.get_children("")]

        def convert(val: str):
            try:
                val = val.replace("%", "").replace("天", "")
                return float(val)
            except Exception:
                return val

        data.sort(key=lambda t: convert(t[0]), reverse=reverse)
        for idx, (_val, k) in enumerate(data):
            tree.move(k, "", idx)
        tree.heading(col, command=lambda: self._sort_tree(tree, col, not reverse))


def main() -> None:  # pragma: no cover - simple runtime entry
    try:
        app = WorldInfoUI()
        app.mainloop()
    except Exception:  # pragma: no cover - runtime only
        traceback.print_exc()
        input("Press Enter to exit...")


if __name__ == "__main__":
    main()
