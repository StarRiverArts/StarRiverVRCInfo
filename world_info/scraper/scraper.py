
"""Retrieve world data from the VRChat API.

This script queries the VRChat worlds endpoint and stores the results as a JSON
list. Keyword search and creator-world search both use the same worlds API,
which is a better fit for a future web app than scraping website HTML.

Authentication headers can be supplied in ``headers.json`` or via command line
options. The credentials file should remain local-only.
"""

from __future__ import annotations

import base64
import json
import datetime as dt
import os
from pathlib import Path
from typing import Dict, List, Optional
import time
from urllib.parse import urlencode

try:
    from world_info.constants import METRIC_COLS
except ModuleNotFoundError:  # pragma: no cover - package path
    from constants import METRIC_COLS

try:
    from openpyxl import Workbook, load_workbook  # type: ignore
except Exception:  # pragma: no cover - optional dependency
    Workbook = None  # type: ignore
    load_workbook = None  # type: ignore


try:
    import requests  # type: ignore
except Exception:  # pragma: no cover - optional dependency
    requests = None  # type: ignore

BASE = Path(__file__).parent
HEADERS_FILE = BASE / "headers.json"
HISTORY_FILE = BASE / "history.json"
HISTORY_TABLE = BASE / "history_table.xlsx"


class VRChatRateLimitError(RuntimeError):
    def __init__(
        self,
        message: str,
        *,
        retry_after_seconds: int = 0,
        url: str | None = None,
        status_code: int = 429,
    ) -> None:
        super().__init__(message)
        self.retry_after_seconds = max(int(retry_after_seconds or 0), 0)
        self.url = url
        self.status_code = status_code


def _use_system_proxy() -> bool:
    return os.getenv("WORLD_INFO_USE_SYSTEM_PROXY", "").strip() == "1"


def _build_http_client():
    if requests is None:
        raise RuntimeError("requests package is required")
    if _use_system_proxy():
        return requests
    session = requests.Session()
    session.trust_env = False
    return session


def _load_headers(cookie: Optional[str] = None,
                  username: Optional[str] = None,
                  password: Optional[str] = None) -> Dict[str, str]:
    """Load HTTP headers from ``headers.json`` and command line options."""

    headers = {
        "User-Agent": os.getenv(
            "WORLD_INFO_USER_AGENT",
            "VRRacingClubTW-WorldInfo/2.0 local-user",
        )
    }

    if HEADERS_FILE.exists():
        with open(HEADERS_FILE, "r", encoding="utf-8") as f:
            try:
                headers.update(json.load(f))
            except json.JSONDecodeError:
                pass

    if cookie:
        headers["Cookie"] = cookie

    if username and password:
        token = base64.b64encode(f"{username}:{password}".encode()).decode()
        headers["Authorization"] = f"Basic {token}"

    return headers


HEADERS: Dict[str, str] = _load_headers()


def _has_auth_headers(headers: Optional[Dict[str, str]]) -> bool:
    payload = headers or HEADERS
    return bool(payload.get("Cookie") or payload.get("Authorization"))


def record_row(world: dict, now: Optional[int] = None) -> List[object]:
    """Return a metrics row for history tables and Excel.

    The first column contains the fetch date so spreadsheets clearly show when
    the data was retrieved.
    visits/favorites are kept as None when the API did not return them so that
    callers can distinguish "not available" from a genuine zero.
    """

    ts_now = (
        dt.datetime.fromtimestamp(now, dt.timezone.utc)
        if isinstance(now, int)
        else dt.datetime.now(dt.timezone.utc)
    )
    fetch_date = ts_now.strftime("%Y/%m/%d")

    pub = _parse_date(world.get("publicationDate"))
    updated = _parse_date(world.get("updated_at"))
    labs = _parse_date(world.get("labsPublicationDate"))

    days_labs_to_pub = (pub - labs).days if pub and labs else ""
    visits = world.get("visits")          # None means API didn't return this field
    favs = world.get("favorites")         # None means API didn't return this field
    _v = visits if visits is not None else 0
    _f = favs if favs is not None else 0
    ratio_vf = f"{round(_f / _v * 100, 2)}%" if _v else ""
    since_update = (ts_now - updated).days if updated else ""
    since_pub = (ts_now - pub).days if pub else 0
    visits_per_day = round(_v / since_pub, 2) if (visits is not None and since_pub > 0) else ""
    labs_to_pub_str = f"{days_labs_to_pub}天" if days_labs_to_pub != "" else ""
    since_update_str = f"{since_update}天" if since_update != "" else ""
    since_pub_str = f"{since_pub}天" if since_pub else ""

    return [
        fetch_date,
        world.get("name"),
        world.get("id"),
        world.get("publicationDate"),
        world.get("updated_at"),
        visits,       # None → empty cell; genuine 0 stays 0
        world.get("capacity"),
        favs,         # None → empty cell; genuine 0 stays 0
        world.get("heat"),
        world.get("popularity"),
        labs_to_pub_str,
        ratio_vf,
        since_update_str,
        since_pub_str,
        visits_per_day,
    ]


def _parse_date(value: Optional[str]) -> Optional[dt.datetime]:
    if not value:
        return None
    try:
        # allow plain dates like "2025/7/12" for manual edits
        if isinstance(value, (int, float)):
            return dt.datetime.fromtimestamp(float(value), dt.timezone.utc)
        if isinstance(value, str) and value.isdigit():
            return dt.datetime.fromtimestamp(float(value), dt.timezone.utc)
        if value.endswith("Z"):
            value = value[:-1] + "+00:00"
        if "T" in value:
            dt_obj = dt.datetime.fromisoformat(value)
        else:
            for fmt in ("%Y/%m/%d", "%Y-%m-%d"):
                try:
                    dt_obj = dt.datetime.strptime(value, fmt)
                    break
                except ValueError:
                    dt_obj = None
            if dt_obj is None:
                return None
        if dt_obj.tzinfo is None:
            dt_obj = dt_obj.replace(tzinfo=dt.timezone.utc)
        return dt_obj
    except Exception:
        return None


def load_history() -> Dict[str, List[dict]]:
    """Load the long-term history file if present."""
    if HISTORY_FILE.exists():
        with open(HISTORY_FILE, "r", encoding="utf-8") as f:
            try:
                return json.load(f)
            except json.JSONDecodeError:
                return {}
    return {}


def update_history(worlds: List[dict], threshold: int = 3600) -> Dict[str, List[dict]]:
    """Append new stats to ``history.json`` unless recorded recently."""
    history = load_history()
    now = int(time.time())
    appended = False
    for w in worlds:
        wid = w.get("id") or w.get("worldId")
        if not wid:
            continue
        recs = history.setdefault(wid, [])
        if recs and now - recs[-1].get("timestamp", 0) < threshold:
            continue
        rec = {
            "timestamp": now,
            "name": w.get("name"),
            "created_at": w.get("created_at"),
            "visits": w.get("visits"),
            "favorites": w.get("favorites"),
            "heat": w.get("heat"),
            "popularity": w.get("popularity"),
            "updated_at": w.get("updated_at"),
            "publicationDate": w.get("publicationDate"),
            "labsPublicationDate": w.get("labsPublicationDate"),
        }
        recs.append(rec)
        row = record_row(w, now)
        _append_history_table(row)
        appended = True
    if appended:
        with open(HISTORY_FILE, "w", encoding="utf-8") as f:
            json.dump(history, f, ensure_ascii=False, indent=2)
    return history

def _append_history_table(row: List[object]) -> None:
    """Append a metrics row to ``history_table.xlsx``."""
    correct_headers = ["爬取日期"] + METRIC_COLS
    if Workbook is None or load_workbook is None:
        raise RuntimeError("openpyxl is required to write Excel logs")
    if HISTORY_TABLE.exists():
        wb = load_workbook(HISTORY_TABLE)
        ws = wb.active
        # Fix header if it was created before 爬取日期 was the first column
        current_first = ws.cell(1, 1).value
        if current_first != "爬取日期":
            for col_idx, h in enumerate(correct_headers, 1):
                ws.cell(1, col_idx).value = h
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(correct_headers)
    ws.append(row)
    wb.save(HISTORY_TABLE)



def _fetch_paginated(base_url: str, limit: int, delay: float,
                     headers: Optional[Dict[str, str]] = None) -> List[dict]:
    """Fetch up to ``limit`` worlds from ``base_url`` using pagination."""
    client = _build_http_client()
    results: List[dict] = []
    offset = 0
    retry_count = 0
    max_retry_count = 1
    while len(results) < limit:
        remaining = min(100, limit - len(results))
        sep = '&' if '?' in base_url else '?'  # handle URLs with no query yet
        url = f"{base_url}{sep}n={remaining}&offset={offset}"
        try:
            r = client.get(url, headers=headers or HEADERS, timeout=30)
            r.raise_for_status()
        except requests.exceptions.HTTPError as e:  # pragma: no cover - runtime only
            if e.response is not None and e.response.status_code == 429:
                retry_count += 1
                retry_after = 0
                header_value = getattr(e.response, "headers", {}).get("Retry-After")
                if header_value:
                    try:
                        retry_after = max(1, int(header_value))
                    except ValueError:
                        retry_after = 0
                if retry_after <= 0:
                    retry_after = 15
                if retry_count > max_retry_count:
                    raise VRChatRateLimitError(
                        "429 Too Many Requests: VRChat rate limited this request. "
                        f"Suggested wait: {retry_after} seconds before retrying.",
                        retry_after_seconds=retry_after,
                        url=url,
                    ) from e
                time.sleep(retry_after)
                continue
            if e.response is not None and e.response.status_code == 401:
                if _has_auth_headers(headers):
                    raise RuntimeError(
                        "401 Unauthorized: VRChat rejected the current auth headers. "
                        "Refresh world_info/scraper/headers.json with a valid Cookie."
                    ) from e
                raise RuntimeError(
                    "401 Unauthorized: this request needs VRChat auth. "
                    "Create world_info/scraper/headers.json with your local Cookie, for example "
                    '{"Cookie": "auth=...; twoFactorAuth=...; machineId=..."}'
                ) from e
            if e.response is not None and e.response.status_code == 403:
                raise RuntimeError(
                    "403 Forbidden: check your cookie or login credentials"
                ) from e
            raise
        retry_count = 0
        chunk = r.json()
        if not isinstance(chunk, list):
            break
        results.extend(chunk)
        if len(chunk) < remaining:
            break
        offset += len(chunk)
        if delay:
            time.sleep(delay)
    return results[:limit]


WORLD_SEARCH_SORTS = {
    "popularity",
    "heat",
    "trust",
    "shuffle",
    "random",
    "favorites",
    "reportScore",
    "reportCount",
    "publicationDate",
    "labsPublicationDate",
    "created",
    "_created_at",
    "updated",
    "_updated_at",
    "order",
    "relevance",
    "magic",
    "name",
}


def _csv_value(values: object) -> str:
    if values in (None, ""):
        return ""
    if isinstance(values, str):
        items = values.split(",")
    elif isinstance(values, (list, tuple, set)):
        items = [str(item) for item in values]
    else:
        items = [str(values)]
    return ",".join(item.strip() for item in items if item and item.strip())


def _optional_query_value(value: object) -> str:
    if value in (None, ""):
        return ""
    text = str(value).strip()
    if text.casefold() in {"none", "null", "undefined"}:
        return ""
    return text


def _search_worlds(params: Dict[str, object], limit: int = 20,
                   delay: float = 1.0,
                   headers: Optional[Dict[str, str]] = None,
                   endpoint: str = "worlds") -> List[dict]:
    query = urlencode({k: v for k, v in params.items() if v not in (None, "")})
    if endpoint not in {"worlds", "worlds/active"}:
        raise ValueError(f"unsupported worlds endpoint: {endpoint}")
    base = f"https://api.vrchat.cloud/api/1/{endpoint}"
    if query:
        base = f"{base}?{query}"
    return _fetch_paginated(base, limit, delay, headers)


def search_worlds(keyword: str, limit: int = 20, delay: float = 1.0,
                  headers: Optional[Dict[str, str]] = None) -> List[dict]:
    if requests is None:
        raise RuntimeError("requests package is required")

    return _search_worlds(
        {"search": keyword, "sort": "relevance", "order": "descending"},
        limit,
        delay,
        headers,
    )


def get_user_worlds(user_id: str, limit: int = 20, delay: float = 1.0,
                    headers: Optional[Dict[str, str]] = None) -> List[dict]:
    """Fetch public worlds created by the given creator user ID."""
    return _search_worlds(
        {"userId": user_id, "sort": "updated", "order": "descending"},
        limit,
        delay,
        headers,
    )


def search_worlds_query(*,
                        search: Optional[str] = None,
                        tags: object = None,
                        notags: object = None,
                        sort: str = "popularity",
                        order: str = "descending",
                        featured: Optional[bool] = None,
                        active: bool = False,
                        release_status: Optional[str] = None,
                        platform: Optional[str] = None,
                        limit: int = 20,
                        delay: float = 1.0,
                        headers: Optional[Dict[str, str]] = None) -> List[dict]:
    """Search public worlds using the VRChat worlds query parameters.

    ``tags`` and ``notags`` are comma-separated API parameters. VRChat treats
    the included tags as an OR match, so any one tag can return a world.
    """

    if requests is None:
        raise RuntimeError("requests package is required")

    sort = (sort or "popularity").strip()
    if sort not in WORLD_SEARCH_SORTS:
        raise ValueError(f"unsupported world search sort: {sort}")
    order = (order or "descending").strip()
    if order not in {"ascending", "descending"}:
        raise ValueError(f"unsupported world search order: {order}")

    params: Dict[str, object] = {
        "search": _optional_query_value(search),
        "tag": _csv_value(tags),
        "notag": _csv_value(notags),
        "sort": sort,
        "order": order,
        "releaseStatus": _optional_query_value(release_status),
        "platform": _optional_query_value(platform),
    }
    if featured is not None:
        params["featured"] = "true" if featured else "false"

    return _search_worlds(
        params,
        limit,
        delay,
        headers,
        endpoint="worlds/active" if active else "worlds",
    )


def extract_info(world: dict) -> Dict[str, object]:
    return {
        "worldId": world.get("id"),
        "世界名稱": world.get("name"),
        "世界ID": world.get("id"),
        "世界大小": world.get("capacity"),
        "上傳日期": world.get("created_at"),
        "更新日期": world.get("updated_at"),
        "實驗室日期": world.get("labsPublicationDate"),
        "發布日期": world.get("publicationDate"),
        "瀏覽人次": world.get("visits"),
        "收藏人次": world.get("favorites"),
        "世界熱度": world.get("heat"),
        "世界熱門度": world.get("popularity"),
        "Tag": world.get("tags"),
        "世界連結": f"https://vrchat.com/home/world/{world.get('id')}",
    }


def fetch_worlds(*,
                 keyword: Optional[str] = None,
                 user_id: Optional[str] = None,
                 limit: int = 20,
                 delay: float = 1.0,
                 headers: Optional[Dict[str, str]] = None) -> List[dict]:
    """High level helper to fetch worlds by keyword or user ID."""
    if keyword:
        return search_worlds(keyword, limit, delay, headers)
    if user_id:
        return get_user_worlds(user_id, limit, delay, headers)
    raise ValueError("keyword or user_id required")


def fetch_world_by_id(world_id: str,
                      headers: Optional[Dict[str, str]] = None) -> Optional[dict]:
    """Fetch full world data for a single world ID.

    The individual world endpoint reliably returns ``visits`` and other
    counters even when the list-search endpoint omits them.
    Returns ``None`` on any network or parse error.
    """
    if requests is None:
        return None
    client = _build_http_client()
    url = f"https://api.vrchat.cloud/api/1/worlds/{world_id}"
    try:
        r = client.get(url, headers=headers or HEADERS, timeout=30)
        r.raise_for_status()
        return r.json()
    except Exception:
        return None


def enrich_visits(worlds: List[dict],
                  headers: Optional[Dict[str, str]] = None,
                  delay: float = 0.5) -> List[dict]:
    """Fill in missing ``visits`` (and related counters) via individual world fetches.

    For each world where ``visits`` is ``None`` (the list endpoint returned no
    data), a separate request to ``/api/1/worlds/{id}`` is made.  This is
    slower but ensures accurate visit counts.

    Parameters
    ----------
    worlds:
        Worlds returned by a list/search endpoint.
    headers:
        Auth headers (must contain a valid ``Cookie``).
    delay:
        Seconds to wait between individual requests (rate-limit courtesy).
    """
    enriched: List[dict] = []
    for w in worlds:
        if w.get("visits") is None:
            wid = w.get("id") or w.get("worldId")
            if wid:
                detail = fetch_world_by_id(wid, headers)
                if detail:
                    # Merge fresh counters from the full response
                    for field in ("visits", "favorites", "heat", "popularity",
                                  "capacity", "publicationDate", "labsPublicationDate",
                                  "updated_at", "created_at"):
                        if detail.get(field) is not None:
                            w = {**w, field: detail[field]}
                if delay:
                    time.sleep(delay)
        enriched.append(w)
    return enriched


def vrchat_check_session(cookie: str) -> dict:
    """Verify whether a saved cookie is still accepted by VRChat.

    Returns {"ok": True, "user": {...}} if valid,
            {"ok": False, "error": "..."} if expired or invalid.
    """
    if requests is None:
        return {"ok": False, "error": "requests package is required"}
    if not cookie:
        return {"ok": False, "error": "cookie が空です"}

    session = requests.Session()
    session.trust_env = False
    req_headers = {
        "User-Agent": HEADERS.get(
            "User-Agent", "VRRacingClubTW-WorldInfo/2.0 local-user"
        ),
        "Cookie": cookie,
    }
    try:
        r = session.get(
            "https://api.vrchat.cloud/api/1/auth/user",
            headers=req_headers,
            timeout=15,
        )
    except Exception as exc:
        return {"ok": False, "error": f"網路錯誤: {exc}"}

    if r.status_code == 401:
        return {"ok": False, "error": "session 已過期，請重新登入"}
    if not r.ok:
        return {"ok": False, "error": f"伺服器錯誤 {r.status_code}"}

    try:
        data = r.json()
    except Exception:
        return {"ok": False, "error": "回應格式錯誤"}

    if "requiresTwoFactorAuth" in data:
        return {"ok": False, "error": "session 需要重新進行 2FA 驗證"}

    return {"ok": True, "user": data}


def vrchat_login(username: str, password: str) -> dict:
    """Authenticate with VRChat API using username and password.

    Returns a dict with keys:
      ok            – True on success
      requires_2fa  – True if a second-factor code is needed
      methods       – list of accepted 2FA methods ("totp", "otp", "emailotp")
      auth_cookie   – raw cookie string holding the partial auth token
      cookie        – full cookie string (only when ok=True and no 2FA required)
      user          – VRChat user dict (only on full success)
      error         – human-readable error message (only on failure)
    """
    if requests is None:
        return {"ok": False, "error": "requests package is required"}

    import uuid as _uuid
    session = requests.Session()
    session.trust_env = False

    token = base64.b64encode(f"{username}:{password}".encode()).decode()
    req_headers = {
        "User-Agent": HEADERS.get(
            "User-Agent", "VRRacingClubTW-WorldInfo/2.0 local-user"
        ),
        "Authorization": f"Basic {token}",
    }

    # Supply a machineId cookie so VRChat doesn't flag the request as suspicious
    machine_id = str(_uuid.uuid4())
    session.cookies.set("machineId", machine_id)

    try:
        r = session.get(
            "https://api.vrchat.cloud/api/1/auth/user",
            headers=req_headers,
            timeout=30,
        )
    except Exception as exc:
        return {"ok": False, "error": f"網路錯誤: {exc}"}

    if r.status_code == 401:
        return {"ok": False, "error": "帳號或密碼錯誤（401）"}
    if r.status_code == 403:
        return {"ok": False, "error": "存取被拒絕（403）"}
    if not r.ok:
        return {"ok": False, "error": f"伺服器錯誤 {r.status_code}"}

    try:
        data = r.json()
    except Exception:
        return {"ok": False, "error": "回應格式錯誤，無法解析 JSON"}

    # Build cookie string from the session jar
    cookie_str = "; ".join(f"{k}={v}" for k, v in session.cookies.items())

    if "requiresTwoFactorAuth" in data:
        methods = data["requiresTwoFactorAuth"]  # e.g. ["totp", "otp"]
        return {
            "ok": False,
            "requires_2fa": True,
            "methods": methods,
            "auth_cookie": cookie_str,
            "error": "需要雙因素驗證",
        }

    return {
        "ok": True,
        "requires_2fa": False,
        "cookie": cookie_str,
        "user": data,
    }


def vrchat_verify_2fa(code: str, method: str, auth_cookie: str) -> dict:
    """Verify a 2FA code obtained after ``vrchat_login``.

    Parameters
    ----------
    code:
        6-digit code from an authenticator app / email.
    method:
        One of "totp", "otp", "emailotp".
    auth_cookie:
        The ``auth_cookie`` value returned by ``vrchat_login``.

    Returns a dict with keys:
      ok      – True on success
      cookie  – complete cookie string ready for use in HTTP headers
      error   – human-readable error (only on failure)
    """
    if requests is None:
        return {"ok": False, "error": "requests package is required"}

    session = requests.Session()
    session.trust_env = False

    # Restore the partial-auth cookies into the session
    for part in auth_cookie.split(";"):
        k, _, v = part.strip().partition("=")
        if k:
            session.cookies.set(k.strip(), v.strip())

    req_headers = {
        "User-Agent": HEADERS.get(
            "User-Agent", "VRRacingClubTW-WorldInfo/2.0 local-user"
        ),
        "Content-Type": "application/json",
    }

    url = f"https://api.vrchat.cloud/api/1/auth/twofactorauth/{method}/verify"
    try:
        r = session.post(url, json={"code": code}, headers=req_headers, timeout=30)
    except Exception as exc:
        return {"ok": False, "error": f"網路錯誤: {exc}"}

    if r.status_code in (400, 401):
        return {"ok": False, "error": "驗證碼錯誤或已過期"}
    if not r.ok:
        return {"ok": False, "error": f"伺服器錯誤 {r.status_code}"}

    cookie_str = "; ".join(f"{k}={v}" for k, v in session.cookies.items())
    return {"ok": True, "cookie": cookie_str}


def main() -> None:
    import argparse

    parser = argparse.ArgumentParser(description="Query VRChat worlds")
    parser.add_argument("--keyword", help="search keyword")
    parser.add_argument("--user", help="creator userId")
    parser.add_argument("--limit", type=int, default=20, help="maximum worlds")
    parser.add_argument("--delay", type=float, default=1.0, help="seconds between requests")
    parser.add_argument("--out", type=Path, default=BASE / "raw_worlds.json", help="output JSON path")
    parser.add_argument("--cookie", help="authentication cookie string")
    parser.add_argument("--username", help="basic auth username")
    parser.add_argument("--password", help="basic auth password")
    args = parser.parse_args()

    global HEADERS
    HEADERS = _load_headers(args.cookie, args.username, args.password)

    if args.keyword:
        worlds = search_worlds(args.keyword, args.limit, args.delay, HEADERS)
    elif args.user:
        worlds = get_user_worlds(args.user, args.limit, args.delay, HEADERS)
    else:
        parser.error("--keyword or --user is required")

    parsed = [extract_info(w) for w in worlds]
    update_history(worlds)
    with open(args.out, "w", encoding="utf-8") as f:
        json.dump(parsed, f, ensure_ascii=False, indent=2)
    print(f"Saved {len(parsed)} worlds to {args.out}")


if __name__ == "__main__":
    main()

