from __future__ import annotations

import datetime as dt
import base64
import copy
import json
import logging
import os
import re
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from world_info.scraper.scraper import (
    VRChatRateLimitError,
    _load_headers,
    _parse_date,
    enrich_visits,
    fetch_worlds,
    search_worlds_query,
    vrchat_check_session,
    vrchat_login,
    vrchat_verify_2fa,
)

from .storage import WorldInfoStorage

logger = logging.getLogger(__name__)

ANALYSIS_CACHE_LIMIT = 40


LEGACY_SOURCE_LABELS = {
    "legacy-raw": "Legacy keyword JSON",
    "legacy-user": "Legacy user JSON",
    "legacy-taiwan": "Legacy Taiwan workbook",
    "legacy-starriver": "Legacy StarRiver workbook",
}

IMPORT_SOURCE_LABELS = {
    "import:legacy-raw": "Imported legacy keyword JSON",
    "import:legacy-user": "Imported legacy user JSON",
    "import:legacy-taiwan": "Imported legacy Taiwan workbook",
    "import:legacy-starriver": "Imported legacy StarRiver workbook",
    "history:legacy": "Imported legacy history",
}


class WorldInfoService:
    def __init__(
        self,
        repo_root: Path | None = None,
        app_root: Path | None = None,
        jobs_path: Path | None = None,
        topics_path: Path | None = None,
        world_properties_path: Path | None = None,
    ) -> None:
        self.repo_root = (repo_root or Path(__file__).resolve().parents[2]).resolve()
        self.app_root = (app_root or Path(__file__).resolve().parents[1]).resolve()
        self.frontend_dir = self.app_root / "frontend"
        self.data_dir = self.app_root / "data"
        self.data_dir.mkdir(parents=True, exist_ok=True)

        self._job_display_cache: dict[str, tuple[float, dict]] = {}
        self.legacy_root = self.repo_root / "world_info"
        self.legacy_scraper_dir = self.legacy_root / "scraper"
        self.legacy_analytics_dir = self.repo_root / "analytics"
        self.storage = WorldInfoStorage(self.data_dir / "world_info.sqlite3")
        self.jobs_path = jobs_path or (self.app_root / "config" / "sync_jobs.json")
        self.topics_path = topics_path or (self.app_root / "config" / "topics.json")
        self.world_properties_path = world_properties_path or (self.app_root / "config" / "world_properties.json")

        self.legacy_sources = {
            "legacy-raw": {
                "path": self.legacy_scraper_dir / "raw_worlds.json",
                "kind": "json",
                "origin": "legacy",
            },
            "legacy-user": {
                "path": self.legacy_scraper_dir / "user_worlds.json",
                "kind": "json",
                "origin": "legacy",
            },
            "legacy-taiwan": {
                "path": self.legacy_scraper_dir / "TaiwanWorlds.xlsx",
                "kind": "xlsx",
                "origin": "legacy",
            },
            "legacy-starriver": {
                "path": self.legacy_scraper_dir / "StarRiverArts.xlsx",
                "kind": "xlsx",
                "origin": "legacy",
            },
        }
        try:
            self._sync_topic_catalog(refresh=False)
        except Exception as exc:
            logger.warning("Topic catalog sync skipped during startup: %s", exc)

    def list_topics(self, *, include_inactive: bool = False) -> list[dict[str, Any]]:
        items = []
        configs = self._load_topic_configs()
        for topic in self.storage.list_topics():
            if not include_inactive and not bool(topic.get("is_active", 1)):
                continue
            resolved = self._resolve_topic_config(topic["topic_key"], configs.get(topic["topic_key"], {}))
            rules = self.storage.list_topic_rules(topic["topic_key"])
            items.append(
                {
                    "topic_key": topic["topic_key"],
                    "label": topic["label"],
                    "description": topic.get("description"),
                    "color": topic.get("color"),
                    "topic_type": resolved.get("topic_type", "job"),
                    "sort_order": topic.get("sort_order", 0),
                    "is_active": bool(topic.get("is_active", 1)),
                    "world_count": topic.get("world_count", 0),
                    "last_seen_at": topic.get("last_seen_at"),
                    "rules": rules,
                    "summary": {
                        "world_count": topic.get("world_count", 0),
                        "tracked_creators": 0,
                        "total_visits": 0,
                        "total_favorites": 0,
                        "new_worlds_7d": 0,
                        "updated_worlds_30d": 0,
                    },
                }
            )
        return items

    def get_topic(self, topic_key: str) -> dict[str, Any]:
        topic = self.storage.get_topic(topic_key)
        if topic is None:
            raise KeyError(f"Unknown topic: {topic_key}")
        resolved = self._resolve_topic_config(topic_key, self._load_topic_configs().get(topic_key, {}))
        dashboard = self.get_topic_dashboard(topic_key)
        return {
            "topic_key": topic["topic_key"],
            "label": topic["label"],
            "description": topic.get("description"),
            "color": topic.get("color"),
            "topic_type": resolved.get("topic_type", "job"),
            "sort_order": topic.get("sort_order", 0),
            "is_active": bool(topic.get("is_active", 1)),
            "rules": self.storage.list_topic_rules(topic_key),
            "summary": dashboard["summary"],
            "trend": dashboard["trend"],
            "top_worlds": dashboard["top_worlds"],
            "top_creators": dashboard["top_creators"],
            "new_worlds": dashboard["new_worlds"],
        }

    def load_topic_worlds(
        self,
        topic_key: str,
        *,
        query: str | None = None,
        tag: str | None = None,
        sort: str = "visits",
        direction: str = "desc",
    ) -> list[dict[str, Any]]:
        membership_rows = self.storage.list_topic_memberships(topic_key)
        membership_map = {row["world_id"]: row for row in membership_rows}
        rules = [rule for rule in self.storage.list_topic_rules(topic_key) if rule.get("is_active", 1)]
        source_rules = [rule for rule in rules if rule.get("rule_type") == "source"]
        other_rules = [rule for rule in rules if rule.get("rule_type") != "source"]
        if source_rules and not other_rules:
            topic_history = self._load_history_for_sources([rule.get("rule_value") or "" for rule in source_rules])
            worlds_by_id: dict[str, dict[str, Any]] = {}
            for rule in source_rules:
                for world in self._load_topic_source_worlds(rule.get("rule_value") or ""):
                    world_id = world.get("id")
                    if world_id:
                        worlds_by_id[world_id] = world
            worlds = list(worlds_by_id.values())
            if query:
                needle = query.casefold()
                worlds = [
                    world
                    for world in worlds
                    if needle in (world.get("name") or "").casefold()
                    or needle in (world.get("id") or "").casefold()
                    or needle in (world.get("author_name") or "").casefold()
                ]
            if tag and tag != "all":
                worlds = [world for world in worlds if tag in world.get("tags", [])]
            worlds = self._sort_worlds(worlds, sort=sort, direction=direction, history=topic_history)
        else:
            worlds = [
                world
                for world in self.load_worlds("db:all", query=query, tag=tag, sort=sort, direction=direction)
                if world.get("id") in membership_map
            ]
        for world in worlds:
            membership = membership_map.get(world.get("id"))
            if membership:
                world["topic_first_seen_at"] = membership["first_seen_at"]
                world["topic_last_seen_at"] = membership["last_seen_at"]
                world["topic_matched_by"] = membership.get("matched_by")
        return worlds

    def get_topic_dashboard(self, topic_key: str) -> dict[str, Any]:
        topic = self.storage.get_topic(topic_key)
        if topic is None:
            raise KeyError(f"Unknown topic: {topic_key}")

        worlds = self.load_topic_worlds(topic_key)
        today = dt.datetime.now(dt.timezone.utc)
        summary = {
            "world_count": len(worlds),
            "total_visits": sum(self._to_int(world.get("visits")) for world in worlds),
            "total_favorites": sum(self._to_int(world.get("favorites")) for world in worlds),
            "tracked_creators": len({world.get("author_id") for world in worlds if world.get("author_id")}),
            "new_worlds_7d": sum(
                1
                for world in worlds
                if self._within_days(world.get("topic_first_seen_at"), 7, today)
            ),
            "updated_worlds_30d": sum(
                1
                for world in worlds
                if self._within_days(world.get("updated_at"), 30, today)
            ),
        }

        top_worlds = worlds[:5]

        creator_rows = {}
        for world in worlds:
            creator_key = world.get("author_id") or world.get("author_name")
            if not creator_key:
                continue
            bucket = creator_rows.setdefault(
                creator_key,
                {
                    "author_id": world.get("author_id"),
                    "author_name": world.get("author_name"),
                    "world_count": 0,
                    "total_visits": 0,
                    "total_favorites": 0,
                },
            )
            bucket["world_count"] += 1
            bucket["total_visits"] += self._to_int(world.get("visits"))
            bucket["total_favorites"] += self._to_int(world.get("favorites"))
        top_creators = sorted(
            creator_rows.values(),
            key=lambda item: (item["world_count"], item["total_visits"]),
            reverse=True,
        )[:5]

        new_worlds = sorted(
            [
                world
                for world in worlds
                if self._within_days(world.get("topic_first_seen_at"), 30, today)
            ],
            key=lambda item: self._date_score(item.get("topic_first_seen_at")),
            reverse=True,
        )[:8]

        trend_rows = {}
        history = self._load_topic_history(topic_key)
        world_ids = {world.get("id") for world in worlds if world.get("id")}
        for world_id in world_ids:
            for entry in history.get(world_id, []):
                date_label = self._date_bucket(entry.get("iso_time"))
                if not date_label:
                    continue
                bucket = trend_rows.setdefault(
                    date_label,
                    {"date": date_label, "world_count": 0, "visits": 0, "favorites": 0},
                )
                bucket["world_count"] += 1
                bucket["visits"] += self._to_int(entry.get("visits"))
                bucket["favorites"] += self._to_int(entry.get("favorites"))
        trend = [trend_rows[key] for key in sorted(trend_rows.keys())][-20:]

        return {
            "topic_key": topic_key,
            "label": topic["label"],
            "summary": summary,
            "trend": trend,
            "top_worlds": top_worlds,
            "top_creators": top_creators,
            "new_worlds": new_worlds,
        }

    def list_sources(self) -> list[dict[str, Any]]:
        items: list[dict[str, Any]] = []

        if self.storage.has_data():
            items.append(
                {
                    "key": "db:all",
                    "label": "db:all",
                    "origin": "db",
                    "path": self._display_path(self.storage.db_path),
                    "available": True,
                    "count": self.storage.count_latest_worlds(),
                    "tags": [],
                }
            )

        for row in self.storage.list_db_sources():
            public_key = self._public_db_source_key(row["source_key"])
            items.append(
                {
                    "key": public_key,
                    "label": self._label_for_db_source(row["source_key"]),
                    "origin": "db",
                    "path": self._display_path(self.storage.db_path),
                    "available": True,
                    "count": row["world_count"],
                    "tags": [],
                    "latest_fetched_at": row["latest_fetched_at"],
                }
            )

        for key, config in self.legacy_sources.items():
            path = Path(config["path"])
            items.append(
                {
                    "key": key,
                    "label": LEGACY_SOURCE_LABELS.get(key, key),
                    "origin": "legacy",
                    "path": self._display_path(path),
                    "available": path.exists(),
                    # Keep source discovery cheap. Full legacy loads belong to explicit
                    # collection views, not to the global boot-time source picker.
                    "count": 0,
                    "tags": [],
                }
            )
        return items

    def list_jobs(self) -> list[dict[str, Any]]:
        items = []
        for job_key, config in sorted(self._load_job_configs().items()):
            resolved = self._resolve_job_config(job_key, config)
            items.append(
                {
                    "job_key": job_key,
                    "label": resolved["label"],
                    "type": resolved["type"],
                    "source": self._public_db_source_key(resolved["source_key"]),
                    "ready": resolved["ready"],
                    "reason": resolved["reason"],
                    "creator_review_enabled": bool(resolved.get("creator_review_enabled")),
                    "keywords": resolved.get("keywords", []),
                    "user_id": resolved.get("user_id"),
                    "limit": resolved.get("limit"),
                    "limit_per_keyword": resolved.get("limit_per_keyword"),
                    "search": resolved.get("search"),
                    "tags": resolved.get("tags", []),
                    "notags": resolved.get("notags", []),
                    "sort": resolved.get("sort"),
                    "order": resolved.get("order"),
                    "featured": resolved.get("featured"),
                    "active": resolved.get("active"),
                    "release_status": resolved.get("release_status"),
                    "platform": resolved.get("platform"),
                    "latest_run": self._decorate_run(self.storage.get_latest_run_for_job(job_key)),
                }
            )
        return items

    def load_communities_workspace(self) -> dict[str, Any]:
        topics = self.list_topics(include_inactive=True)
        view_topics = [item for item in topics if item.get("topic_type") == "view"]
        tracked_topics = [item for item in topics if item.get("topic_type") == "job"]
        groups = self.storage.list_groups(limit=12)
        managed_groups = self.storage.list_managed_groups()
        scheduled_posts = self.storage.list_scheduled_posts(limit=8)
        memberships = self.storage.list_group_world_memberships(limit=200)
        membership_groups = {item.get("group_id") for item in memberships if item.get("group_id")}
        world_lookup = self._load_world_lookup_for_group_memberships(memberships)

        summary = {
            "group_count": self.storage.count_groups(),
            "managed_group_count": self.storage.count_managed_groups(),
            "scheduled_post_count": self.storage.count_scheduled_posts(),
            "scheduled_post_pending_count": self.storage.count_scheduled_posts("pending"),
            "group_world_link_count": len(memberships),
            "linked_group_count": len(membership_groups),
            "tracked_creator_count": self.storage.count_distinct_authors(),
            "tracked_view_count": len(tracked_topics),
            "saved_view_count": len(view_topics),
        }

        directory_items = [
            {
                "group_id": item.get("group_id"),
                "name": item.get("name") or item.get("group_id"),
                "region": item.get("region"),
                "category": item.get("category"),
                "description": item.get("description"),
                "managed_status": item.get("managed_status"),
                "world_count": self._to_int(item.get("world_count")),
                "last_synced_at": item.get("last_synced_at"),
                "external_links": item.get("external_links") or [],
            }
            for item in groups[:8]
        ]

        managed_items = [
            {
                "group_id": item.get("group_id"),
                "name": item.get("name") or item.get("group_id"),
                "workspace_key": item.get("workspace_key"),
                "posting_enabled": bool(item.get("posting_enabled")),
                "notes": item.get("notes"),
                "updated_at": item.get("updated_at"),
            }
            for item in managed_groups[:8]
        ]

        publishing_items = [
            {
                "id": item.get("id"),
                "group_id": item.get("group_id"),
                "group_name": item.get("group_name") or item.get("group_id"),
                "content_type": item.get("content_type"),
                "status": item.get("status"),
                "scheduled_for": item.get("scheduled_for"),
                "payload": item.get("payload") or {},
                "delivered_at": item.get("delivered_at"),
            }
            for item in scheduled_posts
        ]

        membership_items = [self._enrich_group_world_membership(item, world_lookup) for item in memberships[:24]]
        worlds_by_group: list[dict[str, Any]] = []
        membership_map: dict[str, list[dict[str, Any]]] = {}
        for item in membership_items:
            membership_map.setdefault(item.get("group_id") or "", []).append(item)
        for group in groups[:8]:
            group_id = group.get("group_id")
            linked_items = membership_map.get(group_id or "", [])
            if not linked_items:
                continue
            linked_items.sort(
                key=lambda item: (
                    self._to_int(item.get("visits")),
                    self._to_int(item.get("favorites")),
                    self._date_score(item.get("linked_at")),
                ),
                reverse=True,
            )
            worlds_by_group.append(
                {
                    "group_id": group_id,
                    "group_name": group.get("name") or group_id,
                    "world_count": len(linked_items),
                    "top_worlds": linked_items[:4],
                }
            )

        return {
            "generated_at": dt.datetime.now(dt.UTC).isoformat(),
            "summary": summary,
            "directory": {
                "items": directory_items,
                "status": "group tables ready",
            },
            "growth": {
                "status": "waiting for group snapshots",
                "signals": [
                    "total visits / favorites growth by group",
                    "new worlds per group over 7d / 30d",
                    "breakout hit rate and dormant revival inside each group",
                ],
            },
            "worlds": {
                "status": "membership view ready" if membership_items else "no group-world memberships yet",
                "count": len(membership_items),
                "linked_group_count": len(membership_groups),
                "items": membership_items[:12],
                "groups": worlds_by_group,
                "signals": [
                    "group-local breakout candidates",
                    "group quality picks and update winners",
                    "top worlds attached to each group identity",
                ],
            },
            "publishing": {
                "items": publishing_items,
                "managed_groups": managed_items,
                "status": "publishing tables ready",
            },
        }

    def list_groups(self, *, limit: int = 100) -> dict[str, Any]:
        return {
            "count": self.storage.count_groups(),
            "items": self.storage.list_groups(limit=limit),
        }

    def upsert_group(
        self,
        *,
        group_id: Any,
        name: Any,
        region: Any = None,
        category: Any = None,
        description: Any = None,
        managed_status: Any = None,
        external_links: Any = None,
        last_synced_at: Any = None,
    ) -> dict[str, Any]:
        cleaned_group_id = self._clean_optional_text(group_id)
        cleaned_name = self._clean_optional_text(name)
        if not cleaned_group_id:
            raise ValueError("group_id is required")
        if not cleaned_name:
            raise ValueError("name is required")
        if isinstance(external_links, str):
            cleaned_links = self._csv_items(external_links)
        elif isinstance(external_links, list):
            cleaned_links = self._csv_items(external_links)
        else:
            cleaned_links = []
        self.storage.upsert_group(
            group_id=cleaned_group_id,
            name=cleaned_name,
            region=self._clean_optional_text(region),
            category=self._clean_optional_text(category),
            description=self._clean_optional_text(description),
            managed_status=self._clean_optional_text(managed_status) or "observed",
            external_links=cleaned_links,
            last_synced_at=self._clean_optional_text(last_synced_at) or dt.datetime.now(dt.UTC).isoformat(),
        )
        item = self.storage.get_group(cleaned_group_id)
        if item is None:
            raise RuntimeError("Failed to save group")
        return item

    def delete_group(self, group_id: Any) -> dict[str, Any]:
        cleaned_group_id = self._clean_optional_text(group_id)
        if not cleaned_group_id:
            raise ValueError("group_id is required")
        if self.storage.get_group(cleaned_group_id) is None:
            raise KeyError(f"Unknown group: {cleaned_group_id}")
        self.storage.delete_group(cleaned_group_id)
        return {"status": "deleted", "group_id": cleaned_group_id}

    def list_group_world_memberships(
        self,
        *,
        group_id: Any = None,
        limit: int = 200,
    ) -> dict[str, Any]:
        cleaned_group_id = self._clean_optional_text(group_id)
        items = self.storage.list_group_world_memberships(group_id=cleaned_group_id, limit=limit)
        world_lookup = self._load_world_lookup_for_group_memberships(items)
        enriched = [self._enrich_group_world_membership(item, world_lookup) for item in items]
        return {
            "count": len(enriched),
            "items": enriched,
        }

    def upsert_group_world_membership(
        self,
        *,
        group_id: Any,
        world_id: Any,
        membership_role: Any = None,
        source_key: Any = None,
    ) -> dict[str, Any]:
        cleaned_group_id = self._clean_optional_text(group_id)
        cleaned_world_id = self._clean_optional_text(world_id)
        if not cleaned_group_id:
            raise ValueError("group_id is required")
        if not cleaned_world_id:
            raise ValueError("world_id is required")
        if self.storage.get_group(cleaned_group_id) is None:
            raise KeyError(f"Unknown group: {cleaned_group_id}")
        world_lookup = self._load_world_lookup_for_group_memberships(
            [{"world_id": cleaned_world_id, "group_id": cleaned_group_id}]
        )
        if cleaned_world_id not in world_lookup:
            raise KeyError(f"Unknown world: {cleaned_world_id}")
        self.storage.upsert_group_world_membership(
            group_id=cleaned_group_id,
            world_id=cleaned_world_id,
            membership_role=self._clean_optional_text(membership_role) or "member",
            linked_at=dt.datetime.now(dt.UTC).isoformat(),
            source_key=self._clean_optional_text(source_key),
        )
        item = self.storage.get_group_world_membership(group_id=cleaned_group_id, world_id=cleaned_world_id)
        if item is None:
            raise RuntimeError("Failed to save group world membership")
        return self._enrich_group_world_membership(item, world_lookup)

    def delete_group_world_membership(self, *, group_id: Any, world_id: Any) -> dict[str, Any]:
        cleaned_group_id = self._clean_optional_text(group_id)
        cleaned_world_id = self._clean_optional_text(world_id)
        if not cleaned_group_id:
            raise ValueError("group_id is required")
        if not cleaned_world_id:
            raise ValueError("world_id is required")
        if self.storage.get_group_world_membership(group_id=cleaned_group_id, world_id=cleaned_world_id) is None:
            raise KeyError(f"Unknown membership: {cleaned_group_id}/{cleaned_world_id}")
        self.storage.delete_group_world_membership(group_id=cleaned_group_id, world_id=cleaned_world_id)
        return {"status": "deleted", "group_id": cleaned_group_id, "world_id": cleaned_world_id}

    def list_managed_groups(self) -> dict[str, Any]:
        items = self.storage.list_managed_groups()
        return {
            "count": len(items),
            "items": items,
        }

    def upsert_managed_group(
        self,
        *,
        group_id: Any,
        workspace_key: Any = None,
        posting_enabled: Any = False,
        notes: Any = None,
    ) -> dict[str, Any]:
        cleaned_group_id = self._clean_optional_text(group_id)
        if not cleaned_group_id:
            raise ValueError("group_id is required")
        if self.storage.get_group(cleaned_group_id) is None:
            raise KeyError(f"Unknown group: {cleaned_group_id}")
        self.storage.upsert_managed_group(
            group_id=cleaned_group_id,
            workspace_key=self._clean_optional_text(workspace_key),
            posting_enabled=bool(self._optional_bool(posting_enabled) or False),
            notes=self._clean_optional_text(notes),
            updated_at=dt.datetime.now(dt.UTC).isoformat(),
        )
        item = self.storage.get_managed_group(cleaned_group_id)
        if item is None:
            raise RuntimeError("Failed to save managed group")
        return item

    def delete_managed_group(self, group_id: Any) -> dict[str, Any]:
        cleaned_group_id = self._clean_optional_text(group_id)
        if not cleaned_group_id:
            raise ValueError("group_id is required")
        if self.storage.get_managed_group(cleaned_group_id) is None:
            raise KeyError(f"Unknown managed group: {cleaned_group_id}")
        self.storage.delete_managed_group(cleaned_group_id)
        return {"status": "deleted", "group_id": cleaned_group_id}

    def list_scheduled_posts(self, *, limit: int = 50) -> dict[str, Any]:
        return {
            "count": self.storage.count_scheduled_posts(),
            "items": self.storage.list_scheduled_posts(limit=limit),
        }

    def upsert_scheduled_post(
        self,
        *,
        post_id: Any = None,
        group_id: Any,
        content_type: Any,
        status: Any,
        scheduled_for: Any,
        payload: Any = None,
        delivered_at: Any = None,
    ) -> dict[str, Any]:
        cleaned_group_id = self._clean_optional_text(group_id)
        cleaned_content_type = self._clean_optional_text(content_type)
        cleaned_status = self._clean_optional_text(status)
        cleaned_scheduled_for = self._clean_optional_text(scheduled_for)
        if not cleaned_group_id:
            raise ValueError("group_id is required")
        if self.storage.get_group(cleaned_group_id) is None:
            raise KeyError(f"Unknown group: {cleaned_group_id}")
        if not cleaned_content_type:
            raise ValueError("content_type is required")
        if not cleaned_status:
            raise ValueError("status is required")
        if not cleaned_scheduled_for:
            raise ValueError("scheduled_for is required")
        existing_post_id = self._to_optional_int(post_id)
        existing = None
        if existing_post_id is not None:
            existing = self.storage.get_scheduled_post(existing_post_id)
            if existing is None:
                raise KeyError(f"Unknown scheduled post: {existing_post_id}")
        payload_dict: dict[str, Any]
        if payload in (None, ""):
            payload_dict = {}
        elif isinstance(payload, dict):
            payload_dict = payload
        elif isinstance(payload, str):
            try:
                loaded = json.loads(payload)
            except json.JSONDecodeError as exc:
                raise ValueError("payload must be valid JSON") from exc
            if not isinstance(loaded, dict):
                raise ValueError("payload must decode to a JSON object")
            payload_dict = loaded
        else:
            raise ValueError("payload must be a JSON object")
        now = dt.datetime.now(dt.UTC).isoformat()
        saved_post_id = self.storage.upsert_scheduled_post(
            post_id=existing_post_id,
            group_id=cleaned_group_id,
            content_type=cleaned_content_type,
            status=cleaned_status,
            scheduled_for=cleaned_scheduled_for,
            payload=payload_dict,
            created_at=(existing or {}).get("created_at") or now,
            updated_at=now,
            delivered_at=self._clean_optional_text(delivered_at),
        )
        item = self.storage.get_scheduled_post(saved_post_id)
        if item is None:
            raise RuntimeError("Failed to save scheduled post")
        return item

    def delete_scheduled_post(self, post_id: Any) -> dict[str, Any]:
        cleaned_post_id = self._to_optional_int(post_id)
        if cleaned_post_id is None:
            raise ValueError("post_id is required")
        if self.storage.get_scheduled_post(cleaned_post_id) is None:
            raise KeyError(f"Unknown scheduled post: {cleaned_post_id}")
        self.storage.delete_scheduled_post(cleaned_post_id)
        return {"status": "deleted", "post_id": cleaned_post_id}

    def list_job_diagnostics(self) -> list[dict[str, Any]]:
        configs = self._load_job_configs()
        items: list[dict[str, Any]] = []
        for job in self.list_jobs():
            job_key = job["job_key"]
            resolved = self._resolve_job_config(job_key, configs.get(job_key, {}))
            raw_runs = self.storage.list_runs(limit=20, job_key=job_key)
            completed_runs = [run for run in raw_runs if run.get("status") == "completed"]
            latest_completed_run = completed_runs[0] if completed_runs else None
            worlds = (
                [self._normalise_db_world(world) for world in self.storage.load_run_worlds(int(latest_completed_run["id"]))]
                if latest_completed_run
                else []
            )
            current_creator_count = len({world.get("author_id") for world in worlds if world.get("author_id")})
            item = {
                "job_key": job_key,
                "label": job["label"],
                "type": job["type"],
                "source": job["source"],
                "ready": job["ready"],
                "reason": job["reason"],
                "creator_review_enabled": job["creator_review_enabled"],
                "keyword_count": len(resolved.get("keywords", [])),
                "keywords": resolved.get("keywords", []),
                "search": resolved.get("search"),
                "tags": resolved.get("tags", []),
                "notags": resolved.get("notags", []),
                "sort": resolved.get("sort"),
                "order": resolved.get("order"),
                "featured": resolved.get("featured"),
                "active": resolved.get("active"),
                "limit": resolved.get("limit"),
                "limit_per_keyword": resolved.get("limit_per_keyword"),
                "world_blacklist_count": len(self._load_blacklist(resolved.get("blacklist_file"))),
                "creator_whitelist_count": len(resolved.get("include_user_ids", [])),
                "creator_blacklist_count": len(resolved.get("exclude_author_ids", [])),
                "current_world_count": len(worlds),
                "current_creator_count": current_creator_count,
                "latest_run": self._decorate_run(raw_runs[0]) if raw_runs else None,
                "latest_completed_run": self._decorate_run(latest_completed_run) if latest_completed_run else None,
                "source_diff": self.get_job_source_diff(job_key),
            }
            items.append(item)
        return items

    def get_job_source_diff(
        self,
        job_key: str,
        *,
        added_limit: int | None = 5,
        removed_limit: int | None = 5,
        changed_limit: int | None = 5,
    ) -> dict[str, Any]:
        configs = self._load_job_configs()
        if job_key not in configs:
            raise KeyError(f"Unknown job: {job_key}")

        completed_runs = [
            run for run in self.storage.list_runs(limit=20, job_key=job_key)
            if run.get("status") == "completed"
        ]
        latest_run = completed_runs[0] if completed_runs else None
        previous_run = completed_runs[1] if len(completed_runs) > 1 else None
        if latest_run is None or previous_run is None:
            return {
                "job_key": job_key,
                "status": "insufficient_history",
                "latest_run": self._decorate_run(latest_run) if latest_run else None,
                "previous_run": self._decorate_run(previous_run) if previous_run else None,
                "added_count": 0,
                "removed_count": 0,
                "changed_count": 0,
                "added_worlds": [],
                "removed_worlds": [],
                "changed_worlds": [],
                "message": "Need at least two completed runs for a source diff.",
            }

        latest_worlds = {
            world.get("id"): self._normalise_db_world(world)
            for world in self.storage.load_run_worlds(int(latest_run["id"]))
            if world.get("id")
        }
        previous_worlds = {
            world.get("id"): self._normalise_db_world(world)
            for world in self.storage.load_run_worlds(int(previous_run["id"]))
            if world.get("id")
        }

        added_ids = sorted(set(latest_worlds) - set(previous_worlds))
        removed_ids = sorted(set(previous_worlds) - set(latest_worlds))
        changed_worlds = []
        for world_id in sorted(set(latest_worlds) & set(previous_worlds)):
            latest = latest_worlds[world_id]
            previous = previous_worlds[world_id]
            visits_delta = self._to_int(latest.get("visits")) - self._to_int(previous.get("visits"))
            favorites_delta = self._to_int(latest.get("favorites")) - self._to_int(previous.get("favorites"))
            heat_delta = self._to_int(latest.get("heat")) - self._to_int(previous.get("heat"))
            popularity_delta = self._to_int(latest.get("popularity")) - self._to_int(previous.get("popularity"))
            changed_fields = []
            for field in ("name", "author_id", "author_name", "updated_at", "release_status"):
                if (latest.get(field) or None) != (previous.get(field) or None):
                    changed_fields.append(field)
            if visits_delta:
                changed_fields.append("visits")
            if favorites_delta:
                changed_fields.append("favorites")
            if heat_delta:
                changed_fields.append("heat")
            if popularity_delta:
                changed_fields.append("popularity")
            if not changed_fields:
                continue
            changed_worlds.append(
                {
                    "id": world_id,
                    "name": latest.get("name") or previous.get("name") or world_id,
                    "author_name": latest.get("author_name") or previous.get("author_name"),
                    "visits_delta": visits_delta,
                    "favorites_delta": favorites_delta,
                    "heat_delta": heat_delta,
                    "popularity_delta": popularity_delta,
                    "changed_fields": sorted(set(changed_fields)),
                    "latest": self._world_preview(latest),
                    "previous": self._world_preview(previous),
                    "score": abs(visits_delta) + abs(favorites_delta) * 4 + abs(heat_delta) * 2 + abs(popularity_delta) * 2,
                }
            )

        changed_worlds.sort(key=lambda item: (item["score"], item["name"].casefold()), reverse=True)
        for item in changed_worlds:
            item.pop("score", None)

        return {
            "job_key": job_key,
            "status": "ok",
            "latest_run": self._decorate_run(latest_run),
            "previous_run": self._decorate_run(previous_run),
            "added_count": len(added_ids),
            "removed_count": len(removed_ids),
            "changed_count": len(changed_worlds),
            "added_worlds": [
                self._world_preview(latest_worlds[world_id])
                for world_id in (added_ids if added_limit is None else added_ids[:added_limit])
            ],
            "removed_worlds": [
                self._world_preview(previous_worlds[world_id])
                for world_id in (removed_ids if removed_limit is None else removed_ids[:removed_limit])
            ],
            "changed_worlds": changed_worlds if changed_limit is None else changed_worlds[:changed_limit],
            "message": "Compared the latest two completed runs for this job source.",
        }

    def create_job_with_topic(
        self,
        *,
        job_key: str,
        label: str,
        job_type: str,
        keywords: list[str] | None = None,
        user_id: str | None = None,
        limit: int | None = None,
        limit_per_keyword: int | None = None,
        search: str | None = None,
        tags: list[str] | str | None = None,
        notags: list[str] | str | None = None,
        sort: str | None = None,
        order: str | None = None,
        featured: bool | str | None = None,
        active: bool | str | None = None,
        release_status: str | None = None,
        platform: str | None = None,
    ) -> dict[str, Any]:
        job_key = self._slugify(job_key)
        label = str(label).strip() or job_key
        job_type = str(job_type).strip()
        jobs = self._load_job_configs()
        topics = self._load_topic_configs()

        if job_key in jobs:
            raise ValueError(f"job {job_key} already exists")
        if job_key in topics:
            raise ValueError(f"topic {job_key} already exists")

        if job_type == "keywords":
            cleaned_keywords = [str(item).strip() for item in (keywords or []) if str(item).strip()]
            if not cleaned_keywords:
                raise ValueError("keywords job requires at least one keyword")
            jobs[job_key] = {
                "label": label,
                "type": "keywords",
                "source_key": f"job:{job_key}",
                "keywords": cleaned_keywords,
                "limit_per_keyword": max(self._to_int(limit_per_keyword), 1) or 50,
            }
        elif job_type == "user":
            cleaned_user_id = str(user_id or "").strip()
            if not cleaned_user_id.startswith("usr_"):
                raise ValueError("user job requires a VRChat user ID like usr_...")
            jobs[job_key] = {
                "label": label,
                "type": "user",
                "source_key": f"job:{job_key}",
                "user_id": cleaned_user_id,
                "limit": max(self._to_int(limit), 1) or 50,
            }
        elif job_type in {"world_search", "worlds"}:
            cleaned_tags = self._csv_items(tags)
            cleaned_notags = self._csv_items(notags)
            cleaned_search = str(search or "").strip()
            jobs[job_key] = {
                "label": label,
                "type": "world_search",
                "source_key": f"job:{job_key}",
                "search": cleaned_search,
                "tags": cleaned_tags,
                "notags": cleaned_notags,
                "sort": str(sort or "popularity").strip() or "popularity",
                "order": str(order or "descending").strip() or "descending",
                "featured": self._optional_bool(featured),
                "active": bool(self._optional_bool(active) or False),
                "release_status": self._clean_optional_text(release_status),
                "platform": self._clean_optional_text(platform),
                "limit": max(self._to_int(limit), 1) or 50,
            }
        else:
            raise ValueError("job_type must be keywords, user, or world_search")

        next_sort_order = max((int(topic.get("sort_order", 0) or 0) for topic in topics.values()), default=0) + 10
        topics[job_key] = {
            "label": label,
            "description": f"{label} job worlds",
            "color": self._topic_color(job_key, len(topics)),
            "sort_order": next_sort_order,
            "rules": [{"type": "source", "value": f"db:job:{job_key}"}],
        }

        self._write_json(self.jobs_path, jobs)
        self._write_json(self.topics_path, topics)
        self._sync_topic_catalog(refresh=False)
        return {
          "status": "created",
          "job": next(item for item in self.list_jobs() if item["job_key"] == job_key),
          "topic": self.get_topic(job_key),
        }

    def update_job_with_topic(
        self,
        *,
        job_key: str,
        label: str,
        job_type: str,
        keywords: list[str] | None = None,
        user_id: str | None = None,
        limit: int | None = None,
        limit_per_keyword: int | None = None,
        search: str | None = None,
        tags: list[str] | str | None = None,
        notags: list[str] | str | None = None,
        sort: str | None = None,
        order: str | None = None,
        featured: bool | str | None = None,
        active: bool | str | None = None,
        release_status: str | None = None,
        platform: str | None = None,
    ) -> dict[str, Any]:
        job_key = self._slugify(job_key)
        jobs = self._load_job_configs()
        if job_key not in jobs:
            raise KeyError(f"Unknown job: {job_key}")
        label = str(label).strip() or job_key
        job_type = str(job_type).strip()
        existing = dict(jobs.get(job_key) or {})
        source_key = str(existing.get("source_key", f"job:{job_key}")).strip() or f"job:{job_key}"

        if job_type == "keywords":
            cleaned_keywords = [str(item).strip() for item in (keywords or []) if str(item).strip()]
            if not cleaned_keywords:
                raise ValueError("keywords job requires at least one keyword")
            updated = {
                "label": label,
                "type": "keywords",
                "source_key": source_key,
                "keywords": cleaned_keywords,
                "limit_per_keyword": max(self._to_int(limit_per_keyword), 1) or 50,
            }
            for key in (
                "blacklist_file",
                "include_user_ids_file",
                "include_user_ids",
                "exclude_author_ids_file",
                "creator_review_enabled",
                "blacklist_world_name_substrings",
            ):
                if key in existing:
                    updated[key] = existing[key]
        elif job_type == "user":
            cleaned_user_id = str(user_id or "").strip()
            if not cleaned_user_id.startswith("usr_"):
                raise ValueError("user job requires a VRChat user ID like usr_...")
            updated = {
                "label": label,
                "type": "user",
                "source_key": source_key,
                "user_id": cleaned_user_id,
                "limit": max(self._to_int(limit), 1) or 50,
            }
            if "creator_review_enabled" in existing:
                updated["creator_review_enabled"] = existing["creator_review_enabled"]
        elif job_type in {"world_search", "worlds"}:
            updated = {
                "label": label,
                "type": "world_search",
                "source_key": source_key,
                "search": str(search or "").strip(),
                "tags": self._csv_items(tags),
                "notags": self._csv_items(notags),
                "sort": str(sort or "popularity").strip() or "popularity",
                "order": str(order or "descending").strip() or "descending",
                "featured": self._optional_bool(featured),
                "active": bool(self._optional_bool(active) or False),
                "release_status": self._clean_optional_text(release_status),
                "platform": self._clean_optional_text(platform),
                "limit": max(self._to_int(limit), 1) or 50,
            }
        else:
            raise ValueError("job_type must be keywords, user, or world_search")

        jobs[job_key] = updated
        self._write_json(self.jobs_path, jobs)

        topics = self._load_topic_configs()
        if job_key in topics:
            topics[job_key]["label"] = label
            topics[job_key]["description"] = topics[job_key].get("description") or f"{label} job worlds"
            self._write_json(self.topics_path, topics)

        self._sync_topic_catalog(refresh=False)
        return {
            "status": "updated",
            "job": next(item for item in self.list_jobs() if item["job_key"] == job_key),
            "topic": self.get_topic(job_key),
        }

    def upsert_topic(
        self,
        *,
        topic_key: str,
        label: str,
        description: str | None = None,
        topic_type: str = "view",
        color: str | None = None,
        sort_order: int | None = None,
        is_active: bool = True,
        rules: list[dict[str, Any]] | None = None,
    ) -> dict[str, Any]:
        topic_key = self._slugify(topic_key)
        label = str(label).strip() or topic_key
        topic_type = str(topic_type or "view").strip().casefold()
        if topic_type not in {"job", "view"}:
            topic_type = "view"

        topics = self._load_topic_configs()
        if sort_order is None:
            sort_order = self._to_int((topics.get(topic_key) or {}).get("sort_order"))
        else:
            sort_order = self._to_int(sort_order)
        if not sort_order:
            sort_order = max((int(topic.get("sort_order", 0) or 0) for topic in topics.values()), default=0) + 10

        cleaned_rules: list[dict[str, str]] = []
        for rule in rules or []:
            if not isinstance(rule, dict):
                continue
            rule_type = str(rule.get("type") or rule.get("rule_type") or "").strip()
            rule_value = str(rule.get("value") or rule.get("rule_value") or "").strip()
            if rule_type and rule_value:
                cleaned_rules.append({"type": rule_type, "value": rule_value})
        if not cleaned_rules:
            raise ValueError("topic requires at least one rule")

        topics[topic_key] = {
            "label": label,
            "description": str(description or "").strip(),
            "topic_type": topic_type,
            "color": str(color or "").strip() or self._topic_color(topic_key, len(topics)),
            "sort_order": sort_order,
            "is_active": bool(self._optional_bool(is_active) if self._optional_bool(is_active) is not None else True),
            "rules": cleaned_rules,
        }
        self._write_json(self.topics_path, topics)
        self._sync_topic_catalog(refresh=False)
        if self.storage.has_data():
            self._refresh_topic_memberships(topic_keys={topic_key})
        return self.get_topic(topic_key)

    def set_topic_active(self, topic_key: str, is_active: bool) -> dict[str, Any]:
        topics = self._load_topic_configs()
        if topic_key not in topics:
            raise KeyError(f"Unknown topic: {topic_key}")
        topics[topic_key]["is_active"] = bool(is_active)
        self._write_json(self.topics_path, topics)
        self._sync_topic_catalog(refresh=False)
        return self.get_topic(topic_key)

    def delete_topic(self, topic_key: str) -> dict[str, Any]:
        topics = self._load_topic_configs()
        if topic_key not in topics and self.storage.get_topic(topic_key) is None:
            raise KeyError(f"Unknown topic: {topic_key}")
        if topic_key in topics:
            topics.pop(topic_key)
            self._write_json(self.topics_path, topics)
        self.storage.delete_topic(topic_key)
        return {"status": "deleted", "topic_key": topic_key}

    def delete_job(self, job_key: str, *, delete_topic: bool = True) -> dict[str, Any]:
        jobs = self._load_job_configs()
        if job_key not in jobs:
            raise KeyError(f"Unknown job: {job_key}")
        jobs.pop(job_key)
        self._write_json(self.jobs_path, jobs)

        topic_deleted = False
        if delete_topic:
            topics = self._load_topic_configs()
            if job_key in topics or self.storage.get_topic(job_key) is not None:
                if job_key in topics:
                    topics.pop(job_key)
                    self._write_json(self.topics_path, topics)
                self.storage.delete_topic(job_key)
                topic_deleted = True
        return {"status": "deleted", "job_key": job_key, "topic_deleted": topic_deleted}

    def list_runs(self, limit: int = 20) -> list[dict[str, Any]]:
        return [self._decorate_run(run) for run in self.storage.list_runs(limit=limit)]

    def record_rate_limit_event(
        self,
        *,
        error: VRChatRateLimitError,
        source_key: str | None,
        job_key: str | None,
        trigger_type: str | None,
        query_kind: str | None = None,
        query_value: str | None = None,
    ) -> dict[str, Any]:
        now = dt.datetime.now(dt.timezone.utc)
        retry_after_seconds = max(int(getattr(error, "retry_after_seconds", 0) or 0), 0)
        cooldown_seconds = self._calculate_rate_limit_cooldown(retry_after_seconds=retry_after_seconds, now=now)
        cooldown_until = (now + dt.timedelta(seconds=cooldown_seconds)).isoformat() if cooldown_seconds else now.isoformat()
        event_id = self.storage.insert_rate_limit_event(
            event_at=now.isoformat(),
            source_key=source_key,
            job_key=job_key,
            trigger_type=trigger_type,
            query_kind=query_kind,
            query_value=query_value,
            retry_after_seconds=retry_after_seconds,
            cooldown_seconds=cooldown_seconds,
            cooldown_until=cooldown_until,
            error_text=str(error),
        )
        return {
            "event_id": event_id,
            "retry_after_seconds": retry_after_seconds,
            "cooldown_seconds": cooldown_seconds,
            "cooldown_until": cooldown_until,
            "message": str(error),
        }

    def list_rate_limit_events(self, limit: int = 20) -> dict[str, Any]:
        items = self.storage.list_rate_limit_events(limit=limit)
        now = dt.datetime.now(dt.timezone.utc)
        count_24h = self.storage.count_rate_limit_events_since((now - dt.timedelta(hours=24)).isoformat())
        active_event = next(
            (
                item
                for item in items
                if item.get("cooldown_until")
                and (_parse_date(item.get("cooldown_until")) or now) > now
            ),
            None,
        )
        active_until = active_event.get("cooldown_until") if active_event else None
        remaining_seconds = 0
        if active_until:
            active_dt = _parse_date(active_until)
            if active_dt is not None:
                remaining_seconds = max(int((active_dt - now).total_seconds()), 0)
        summary = {
            "count_24h": count_24h,
            "active_cooldown_until": active_until,
            "active_cooldown_remaining_seconds": remaining_seconds,
            "latest_retry_after_seconds": int(active_event.get("retry_after_seconds") or 0) if active_event else 0,
            "strategy_hint": self._build_rate_limit_strategy_hint(
                count_24h=count_24h,
                remaining_seconds=remaining_seconds,
            ),
        }
        return {"summary": summary, "items": items}

    def list_query_analytics(self, limit_runs: int = 12) -> dict[str, Any]:
        recent_runs = self.storage.list_runs(limit=limit_runs)
        decorated_runs = [self._decorate_run(run) for run in recent_runs]
        run_ids = [int(run["id"]) for run in recent_runs if run.get("id") is not None]
        query_rows = self.storage.list_run_queries(run_ids)
        queries_by_run: dict[int, list[dict[str, Any]]] = {}
        for row in query_rows:
            queries_by_run.setdefault(int(row["run_id"]), []).append(dict(row))
        query_ids = [int(row["id"]) for row in query_rows if row.get("id") is not None]
        hit_rows = self.storage.list_run_query_hits(query_ids)
        hits_by_query: dict[int, list[dict[str, Any]]] = {}
        world_ids: set[str] = set()
        for row in hit_rows:
            query_id = int(row["run_query_id"])
            hits_by_query.setdefault(query_id, []).append(dict(row))
            world_id = str(row.get("world_id") or "").strip()
            if world_id:
                world_ids.add(world_id)
        topic_rows = self.storage.list_topic_memberships_for_worlds(world_ids)
        topics_by_world: dict[str, list[dict[str, Any]]] = {}
        for row in topic_rows:
            world_id = str(row.get("world_id") or "").strip()
            if not world_id:
                continue
            topics_by_world.setdefault(world_id, []).append(
                {
                    "topic_key": row.get("topic_key"),
                    "label": row.get("topic_label") or row.get("topic_key"),
                    "is_active": bool(row.get("is_active", 1)),
                }
            )

        run_items: list[dict[str, Any]] = []
        total_queries = 0
        total_hits = 0
        total_new_worlds = 0
        for run in decorated_runs:
            run_id = int(run["id"])
            query_items: list[dict[str, Any]] = []
            for query in queries_by_run.get(run_id, []):
                query_id = int(query["id"])
                hits = hits_by_query.get(query_id, [])
                topic_counter: Counter[str] = Counter()
                topic_meta: dict[str, dict[str, Any]] = {}
                sample_hits: list[dict[str, Any]] = []
                for hit in hits:
                    world_topics = topics_by_world.get(str(hit.get("world_id") or ""), [])
                    if not world_topics:
                        topic_counter["unmatched"] += 1
                        topic_meta.setdefault(
                            "unmatched",
                            {"topic_key": "unmatched", "label": "Unmatched", "is_active": False},
                        )
                    else:
                        seen_topic_keys: set[str] = set()
                        for topic in world_topics:
                            topic_key = str(topic.get("topic_key") or "")
                            if not topic_key or topic_key in seen_topic_keys:
                                continue
                            seen_topic_keys.add(topic_key)
                            topic_counter[topic_key] += 1
                            topic_meta[topic_key] = topic
                    if len(sample_hits) < 6:
                        sample_hits.append(
                            {
                                "world_id": hit.get("world_id"),
                                "world_name": hit.get("world_name"),
                                "author_id": hit.get("author_id"),
                                "rank_index": hit.get("rank_index"),
                                "is_new_global": bool(hit.get("is_new_global")),
                                "topics": world_topics,
                            }
                        )
                top_topics = [
                    {
                        "topic_key": topic_key,
                        "label": topic_meta.get(topic_key, {}).get("label") or topic_key,
                        "is_active": bool(topic_meta.get(topic_key, {}).get("is_active", False)),
                        "count": count,
                    }
                    for topic_key, count in topic_counter.most_common(5)
                ]
                query_items.append(
                    {
                        "id": query_id,
                        "query_index": query.get("query_index", 0),
                        "query_kind": query.get("query_kind"),
                        "query_value": query.get("query_value"),
                        "query_label": query.get("query_label") or query.get("query_value"),
                        "query_payload": query.get("query_payload", {}),
                        "result_count": query.get("result_count", 0),
                        "kept_count": query.get("kept_count", 0),
                        "new_world_count": query.get("new_world_count", 0),
                        "tracked_hit_count": len(hits),
                        "top_topics": top_topics,
                        "sample_hits": sample_hits,
                    }
                )
                total_queries += 1
                total_hits += int(query.get("kept_count", 0) or 0)
                total_new_worlds += int(query.get("new_world_count", 0) or 0)
            if not query_items:
                query_items = self._infer_run_queries_for_display(run)
                if query_items:
                    total_queries += len(query_items)
            run_items.append(
                {
                    **run,
                    "tracked_query_count": len(query_items),
                    "tracking_status": "tracked" if queries_by_run.get(run_id) else "legacy_inferred" if query_items else "legacy",
                    "queries": query_items,
                }
            )

        return {
            "summary": {
                "run_count": len(run_items),
                "query_count": total_queries,
                "tracked_world_hits": total_hits,
                "new_world_hits": total_new_worlds,
            },
            "items": run_items,
        }

    def _infer_run_queries_for_display(self, run: dict[str, Any]) -> list[dict[str, Any]]:
        source_key = str(run.get("source_key") or "")
        job_key = str(run.get("job_key") or "").strip()
        inferred: list[dict[str, Any]] = []
        if job_key:
            config = self._load_job_configs().get(job_key)
            if config:
                resolved = self._resolve_job_config(job_key, config)
                if resolved["type"] == "keywords":
                    for keyword in resolved.get("keywords", []):
                        inferred.append(self._legacy_query_display_item("keyword", keyword))
                    for user_id in resolved.get("include_user_ids", []):
                        inferred.append(self._legacy_query_display_item("user", user_id))
                elif resolved["type"] == "user" and resolved.get("user_id"):
                    inferred.append(self._legacy_query_display_item("user", resolved["user_id"]))
                elif resolved["type"] == "world_search":
                    inferred.append(self._legacy_query_display_item("world_search", resolved.get("label") or job_key))
        elif source_key.startswith("manual:keyword:"):
            inferred.append(self._legacy_query_display_item("keyword", run.get("query_label") or source_key.removeprefix("manual:keyword:")))
        elif source_key.startswith("manual:user:"):
            inferred.append(self._legacy_query_display_item("user", run.get("query_label") or source_key.removeprefix("manual:user:")))
        elif source_key.startswith("manual:fixed:"):
            for item in self._csv_items(run.get("query_label")):
                inferred.append(self._legacy_query_display_item("keyword", item))
        elif source_key.startswith("manual:world_search:"):
            inferred.append(self._legacy_query_display_item("world_search", run.get("query_label") or source_key.removeprefix("manual:world_search:")))
        return inferred

    def _legacy_query_display_item(self, query_kind: str, query_value: Any) -> dict[str, Any]:
        label = str(query_value or "").strip()
        return {
            "id": None,
            "query_index": 0,
            "query_kind": query_kind,
            "query_value": label,
            "query_label": label,
            "query_payload": {},
            "result_count": 0,
            "kept_count": 0,
            "new_world_count": 0,
            "tracked_hit_count": 0,
            "top_topics": [],
            "sample_hits": [],
            "legacy_inferred": True,
        }

    def get_dashboard(self, job_key: str) -> dict[str, Any]:
        source = f"db:job:{job_key}"
        last_run = self.storage.get_latest_run_for_job(job_key)
        worlds = self.load_worlds(source, sort="visits", direction="desc")
        payload = self._build_scope_dashboard_payload(
            label=source,
            worlds=worlds,
            last_run=self._decorate_run(last_run) if last_run else None,
            history_source=source,
        )
        payload["job_key"] = job_key
        return payload

    def get_scope_dashboard(self, *, source: str | None = None, topic_key: str | None = None) -> dict[str, Any]:
        if topic_key:
            topic = self.storage.get_topic(topic_key)
            if topic is None:
                raise KeyError(f"Unknown topic: {topic_key}")
            worlds = self.load_topic_worlds(topic_key, sort="visits", direction="desc")
            payload = self._build_scope_dashboard_payload(
                label=topic["label"],
                worlds=worlds,
                history_entries=self._load_topic_history(topic_key),
            )
            payload["topic_key"] = topic_key
            return payload

        resolved_source = source or "db:all"
        worlds = self.load_worlds(resolved_source, sort="visits", direction="desc")
        last_run = None
        if resolved_source.startswith("db:job:"):
            job_key = resolved_source.removeprefix("db:job:")
            latest = self.storage.get_latest_run_for_job(job_key)
            last_run = self._decorate_run(latest) if latest else None
        payload = self._build_scope_dashboard_payload(
            label=resolved_source,
            worlds=worlds,
            last_run=last_run,
            history_source=resolved_source,
        )
        payload["source"] = resolved_source
        return payload

    def _build_scope_dashboard_payload(
        self,
        *,
        label: str,
        worlds: list[dict[str, Any]],
        last_run: dict[str, Any] | None = None,
        history_source: str | None = None,
        history_entries: dict[str, list[dict[str, Any]]] | None = None,
    ) -> dict[str, Any]:
        total_visits = sum(self._to_int(world.get("visits")) for world in worlds)
        total_favorites = sum(self._to_int(world.get("favorites")) for world in worlds)
        latest_fetched_at = max((str(world.get("fetched_at") or "") for world in worlds), default="") or None
        history_all = history_entries if history_entries is not None else self.load_history(source=history_source)
        changes = []
        for world in worlds:
            world_id = world.get("id")
            if not world_id:
                continue
            points = history_all.get(world_id, [])
            if len(points) < 2:
                continue
            earliest = next((point for point in points if point.get("visits") is not None), None)
            latest = next((point for point in reversed(points) if point.get("visits") is not None), None)
            if not earliest or not latest or earliest is latest:
                continue
            changes.append(
                {
                    "id": world_id,
                    "name": world.get("name"),
                    "delta": self._to_int(latest.get("visits")) - self._to_int(earliest.get("visits")),
                    "visits": self._to_int(world.get("visits")),
                    "thumbnail_url": world.get("thumbnail_url"),
                }
            )
        changes.sort(key=lambda item: (item["delta"], item["visits"]), reverse=True)
        return {
            "label": label,
            "worlds": worlds[:8],
            "stats": {
                "world_count": len(worlds),
                "total_visits": total_visits,
                "total_favorites": total_favorites,
                "avg_visits": round(total_visits / len(worlds)) if worlds else 0,
            },
            "last_run": last_run,
            "last_sync_at": (last_run or {}).get("finished_at") or latest_fetched_at,
            "top_movers": changes[:5],
        }

    def load_worlds(
        self,
        source: str,
        *,
        query: str | None = None,
        tag: str | None = None,
        sort: str = "visits",
        direction: str = "desc",
        dedupe: bool = True,
    ) -> list[dict[str, Any]]:
        if source == "db:all" or source.startswith("db:"):
            source_key = None if source == "db:all" else source.removeprefix("db:")
            worlds = self.storage.load_latest_worlds(source_key)
            worlds = [self._normalise_db_world(world) for world in worlds]
            if source_key and source_key.startswith("job:"):
                job_key = source_key.removeprefix("job:")
                resolved = self._resolve_job_for_display(job_key)
                whitelist = resolved["whitelist"]
                exclude_ids = resolved["exclude_ids"]
                name_bl = resolved["name_bl"]
                if whitelist or exclude_ids or name_bl:
                    worlds = [
                        w for w in worlds
                        if not self._should_exclude_world(
                            w,
                            exclude_author_ids=exclude_ids,
                            whitelist_author_ids=whitelist,
                            name_blacklist=name_bl,
                        )
                    ]
        else:
            config = self.legacy_sources.get(source)
            if config is None:
                raise KeyError(f"Unknown source: {source}")
            path = Path(config["path"])
            if config["kind"] == "xlsx":
                worlds = self._load_workbook_worlds(path, source)
            else:
                raw_items = self._read_json(path, default=[])
                worlds = [self._normalise_api_world(item, source) for item in raw_items if isinstance(item, dict)]

        if dedupe:
            worlds = self._dedupe_worlds(worlds)

        world_properties = self._load_world_properties()
        worlds = [self._apply_world_properties(world, properties=world_properties) for world in worlds]

        if query:
            needle = query.casefold()
            worlds = [
                world
                for world in worlds
                if needle in (world.get("name") or "").casefold()
                or needle in (world.get("id") or "").casefold()
                or needle in (world.get("author_name") or "").casefold()
            ]

        if tag and tag != "all":
            worlds = [world for world in worlds if tag in world.get("tags", [])]

        history = self.load_history(source=source)
        return self._sort_worlds(worlds, sort=sort, direction=direction, history=history)

    def load_history(
        self,
        world_id: str | None = None,
        source: str | None = None,
    ) -> dict[str, list[dict[str, Any]]]:
        merged: dict[str, list[dict[str, Any]]] = {}

        source_key = None
        include_legacy_history = True
        if source and source.startswith("db:") and source != "db:all":
            source_key = source.removeprefix("db:")
            include_legacy_history = False

        legacy_history = self._read_json(self.legacy_scraper_dir / "history.json", default={})
        if include_legacy_history and isinstance(legacy_history, dict):
            for wid, entries in legacy_history.items():
                if world_id and wid != world_id:
                    continue
                if not isinstance(entries, list):
                    continue
                bucket = merged.setdefault(wid, [])
                for entry in entries:
                    if isinstance(entry, dict):
                        bucket.append(self._normalise_history_entry(wid, entry, "legacy"))

        for wid, entries in self.storage.load_history_points(world_id, source_key=source_key).items():
            bucket = merged.setdefault(wid, [])
            for entry in entries:
                bucket.append(self._normalise_db_history_entry(wid, entry))

        for wid, entries in merged.items():
            deduped: dict[tuple[Any, ...], dict[str, Any]] = {}
            for entry in entries:
                key = (
                    entry.get("timestamp"),
                    entry.get("visits"),
                    entry.get("favorites"),
                    entry.get("heat"),
                    entry.get("popularity"),
                )
                deduped[key] = entry
            merged[wid] = sorted(deduped.values(), key=lambda item: item.get("timestamp") or 0)
        return merged

    def load_history_summary(self, source: str | None = None) -> list[dict[str, Any]]:
        history = self.load_history(source=source)
        items: list[dict[str, Any]] = []
        for wid, entries in history.items():
            if not entries:
                continue
            latest = entries[-1]
            items.append(
                {
                    "world_id": wid,
                    "name": latest.get("name") or wid,
                    "points": len(entries),
                    "latest": latest,
                }
            )
        items.sort(key=lambda item: item["latest"].get("timestamp") or 0, reverse=True)
        return items

    def _load_history_for_sources(self, sources: list[str]) -> dict[str, list[dict[str, Any]]]:
        merged: dict[str, list[dict[str, Any]]] = {}
        seen_sources: set[str] = set()
        for source in sources:
            cleaned = self._clean_optional_text(source)
            if not cleaned:
                continue
            normalized = cleaned if cleaned.startswith("db:") else self._public_db_source_key(cleaned)
            if normalized in seen_sources:
                continue
            seen_sources.add(normalized)
            history = self.load_history(source=normalized)
            for world_id, entries in history.items():
                merged.setdefault(world_id, []).extend(entries)

        for world_id, entries in merged.items():
            deduped: dict[tuple[Any, ...], dict[str, Any]] = {}
            for entry in entries:
                key = (
                    entry.get("timestamp"),
                    entry.get("visits"),
                    entry.get("favorites"),
                    entry.get("heat"),
                    entry.get("popularity"),
                )
                deduped[key] = entry
            merged[world_id] = sorted(deduped.values(), key=lambda item: item.get("timestamp") or 0)
        return merged

    def _load_topic_history(self, topic_key: str) -> dict[str, list[dict[str, Any]]]:
        rules = [rule for rule in self.storage.list_topic_rules(topic_key) if rule.get("is_active", 1)]
        source_rules = [rule for rule in rules if rule.get("rule_type") == "source"]
        other_rules = [rule for rule in rules if rule.get("rule_type") != "source"]
        if source_rules and not other_rules:
            return self._load_history_for_sources([rule.get("rule_value") or "" for rule in source_rules])
        return self.load_history()

    def load_daily_stats(self) -> list[dict[str, Any]]:
        stats: list[dict[str, Any]] = []

        for path in sorted(self.legacy_analytics_dir.glob("daily_stats_*.xlsx")):
            rows = self._read_daily_stats_xlsx(path)
            stats.append(
                {
                    "source": path.stem.removeprefix("daily_stats_"),
                    "origin": "legacy",
                    "path": self._display_path(path),
                    "rows": rows,
                    "latest": rows[-1] if rows else None,
                }
            )

        grouped_db_stats: dict[str, list[dict[str, Any]]] = {}
        for row in self.storage.list_daily_stats():
            grouped_db_stats.setdefault(row["source_key"], []).append(
                {
                    "date": row["date"],
                    "total_worlds": self._to_int(row["total_worlds"]),
                    "new_worlds_today": self._to_int(row["new_worlds_today"]),
                }
            )

        for source_key, rows in grouped_db_stats.items():
            rows.sort(key=lambda item: item["date"])
            stats.append(
                {
                    "source": self._public_db_source_key(source_key),
                    "origin": "db",
                    "path": self._display_path(self.storage.db_path),
                    "rows": rows,
                    "latest": rows[-1] if rows else None,
                }
            )

        stats.sort(
            key=lambda item: ((item.get("latest") or {}).get("date") or "", item["source"]),
            reverse=True,
        )
        return stats

    def load_collection_insights(
        self,
        *,
        source: str | None = None,
        topic_key: str | None = None,
        limit: int = 12,
        allow_cache: bool = True,
    ) -> dict[str, Any]:
        resolved_source: str | None = None
        effective_limit = limit
        if topic_key:
            topic = self.get_topic(topic_key)
            worlds = self.load_topic_worlds(topic_key)
            label = topic["label"]
            scope = {"topic_key": topic_key, "label": label}
            history = self._load_topic_history(topic_key)
        else:
            resolved_source = source or "db:all"
            effective_limit = max(limit, ANALYSIS_CACHE_LIMIT)
            storage = getattr(self, "storage", None)
            if allow_cache and storage is not None:
                cached = storage.get_analysis_cache(resolved_source)
                if cached and isinstance(cached.get("payload"), dict):
                    return self._limit_collection_insights_payload(cached["payload"], limit)
            worlds = self.load_worlds(resolved_source)
            label = resolved_source
            scope = {"source": resolved_source, "label": label}
            history = self.load_history(source=resolved_source)
        insights_by_id: dict[str, dict[str, Any]] = {}
        author_buckets: dict[str, dict[str, Any]] = {}
        growth_rows: list[dict[str, Any]] = []
        rising_now_rows: list[dict[str, Any]] = []
        new_hot_rows: list[dict[str, Any]] = []
        worth_watching_rows: list[dict[str, Any]] = []
        dormant_revival_rows: list[dict[str, Any]] = []
        update_effect_rows: list[dict[str, Any]] = []
        anomaly_rows: list[dict[str, Any]] = []
        signal_rows: list[dict[str, Any]] = []
        summary_total_visits = 0
        summary_total_favorites = 0
        summary_creators: set[str] = set()
        summary_last_seen_at: str | None = None

        for world in worlds:
            world_id = world.get("id")
            if not world_id:
                continue
            summary_total_visits += self._to_int(world.get("visits"))
            summary_total_favorites += self._to_int(world.get("favorites"))
            creator_key = (world.get("author_id") or world.get("author_name") or "").strip()
            if creator_key:
                summary_creators.add(creator_key)
            fetched_at = self._clean_optional_text(world.get("fetched_at"))
            if fetched_at and (summary_last_seen_at is None or fetched_at > summary_last_seen_at):
                summary_last_seen_at = fetched_at
            trend = self._build_world_trend_metrics(world, history.get(world_id, []))
            insights_by_id[world_id] = trend
            row_base = {
                "id": world_id,
                "name": world.get("name") or world_id,
                "author_name": world.get("author_name"),
                "author_id": world.get("author_id"),
                "visits": self._to_optional_int(world.get("visits")),
                "favorites": self._to_optional_int(world.get("favorites")),
                "heat": self._to_optional_int(world.get("heat")),
                "popularity": self._to_optional_int(world.get("popularity")),
                "days_since_publication": trend.get("days_since_publication"),
                "days_since_update": trend.get("days_since_update"),
                "visits_delta_1d": trend.get("visits_delta_1d"),
                "visits_delta_7d": trend.get("visits_delta_7d"),
                "visits_delta_prev_7d": trend.get("visits_delta_prev_7d"),
                "visits_delta_30d": trend.get("visits_delta_30d"),
                "visits_growth_1d": trend.get("visits_growth_1d"),
                "visits_growth_7d": trend.get("visits_growth_7d"),
                "favorites_delta_1d": trend.get("favorites_delta_1d"),
                "favorites_delta_7d": trend.get("favorites_delta_7d"),
                "favorite_rate": trend.get("favorite_rate"),
                "publication_visits_per_day": trend.get("publication_visits_per_day"),
                "since_update_visits_per_day": trend.get("since_update_visits_per_day"),
                "tags": trend.get("tags", []),
                "thumbnail_url": world.get("thumbnail_url"),
                "world_url": world.get("world_url"),
            }
            growth_rows.append(
                {
                    **row_base,
                    "momentum_score": trend.get("momentum_score"),
                    "discovery_reason": self._discovery_reason(trend, mode="momentum"),
                    "score": self._trend_score(trend),
                }
            )
            rising_now_score = self._rising_now_score(trend)
            rising_now_rows.append(
                {
                    **row_base,
                    "rising_now_score": rising_now_score,
                    "discovery_reason": self._discovery_reason(trend, mode="rising_now"),
                    "score": rising_now_score,
                }
            )
            new_hot_rows.append(
                {
                    **row_base,
                    "breakout_score": trend.get("breakout_score"),
                    "new_hot_score": trend.get("new_hot_score"),
                    "discovery_reason": self._discovery_reason(trend, mode="new_hot"),
                    "score": self._trend_sort_metric(world, trend, "new_hot"),
                }
            )
            worth_watching_rows.append(
                {
                    **row_base,
                    "worth_watching_score": trend.get("worth_watching_score"),
                    "discovery_reason": self._discovery_reason(trend, mode="worth_watching"),
                    "score": self._trend_sort_metric(world, trend, "worth_watching"),
                }
            )
            dormant_revival_score = self._dormant_revival_score(trend)
            if dormant_revival_score > 0:
                dormant_revival_rows.append(
                    {
                        **row_base,
                        "update_gap_days": trend.get("update_gap_days"),
                        "update_effect_tag": trend.get("update_effect_tag"),
                        "since_update_visits_delta": trend.get("since_update_visits_delta"),
                        "since_update_visits_per_day": trend.get("since_update_visits_per_day"),
                        "dormant_revival_score": dormant_revival_score,
                        "discovery_reason": self._discovery_reason(trend, mode="revival"),
                        "score": dormant_revival_score,
                    }
                )
            signal_rows.append(
                {
                    "id": world_id,
                    "name": world.get("name") or world_id,
                    "author_name": world.get("author_name"),
                    "author_id": world.get("author_id"),
                    "visits": self._to_optional_int(world.get("visits")),
                    "favorites": self._to_optional_int(world.get("favorites")),
                    "heat": self._to_optional_int(world.get("heat")),
                    "popularity": self._to_optional_int(world.get("popularity")),
                    "favorite_rate": trend.get("favorite_rate"),
                    "visits_delta_30d": trend.get("visits_delta_30d"),
                    "favorites_delta_7d": trend.get("favorites_delta_7d"),
                    "days_since_publication": trend.get("days_since_publication"),
                    "days_since_update": trend.get("days_since_update"),
                    "world_url": world.get("world_url"),
                }
            )
            author_key = world.get("author_id") or world.get("author_name") or "__unknown__"
            bucket = author_buckets.setdefault(
                author_key,
                {
                    "author_id": world.get("author_id"),
                    "author_name": world.get("author_name"),
                    "world_count": 0,
                    "active_worlds_30d": 0,
                    "total_visits": 0,
                    "total_favorites": 0,
                    "recent_visits_delta_30d": 0,
                    "recent_visits_delta_7d": 0,
                    "favorite_rate_sum": 0.0,
                    "favorite_rate_count": 0,
                    "top_world_name": None,
                    "top_world_visits": -1,
                    "breakout_worlds": 0,
                    "rising_worlds": 0,
                    "worth_watching_worlds": 0,
                    "lead_world_name": None,
                    "lead_world_score": -1.0,
                },
            )
            bucket["world_count"] += 1
            bucket["total_visits"] += self._to_int(world.get("visits"))
            bucket["total_favorites"] += self._to_int(world.get("favorites"))
            bucket["recent_visits_delta_30d"] += self._to_int(trend.get("visits_delta_30d"))
            bucket["recent_visits_delta_7d"] += self._to_int(trend.get("visits_delta_7d"))
            if self._to_int(trend.get("visits_delta_30d")) > 0:
                bucket["active_worlds_30d"] += 1
            if trend.get("favorite_rate") is not None:
                bucket["favorite_rate_sum"] += float(trend["favorite_rate"])
                bucket["favorite_rate_count"] += 1
            if (self._to_float(trend.get("new_hot_score")) or 0.0) >= 180:
                bucket["breakout_worlds"] += 1
            if rising_now_score >= 140:
                bucket["rising_worlds"] += 1
            if (self._to_float(trend.get("worth_watching_score")) or 0.0) >= 180:
                bucket["worth_watching_worlds"] += 1
            if self._to_int(world.get("visits")) > bucket["top_world_visits"]:
                bucket["top_world_visits"] = self._to_int(world.get("visits"))
                bucket["top_world_name"] = world.get("name") or world_id
            lead_world_score = max(
                self._to_float(trend.get("momentum_score")) or 0.0,
                self._to_float(trend.get("new_hot_score")) or 0.0,
                rising_now_score,
            )
            if lead_world_score > bucket["lead_world_score"]:
                bucket["lead_world_score"] = lead_world_score
                bucket["lead_world_name"] = world.get("name") or world_id

            days_since_update = trend.get("days_since_update")
            if (
                trend.get("since_update_visits_delta") is not None
                and isinstance(days_since_update, int)
                and 0 <= days_since_update <= 120
            ):
                update_effect_rows.append(
                    {
                        "id": world_id,
                        "name": world.get("name") or world_id,
                        "author_name": world.get("author_name"),
                        "updated_at": world.get("updated_at"),
                        "days_since_update": trend.get("days_since_update"),
                        "visits_delta_1d": trend.get("visits_delta_1d"),
                        "favorites_delta_1d": trend.get("favorites_delta_1d"),
                        "visits_delta_7d": trend.get("visits_delta_7d"),
                        "since_update_visits_delta": trend.get("since_update_visits_delta"),
                        "since_update_favorites_delta": trend.get("since_update_favorites_delta"),
                        "since_update_visits_per_day": trend.get("since_update_visits_per_day"),
                        "since_update_favorites_per_day": trend.get("since_update_favorites_per_day"),
                        "update_gap_days": trend.get("update_gap_days"),
                        "update_effect_tag": trend.get("update_effect_tag"),
                        "update_effectiveness_score": self._update_effectiveness_score(trend),
                        "world_url": world.get("world_url"),
                    }
                )
            if self._is_notable_anomaly(trend):
                anomaly_rows.append(
                    {
                        "id": world_id,
                        "name": world.get("name") or world_id,
                        "author_name": world.get("author_name"),
                        "visits": self._to_optional_int(world.get("visits")),
                        "favorites": self._to_optional_int(world.get("favorites")),
                        "visits_delta_1d": trend.get("visits_delta_1d"),
                        "favorites_delta_1d": trend.get("favorites_delta_1d"),
                        "visits_delta_7d": trend.get("visits_delta_7d"),
                        "visits_delta_prev_7d": trend.get("visits_delta_prev_7d"),
                        "favorites_delta_7d": trend.get("favorites_delta_7d"),
                        "days_since_update": trend.get("days_since_update"),
                        "anomaly_ratio": self._anomaly_ratio(trend.get("visits_delta_7d"), trend.get("visits_delta_prev_7d")),
                        "anomaly_score": self._anomaly_score(trend),
                        "world_url": world.get("world_url"),
                    }
                )

        growth_rows.sort(key=lambda item: (item["score"], self._to_int(item["visits_delta_7d"])), reverse=True)
        rising_now_rows.sort(key=lambda item: (float(item["score"] or 0.0), self._to_int(item["visits_delta_1d"])), reverse=True)
        new_hot_rows.sort(key=lambda item: (float(item["score"] or 0.0), self._to_int(item["visits_delta_1d"])), reverse=True)
        worth_watching_rows.sort(
            key=lambda item: (
                float(item["score"] or 0.0),
                self._to_int(item["favorites_delta_7d"]),
                self._to_int(item["visits_delta_7d"]),
            ),
            reverse=True,
        )
        dormant_revival_rows.sort(
            key=lambda item: (
                float(item["score"] or 0.0),
                self._to_int(item["visits_delta_7d"]),
                self._to_int(item["since_update_visits_delta"]),
            ),
            reverse=True,
        )
        for row in growth_rows:
            row.pop("score", None)
        for rows in (rising_now_rows, new_hot_rows, worth_watching_rows, dormant_revival_rows):
            for row in rows:
                row.pop("score", None)

        authors = sorted(
            (
                {
                    **bucket,
                    "average_favorite_rate": round(bucket["favorite_rate_sum"] / bucket["favorite_rate_count"], 2)
                    if bucket["favorite_rate_count"]
                    else None,
                    "top_world_share": round((bucket["top_world_visits"] / bucket["total_visits"]) * 100, 2)
                    if bucket["total_visits"] > 0 and bucket["top_world_visits"] >= 0
                    else None,
                }
                for bucket in author_buckets.values()
            ),
            key=lambda item: (item["recent_visits_delta_30d"], item["total_visits"]),
            reverse=True,
        )
        for item in authors:
            item.pop("favorite_rate_sum", None)
            item.pop("favorite_rate_count", None)
            item.pop("top_world_visits", None)

        creator_momentum = sorted(
            (
                {
                    **bucket,
                    "average_favorite_rate": round(bucket["favorite_rate_sum"] / bucket["favorite_rate_count"], 2)
                    if bucket["favorite_rate_count"]
                    else None,
                    "top_world_share": round((bucket["top_world_visits"] / bucket["total_visits"]) * 100, 2)
                    if bucket["total_visits"] > 0 and bucket["top_world_visits"] >= 0
                    else None,
                    "creator_momentum_score": self._creator_momentum_score(bucket),
                }
                for bucket in author_buckets.values()
            ),
            key=lambda item: (float(item["creator_momentum_score"] or 0.0), item["recent_visits_delta_7d"], item["recent_visits_delta_30d"]),
            reverse=True,
        )
        for item in creator_momentum:
            item.pop("favorite_rate_sum", None)
            item.pop("favorite_rate_count", None)
            item.pop("top_world_visits", None)
            item.pop("lead_world_score", None)

        update_effect_rows.sort(
            key=lambda item: (
                float(item.get("update_effectiveness_score") or 0.0),
                self._to_int(item.get("since_update_visits_delta")),
            ),
            reverse=True,
        )
        anomaly_rows.sort(
            key=lambda item: (
                float(item.get("anomaly_score") or 0.0),
                self._to_int(item.get("visits_delta_7d")),
            ),
            reverse=True,
        )

        update_summary = {
            "tracked_recent_updates": len(update_effect_rows),
            "active_updates": sum(1 for item in update_effect_rows if item.get("update_effect_tag") == "ACTIVE"),
            "steady_updates": sum(1 for item in update_effect_rows if item.get("update_effect_tag") == "STEADY FLOW"),
            "silent_updates": sum(1 for item in update_effect_rows if item.get("update_effect_tag") == "SILENCE UPDATE"),
            "avg_visits_per_day": self._average([item.get("since_update_visits_per_day") for item in update_effect_rows]),
        }
        anomaly_summary = {
            "tracked_anomalies": len(anomaly_rows),
            "high_ratio": sum(1 for item in anomaly_rows if (item.get("anomaly_ratio") or 0) >= 2.0),
            "strong_1d": sum(1 for item in anomaly_rows if self._to_int(item.get("visits_delta_1d")) >= 30),
            "avg_ratio": self._average([item.get("anomaly_ratio") for item in anomaly_rows]),
        }
        summary = {
            "world_count": len(worlds),
            "total_visits": summary_total_visits,
            "total_favorites": summary_total_favorites,
            "tracked_creators": len(summary_creators),
            "last_seen_at": summary_last_seen_at,
            "new_worlds_14d": sum(
                1
                for trend in insights_by_id.values()
                if isinstance(trend.get("days_since_publication"), int)
                and 0 <= int(trend["days_since_publication"]) <= 14
            ),
            "updated_worlds_30d": sum(
                1
                for trend in insights_by_id.values()
                if isinstance(trend.get("days_since_update"), int)
                and 0 <= int(trend["days_since_update"]) <= 30
            ),
        }
        briefing = {
            "momentum": growth_rows[: min(effective_limit, 6)],
            "rising_now": rising_now_rows[: min(effective_limit, 6)],
            "new_worlds": new_hot_rows[: min(effective_limit, 6)],
            "worth_watching": worth_watching_rows[: min(effective_limit, 6)],
        }

        payload = {
            **scope,
            "generated_at": dt.datetime.now(dt.timezone.utc).isoformat(),
            "summary": summary,
            "briefing": briefing,
            "world_insights": insights_by_id,
            "growth_leaderboard": growth_rows[:effective_limit],
            "rising_now_leaderboard": rising_now_rows[:effective_limit],
            "new_hot_leaderboard": new_hot_rows[:effective_limit],
            "worth_watching_leaderboard": worth_watching_rows[:effective_limit],
            "dormant_revival_leaderboard": dormant_revival_rows[:effective_limit],
            "authors": authors[:effective_limit],
            "creator_momentum": creator_momentum[:effective_limit],
            "anomalies": {
                "items": anomaly_rows[:effective_limit],
                "summary": anomaly_summary,
            },
            "update_effectiveness": {
                "enabled": bool(update_effect_rows),
                "items": update_effect_rows[:effective_limit],
                "summary": update_summary,
            },
            "signals": self._build_signal_analysis(signal_rows, limit=effective_limit),
            "performance": {
                "enabled": bool(update_effect_rows),
                "items": update_effect_rows[:effective_limit],
                "summary": update_summary,
            },
        }
        storage = getattr(self, "storage", None)
        if resolved_source and storage is not None:
            storage.upsert_analysis_cache(
                scope_key=resolved_source,
                scope_type="source",
                updated_at=payload["generated_at"],
                payload=payload,
            )
        return self._limit_collection_insights_payload(payload, limit)

    def list_event_feed(self, *, limit: int = 50, recency_days: int = 7) -> dict[str, Any]:
        now = dt.datetime.now(dt.timezone.utc)
        events: list[dict[str, Any]] = []
        seen_keys: set[tuple[str, str, str]] = set()
        worlds_cache: dict[str, dict[str, dict[str, Any]]] = {}

        def load_world_map(source_key: str) -> dict[str, dict[str, Any]]:
            cached = worlds_cache.get(source_key)
            if cached is not None:
                return cached
            try:
                worlds = self.load_worlds(source_key)
            except KeyError:
                worlds = []
            mapping = {
                str(world.get("id")): world
                for world in worlds
                if world.get("id")
            }
            worlds_cache[source_key] = mapping
            return mapping

        def append_event(
            *,
            event_type: str,
            job: dict[str, Any],
            world: dict[str, Any] | None,
            occurred_at: str | None,
            summary: str,
            severity: float,
            delta: dict[str, Any] | None = None,
        ) -> None:
            world_id = str((world or {}).get("id") or "")
            event_key = (job["job_key"], event_type, world_id)
            if event_key in seen_keys:
                return
            seen_keys.add(event_key)
            payload = self._event_world_preview(world or {})
            payload.update(
                {
                    "event_key": f"{job['job_key']}:{event_type}:{world_id or 'job'}",
                    "type": event_type,
                    "job_key": job["job_key"],
                    "label": job["label"],
                    "source": job["source"],
                    "topic_key": job["job_key"] if self.storage.get_topic(job["job_key"]) else None,
                    "occurred_at": occurred_at,
                    "detected_at": (job.get("latest_run") or {}).get("finished_at") or (job.get("latest_run") or {}).get("started_at"),
                    "summary": summary,
                    "severity": round(float(severity), 2),
                    "delta": delta or {},
                }
            )
            events.append(payload)

        jobs = self.list_jobs()
        for job in jobs:
            source = job["source"]
            world_map = load_world_map(source)
            diff = self.get_job_source_diff(job["job_key"], added_limit=20, removed_limit=0, changed_limit=20)
            latest_run = diff.get("latest_run") or job.get("latest_run")

            for added in diff.get("added_worlds", []):
                current_world = world_map.get(str(added.get("id"))) or added
                occurred_at = self._new_world_event_date(current_world) or ((latest_run or {}).get("finished_at"))
                append_event(
                    event_type="new_upload",
                    job=job,
                    world=current_world,
                    occurred_at=occurred_at,
                    summary=(
                        f"New in {job['label']} "
                        f"/ visits {self._to_int(current_world.get('visits')):,} "
                        f"/ favorites {self._to_int(current_world.get('favorites')):,}"
                    ),
                    severity=80 + min(self._to_int(current_world.get("visits")) / 25, 20),
                    delta={
                        "visits": self._to_int(current_world.get("visits")),
                        "favorites": self._to_int(current_world.get("favorites")),
                    },
                )

            for changed in diff.get("changed_worlds", []):
                current_world = world_map.get(str(changed.get("id"))) or changed.get("latest") or {}
                spike = self._compute_diff_spike(changed)
                if self._is_notable_diff_spike(changed, spike):
                    append_event(
                        event_type="traffic_spike",
                        job=job,
                        world=current_world,
                        occurred_at=(latest_run or {}).get("finished_at"),
                        summary=(
                            f"Visits {self._signed_number(changed.get('visits_delta'))}"
                            f" / fav {self._signed_number(changed.get('favorites_delta'))}"
                            f" / heat {self._signed_number(changed.get('heat_delta'))}"
                            f" / pop {self._signed_number(changed.get('popularity_delta'))}"
                        ),
                        severity=spike["score"],
                        delta={
                            "visits_delta": self._to_int(changed.get("visits_delta")),
                            "favorites_delta": self._to_int(changed.get("favorites_delta")),
                            "heat_delta": self._to_int(changed.get("heat_delta")),
                            "popularity_delta": self._to_int(changed.get("popularity_delta")),
                            "visits_growth": spike.get("visits_growth"),
                            "favorites_growth": spike.get("favorites_growth"),
                        },
                    )
                if "updated_at" in (changed.get("changed_fields") or []):
                    update_summary = "World update detected"
                    if self._to_int(changed.get("visits_delta")):
                        update_summary += f" / visits {self._signed_number(changed.get('visits_delta'))}"
                    if self._to_int(changed.get("heat_delta")):
                        update_summary += f" / heat {self._signed_number(changed.get('heat_delta'))}"
                    append_event(
                        event_type="new_update",
                        job=job,
                        world=current_world,
                        occurred_at=current_world.get("updated_at") or (changed.get("latest") or {}).get("updated_at"),
                        summary=update_summary,
                        severity=55 + max(0, self._to_int(changed.get("visits_delta")) / 10),
                        delta={
                            "visits_delta": self._to_int(changed.get("visits_delta")),
                            "favorites_delta": self._to_int(changed.get("favorites_delta")),
                            "heat_delta": self._to_int(changed.get("heat_delta")),
                            "popularity_delta": self._to_int(changed.get("popularity_delta")),
                            "changed_fields": changed.get("changed_fields") or [],
                        },
                    )

            for world in world_map.values():
                world_id = str(world.get("id") or "")
                new_world_event_date = self._new_world_event_date(world)
                if world_id and self._within_days(new_world_event_date, recency_days, now):
                    append_event(
                        event_type="new_upload",
                        job=job,
                        world=world,
                        occurred_at=new_world_event_date,
                        summary=(
                            f"Published recently"
                            f" / visits {self._to_int(world.get('visits')):,}"
                            f" / favorites {self._to_int(world.get('favorites')):,}"
                        ),
                        severity=60 + min(self._to_int(world.get("visits")) / 40, 15),
                        delta={
                            "visits": self._to_int(world.get("visits")),
                            "favorites": self._to_int(world.get("favorites")),
                        },
                    )
                if world_id and self._within_days(world.get("updated_at"), recency_days, now):
                    append_event(
                        event_type="new_update",
                        job=job,
                        world=world,
                        occurred_at=world.get("updated_at"),
                        summary=(
                            f"Updated recently"
                            f" / visits {self._to_int(world.get('visits')):,}"
                            f" / heat {self._to_int(world.get('heat')):,}"
                            f" / popularity {self._to_int(world.get('popularity')):,}"
                        ),
                        severity=45 + min(self._to_int(world.get("heat")) * 3, 18),
                        delta={
                            "visits": self._to_int(world.get("visits")),
                            "favorites": self._to_int(world.get("favorites")),
                            "heat": self._to_int(world.get("heat")),
                            "popularity": self._to_int(world.get("popularity")),
                        },
                    )

        events.sort(
            key=lambda item: (
                self._date_score(item.get("occurred_at") or item.get("detected_at")),
                item.get("severity", 0),
                item.get("label", "").casefold(),
                item.get("name", "").casefold(),
            ),
            reverse=True,
        )
        counts = Counter(item["type"] for item in events)
        return {
            "generated_at": now.isoformat(),
            "recency_days": recency_days,
            "summary": {
                "total": len(events),
                "spikes": counts.get("traffic_spike", 0),
                "uploads": counts.get("new_upload", 0),
                "updates": counts.get("new_update", 0),
            },
            "items": events[:limit],
        }

    def _limit_collection_insights_payload(self, payload: dict[str, Any], limit: int) -> dict[str, Any]:
        if limit <= 0:
            return copy.deepcopy(payload)
        limited = copy.deepcopy(payload)
        for key in (
            "growth_leaderboard",
            "rising_now_leaderboard",
            "new_hot_leaderboard",
            "worth_watching_leaderboard",
            "dormant_revival_leaderboard",
            "authors",
            "creator_momentum",
        ):
            if isinstance(limited.get(key), list):
                limited[key] = limited[key][:limit]
        briefing = limited.get("briefing")
        if isinstance(briefing, dict):
            for key in ("momentum", "rising_now", "new_worlds", "worth_watching"):
                if isinstance(briefing.get(key), list):
                    briefing[key] = briefing[key][:limit]
        for key in ("anomalies", "update_effectiveness", "performance"):
            section = limited.get(key)
            if isinstance(section, dict) and isinstance(section.get("items"), list):
                section["items"] = section["items"][:limit]
        signals = limited.get("signals")
        if isinstance(signals, dict):
            for key in ("correlations", "charts", "leaderboards"):
                if isinstance(signals.get(key), list):
                    signals[key] = signals[key][:limit]
            summary = signals.get("summary")
            if isinstance(summary, dict) and isinstance(summary.get("top_signals"), list):
                summary["top_signals"] = summary["top_signals"][:limit]
        return limited

    def refresh_analysis_cache(self, source: str, *, source_run_id: int | None = None) -> None:
        payload = self.load_collection_insights(
            source=source,
            limit=ANALYSIS_CACHE_LIMIT,
            allow_cache=False,
        )
        self.storage.upsert_analysis_cache(
            scope_key=source,
            scope_type="source",
            updated_at=payload.get("generated_at") or dt.datetime.now(dt.timezone.utc).isoformat(),
            payload=payload,
            source_run_id=source_run_id,
        )

    def search_keyword(
        self,
        *,
        keyword: str,
        limit: int = 50,
        cookie: str | None = None,
        username: str | None = None,
        password: str | None = None,
    ) -> dict[str, Any]:
        source_key = f"manual:keyword:{self._slugify(keyword)}"
        headers = _load_headers(cookie, username, password)
        worlds, warnings, meta = self._prepare_sync_worlds(
            fetch_worlds(keyword=keyword, limit=limit, headers=headers),
            headers=headers,
        )
        return self._store_sync_result(
            source_key=source_key,
            job_key=None,
            trigger_type="manual",
            query_label=keyword,
            worlds=worlds,
            warnings=warnings,
            meta=meta,
            query_batches=[
                self._make_query_batch(
                    kind="keyword",
                    value=keyword,
                    label=keyword,
                    worlds=worlds,
                    payload={"keyword": keyword, "limit": limit},
                )
            ],
        )

    def search_user(
        self,
        *,
        user_id: str,
        limit: int = 50,
        cookie: str | None = None,
        username: str | None = None,
        password: str | None = None,
    ) -> dict[str, Any]:
        source_key = f"manual:user:{self._slugify(user_id)}"
        headers = _load_headers(cookie, username, password)
        worlds, warnings, meta = self._prepare_sync_worlds(
            fetch_worlds(user_id=user_id, limit=limit, headers=headers),
            headers=headers,
        )
        return self._store_sync_result(
            source_key=source_key,
            job_key=None,
            trigger_type="manual",
            query_label=user_id,
            worlds=worlds,
            warnings=warnings,
            meta=meta,
            query_batches=[
                self._make_query_batch(
                    kind="user",
                    value=user_id,
                    label=user_id,
                    worlds=worlds,
                    payload={"user_id": user_id, "limit": limit},
                )
            ],
        )

    def search_worlds(
        self,
        *,
        search: str | None = None,
        tags: list[str] | str | None = None,
        notags: list[str] | str | None = None,
        sort: str = "popularity",
        order: str = "descending",
        featured: bool | str | None = None,
        active: bool | str = False,
        release_status: str | None = None,
        platform: str | None = None,
        limit: int = 50,
        source_name: str | None = None,
        cookie: str | None = None,
        username: str | None = None,
        password: str | None = None,
    ) -> dict[str, Any]:
        headers = _load_headers(cookie, username, password)
        tags_list = self._csv_items(tags)
        notag_list = self._csv_items(notags)
        featured_value = self._optional_bool(featured)
        active_value = bool(self._optional_bool(active) or False)
        release_status_value = self._clean_optional_text(release_status)
        platform_value = self._clean_optional_text(platform)
        label = source_name or self._world_search_label(
            search=search,
            tags=tags_list,
            sort=sort,
            active=active_value,
            featured=featured_value,
        )
        worlds, warnings, meta = self._prepare_sync_worlds(
            search_worlds_query(
                search=search,
                tags=tags_list,
                notags=notag_list,
                sort=sort,
                order=order,
                featured=featured_value,
                active=active_value,
                release_status=release_status_value,
                platform=platform_value,
                limit=limit,
                headers=headers,
            ),
            headers=headers,
        )
        meta.update(
            {
                "search": (search or "").strip(),
                "tags": tags_list,
                "notags": notag_list,
                "sort": sort,
                "order": order,
                "featured": featured_value,
                "active": active_value,
                "release_status": release_status_value,
                "platform": platform_value,
            }
        )
        return self._store_sync_result(
            source_key=f"manual:world_search:{self._slugify(label)}",
            job_key=None,
            trigger_type="manual",
            query_label=label,
            worlds=worlds,
            warnings=warnings,
            meta=meta,
            query_batches=[
                self._make_query_batch(
                    kind="world_search",
                    value=label,
                    label=label,
                    worlds=worlds,
                    payload=meta,
                )
            ],
        )

    def search_fixed_keywords(
        self,
        *,
        keywords: list[str],
        blacklist: set[str] | None = None,
        limit_per_keyword: int = 50,
        source_name: str = "fixed-keywords",
        cookie: str | None = None,
        username: str | None = None,
        password: str | None = None,
    ) -> dict[str, Any]:
        headers = _load_headers(cookie, username, password)
        blacklist = blacklist or set()
        all_worlds: list[dict[str, Any]] = []
        query_batches: list[dict[str, Any]] = []
        for keyword in keywords:
            cleaned = keyword.strip()
            if not cleaned or cleaned in blacklist:
                continue
            batch_worlds = fetch_worlds(keyword=cleaned, limit=limit_per_keyword, headers=headers)
            all_worlds.extend(batch_worlds)
            query_batches.append(
                self._make_query_batch(
                    kind="keyword",
                    value=cleaned,
                    label=cleaned,
                    worlds=batch_worlds,
                    payload={"keyword": cleaned, "limit": limit_per_keyword},
                )
            )
        source_key = f"manual:fixed:{self._slugify(source_name)}"
        worlds, warnings, meta = self._prepare_sync_worlds(all_worlds, headers=headers)
        return self._store_sync_result(
            source_key=source_key,
            job_key=None,
            trigger_type="manual",
            query_label=", ".join(keywords),
            worlds=worlds,
            warnings=warnings,
            meta=meta,
            query_batches=query_batches,
        )

    def run_job(
        self,
        job_key: str,
        *,
        cookie: str | None = None,
        username: str | None = None,
        password: str | None = None,
        trigger_type: str = "job_manual",
    ) -> dict[str, Any]:
        configs = self._load_job_configs()
        if job_key not in configs:
            raise KeyError(f"Unknown job: {job_key}")

        resolved = self._resolve_job_config(job_key, configs[job_key])
        if not resolved["ready"]:
            raise ValueError(resolved["reason"] or f"Job {job_key} is not ready")

        headers = _load_headers(cookie, username, password)
        worlds: list[dict[str, Any]]
        query_batches: list[dict[str, Any]] = []
        if resolved["type"] == "keywords":
            keywords = resolved["keywords"]
            blacklist = self._load_blacklist(resolved.get("blacklist_file"))
            include_user_ids = resolved.get("include_user_ids", [])
            all_worlds = []
            for keyword in keywords:
                batch_worlds = fetch_worlds(keyword=keyword, limit=resolved["limit_per_keyword"], headers=headers)
                all_worlds.extend(batch_worlds)
                query_batches.append(
                    self._make_query_batch(
                        kind="keyword",
                        value=keyword,
                        label=keyword,
                        worlds=batch_worlds,
                        payload={"keyword": keyword, "limit": resolved["limit_per_keyword"]},
                    )
                )
            # Removed: Active search for include_user_ids to reduce API calls
            # Instead, rely on filtering during post-processing
            include_user_ids_set = set(resolved.get("include_user_ids", []))
            exclude_author_ids = resolved.get("exclude_author_ids", set())
            name_blacklist = resolved.get("blacklist_world_name_substrings", [])
            worlds = []
            filtered_batches: list[dict[str, Any]] = []
            for batch in query_batches:
                kept_worlds = []
                for world in batch.get("worlds", []):
                    world_id = world.get("id") or world.get("worldId")
                    if world_id in blacklist:
                        continue
                    if self._should_exclude_world(
                        world,
                        exclude_author_ids=exclude_author_ids,
                        whitelist_author_ids=include_user_ids_set,
                        name_blacklist=name_blacklist,
                    ):
                        continue
                    kept_worlds.append(world)
                    worlds.append(world)
                filtered_batches.append({**batch, "worlds": kept_worlds})
            query_batches = filtered_batches
        elif resolved["type"] == "user":
            worlds = fetch_worlds(
                user_id=resolved["user_id"],
                limit=resolved["limit"],
                headers=headers,
            )
            query_batches = [
                self._make_query_batch(
                    kind="user",
                    value=resolved["user_id"],
                    label=resolved["user_id"],
                    worlds=worlds,
                    payload={"user_id": resolved["user_id"], "limit": resolved["limit"]},
                )
            ]
        elif resolved["type"] == "world_search":
            worlds = search_worlds_query(
                search=resolved.get("search"),
                tags=resolved.get("tags", []),
                notags=resolved.get("notags", []),
                sort=resolved.get("sort", "popularity"),
                order=resolved.get("order", "descending"),
                featured=resolved.get("featured"),
                active=bool(resolved.get("active")),
                release_status=resolved.get("release_status"),
                platform=resolved.get("platform"),
                limit=resolved["limit"],
                headers=headers,
            )
            query_batches = [
                self._make_query_batch(
                    kind="world_search",
                    value=resolved["label"],
                    label=resolved["label"],
                    worlds=worlds,
                    payload={
                        "search": resolved.get("search"),
                        "tags": resolved.get("tags", []),
                        "notags": resolved.get("notags", []),
                        "sort": resolved.get("sort", "popularity"),
                        "order": resolved.get("order", "descending"),
                        "featured": resolved.get("featured"),
                        "active": bool(resolved.get("active")),
                        "release_status": resolved.get("release_status"),
                        "platform": resolved.get("platform"),
                        "limit": resolved["limit"],
                    },
                )
            ]
            include_user_ids = resolved.get("include_user_ids", [])
            for user_id in include_user_ids:
                creator_worlds = fetch_worlds(
                    user_id=user_id,
                    limit=resolved["limit"],
                    headers=headers,
                )
                worlds.extend(creator_worlds)
                query_batches.append(
                    self._make_query_batch(
                        kind="user",
                        value=user_id,
                        label=user_id,
                        worlds=creator_worlds,
                        payload={
                            "user_id": user_id,
                            "limit": resolved["limit"],
                            "source": "world_search_whitelist",
                        },
                    )
                )
        else:
            raise ValueError(f"Unsupported job type: {resolved['type']}")

        worlds, warnings, meta = self._prepare_sync_worlds(worlds, headers=headers)
        result = self._store_sync_result(
            source_key=resolved["source_key"],
            job_key=job_key,
            trigger_type=trigger_type,
            query_label=resolved["label"],
            worlds=worlds,
            warnings=warnings,
            meta=meta,
            query_batches=query_batches,
        )
        result["job_key"] = job_key
        return result

    def check_auth_status(
        self,
        *,
        cookie: str | None = None,
        username: str | None = None,
        password: str | None = None,
    ) -> dict[str, Any]:
        headers = _load_headers(cookie, username, password)
        cookie_value = str(headers.get("Cookie", "")).strip()
        authorization = str(headers.get("Authorization", "")).strip()

        if cookie_value:
            result = vrchat_check_session(cookie_value)
            if result.get("ok"):
                user = result.get("user") or {}
                return {
                    "status": "ok",
                    "mode": "cookie",
                    "message": f"session valid for {user.get('displayName') or user.get('id') or 'VRChat user'}",
                    "user": {
                        "id": user.get("id"),
                        "display_name": user.get("displayName"),
                    },
                }
            return {
                "status": "warning",
                "mode": "cookie",
                "message": result.get("error") or "session check failed",
            }

        if authorization:
            return {
                "status": "warning",
                "mode": "basic",
                "message": "basic auth configured but not session-verified; prefer a VRChat Cookie for stable access",
            }

        return {
            "status": "idle",
            "mode": "none",
            "message": "no auth configured",
        }

    def login_with_vrchat(self, *, username: str, password: str) -> dict[str, Any]:
        username = username.strip()
        if not username or not password:
            raise ValueError("username and password are required")
        result = vrchat_login(username, password)
        if result.get("ok"):
            user = result.get("user") or {}
            return {
                "status": "ok",
                "cookie": result.get("cookie"),
                "user": {
                    "id": user.get("id"),
                    "display_name": user.get("displayName"),
                },
                "message": "login successful",
            }
        if result.get("requires_2fa"):
            methods = [
                self._normalise_2fa_method(method)
                for method in (result.get("methods") or [])
            ]
            methods = [method for method in methods if method]
            return {
                "status": "requires_2fa",
                "methods": methods or ["totp"],
                "auth_cookie": result.get("auth_cookie"),
                "message": "two-factor code required",
            }
        raise ValueError(result.get("error") or "login failed")

    def verify_vrchat_2fa(
        self,
        *,
        code: str,
        method: str,
        auth_cookie: str,
    ) -> dict[str, Any]:
        cleaned_code = code.strip().replace(" ", "")
        method = self._normalise_2fa_method(method)
        if not cleaned_code:
            raise ValueError("code is required")
        if method not in {"totp", "otp", "emailotp"}:
            raise ValueError("method must be one of: totp, otp, emailotp")
        if not auth_cookie.strip():
            raise ValueError("auth_cookie is required")

        result = vrchat_verify_2fa(cleaned_code, method, auth_cookie)
        if not result.get("ok"):
            raise ValueError(result.get("error") or "2FA verification failed")

        cookie = str(result.get("cookie") or "")
        status = self.check_auth_status(cookie=cookie)
        return {
            "status": "ok",
            "cookie": cookie,
            "session": status,
            "message": "2FA verified",
        }

    def persist_server_auth(
        self,
        *,
        cookie: str | None = None,
        username: str | None = None,
        password: str | None = None,
    ) -> dict[str, Any]:
        cookie_value = str(cookie or "").strip()
        username_value = str(username or "").strip()
        password_value = str(password or "")
        if not cookie_value and not (username_value and password_value):
            raise ValueError("cookie or username/password is required")

        headers_path = self.legacy_scraper_dir / "headers.json"
        payload = self._read_json(headers_path, default={})
        if not isinstance(payload, dict):
            payload = {}

        payload.pop("Cookie", None)
        payload.pop("Authorization", None)

        mode = "cookie"
        if cookie_value:
            payload["Cookie"] = cookie_value
        else:
            token = base64.b64encode(f"{username_value}:{password_value}".encode("utf-8")).decode("ascii")
            payload["Authorization"] = f"Basic {token}"
            mode = "basic"

        self._write_json(headers_path, payload)
        status = self.check_auth_status(
            cookie=cookie_value or None,
            username=username_value or None,
            password=password_value or None,
        )
        return {
            "status": "saved",
            "mode": mode,
            "path": self._display_path(headers_path),
            "auth_status": status,
        }

    def clear_server_auth(self) -> dict[str, Any]:
        headers_path = self.legacy_scraper_dir / "headers.json"
        payload = self._read_json(headers_path, default={})
        if not isinstance(payload, dict):
            payload = {}
        payload.pop("Cookie", None)
        payload.pop("Authorization", None)

        if headers_path.exists():
            self._write_json(headers_path, payload)

        return {
            "status": "cleared",
            "path": self._display_path(headers_path),
        }

    def import_legacy_data(self) -> dict[str, Any]:
        imported_sources: list[dict[str, Any]] = []
        world_index = self._build_legacy_world_index()

        legacy_batches = [
            (
                "import:legacy-raw",
                "Legacy keyword JSON import",
                self._load_legacy_json_batch("legacy-raw"),
            ),
            (
                "import:legacy-user",
                "Legacy user JSON import",
                self._load_legacy_json_batch("legacy-user"),
            ),
            (
                "import:legacy-taiwan",
                "Legacy Taiwan workbook import",
                self._load_legacy_workbook_batch("legacy-taiwan"),
            ),
            (
                "import:legacy-starriver",
                "Legacy StarRiver workbook import",
                self._load_legacy_workbook_batch("legacy-starriver"),
            ),
        ]

        for source_key, label, worlds in legacy_batches:
            imported_sources.append(
                self._import_world_batch(
                    source_key=source_key,
                    query_label=label,
                    worlds=worlds,
                )
            )

        history_result = self._import_history_batch(world_index)
        if history_result is not None:
            imported_sources.append(history_result)

        daily_stats_rows = self._import_legacy_daily_stats()
        self._refresh_topic_memberships()

        return {
            "status": "completed",
            "imported_at": dt.datetime.now(dt.timezone.utc).isoformat(),
            "sources": imported_sources,
            "daily_stats_rows": daily_stats_rows,
            "runs": self.list_runs(limit=12),
            "topics": self.list_topics(),
        }

    def update_world_record(
        self,
        *,
        source: str,
        world_id: str,
        changes: dict[str, Any],
    ) -> dict[str, Any]:
        source_key = self._editable_source_key(source)
        current = next(
            (item for item in self.storage.load_latest_worlds(source_key) if item.get("id") == world_id),
            None,
        )
        if current is None:
            raise KeyError(f"World not found in source {source}: {world_id}")

        editable = self._clone_world(self._normalise_db_world(current))
        editable["_db_source_key"] = source_key

        text_fields = {
            "name",
            "author_id",
            "author_name",
            "updated_at",
            "publication_date",
            "labs_publication_date",
            "release_status",
        }
        numeric_fields = {"visits", "favorites", "heat", "popularity", "capacity"}

        for field in text_fields:
            if field in changes:
                editable[field] = self._clean_optional_text(changes.get(field))

        for field in numeric_fields:
            if field in changes:
                editable[field] = self._to_optional_int(changes.get(field))

        if "tags" in changes:
            raw_tags = changes.get("tags")
            if isinstance(raw_tags, str):
                editable["tags"] = [item.strip() for item in raw_tags.split(",") if item.strip()]
            elif isinstance(raw_tags, list):
                editable["tags"] = [str(item).strip() for item in raw_tags if str(item).strip()]

        if "portal_links" in changes:
            portal_links = self._normalise_portal_links(changes.get("portal_links"))
            editable["portal_links"] = portal_links
            self._update_world_properties(
                world_id,
                {
                    "portal_links": portal_links,
                },
            )

        editable["metrics"] = self._calculate_metrics_for_world(editable)

        started_at = dt.datetime.now(dt.timezone.utc).isoformat()
        run_id = self.storage.create_run(
            source_key=source_key,
            job_key=None,
            trigger_type="edit",
            query_label=f"Edit {world_id}",
            started_at=started_at,
        )
        try:
            self.storage.insert_world_snapshots(
                run_id=run_id,
                source_key=source_key,
                fetched_at=started_at,
                worlds=[editable],
            )
            self.storage.finish_run(
                run_id,
                status="completed",
                finished_at=dt.datetime.now(dt.timezone.utc).isoformat(),
                world_count=1,
            )
        except Exception as exc:
            self.storage.finish_run(
                run_id,
                status="failed",
                finished_at=dt.datetime.now(dt.timezone.utc).isoformat(),
                error_text=str(exc),
            )
            raise

        self._refresh_topic_memberships()
        portal_links_saved_to = self._display_path(self.world_properties_path) if "portal_links" in changes else None
        refreshed = next(
            (item for item in self.load_worlds(source, dedupe=False) if item.get("id") == world_id),
            editable,
        )
        return {
            "status": "updated",
            "source": source,
            "world": refreshed,
            "run_id": run_id,
            "portal_links_count": len(self._normalise_portal_links(refreshed.get("portal_links"))),
            "portal_links_saved_to": portal_links_saved_to,
        }

    def delete_world_record(self, *, source: str, world_id: str) -> dict[str, Any]:
        source_key = self._editable_source_key(source)
        started_at = dt.datetime.now(dt.timezone.utc).isoformat()
        run_id = self.storage.create_run(
            source_key=source_key,
            job_key=None,
            trigger_type="delete",
            query_label=f"Delete {world_id}",
            started_at=started_at,
        )
        try:
            self.storage.delete_world_snapshots(source_key, world_id)
            self.storage.finish_run(
                run_id,
                status="completed",
                finished_at=dt.datetime.now(dt.timezone.utc).isoformat(),
                world_count=0,
            )
        except Exception as exc:
            self.storage.finish_run(
                run_id,
                status="failed",
                finished_at=dt.datetime.now(dt.timezone.utc).isoformat(),
                error_text=str(exc),
            )
            raise
        self._refresh_topic_memberships()
        return {
            "status": "deleted",
            "source": source,
            "world_id": world_id,
            "run_id": run_id,
        }

    def list_job_blacklist(self, job_key: str) -> dict[str, Any]:
        resolved = self._resolve_job_config_for_blacklist(job_key)
        return {
            "job_key": job_key,
            "path": self._display_path(resolved["blacklist_path"]),
            "items": sorted(self._load_blacklist(resolved["blacklist_path"])),
        }

    def add_job_blacklist_entry(self, *, job_key: str, world_id: str) -> dict[str, Any]:
        resolved = self._resolve_job_config_for_blacklist(job_key)
        cleaned = self._clean_optional_text(world_id)
        if not cleaned:
            raise ValueError("world_id is required")
        existing = set(self._load_blacklist(resolved["blacklist_path"]))
        existing.add(cleaned)
        self._write_blacklist(resolved["blacklist_path"], existing)
        removed = 0
        if str(resolved.get("source_key", "")).strip():
            before = sum(
                1
                for item in self.storage.load_latest_worlds(resolved["source_key"])
                if item.get("id") == cleaned
            )
            removed = self.storage.delete_world_snapshots(resolved["source_key"], cleaned)
            if removed is None:
                after = sum(
                    1
                    for item in self.storage.load_latest_worlds(resolved["source_key"])
                    if item.get("id") == cleaned
                )
                removed = max(before - after, 0)
            self._refresh_topic_memberships()
        return {
            "status": "added",
            "job_key": job_key,
            "world_id": cleaned,
            "removed_from_db": removed,
            "items": sorted(existing),
            "path": self._display_path(resolved["blacklist_path"]),
        }

    def remove_job_blacklist_entry(self, *, job_key: str, world_id: str) -> dict[str, Any]:
        resolved = self._resolve_job_config_for_blacklist(job_key)
        cleaned = self._clean_optional_text(world_id)
        if not cleaned:
            raise ValueError("world_id is required")
        existing = set(self._load_blacklist(resolved["blacklist_path"]))
        existing.discard(cleaned)
        self._write_blacklist(resolved["blacklist_path"], existing)
        return {
            "status": "removed",
            "job_key": job_key,
            "world_id": cleaned,
            "items": sorted(existing),
            "path": self._display_path(resolved["blacklist_path"]),
        }

    def _resolve_job_for_display(self, job_key: str) -> dict[str, Any]:
        import os
        job_configs = self._read_json(self.jobs_path, default={})
        job_cfg = job_configs.get(job_key, {})
        def _mtime(rel: str | None) -> float:
            if not rel:
                return 0.0
            p = self.repo_root / rel
            return p.stat().st_mtime if p.exists() else 0.0
        cache_mtime = (
            os.path.getmtime(self.jobs_path) if self.jobs_path.exists() else 0.0,
            _mtime(job_cfg.get("include_user_ids_file")),
            _mtime(job_cfg.get("exclude_author_ids_file")),
        )
        cache_entry = self._job_display_cache.get(job_key)
        if cache_entry and cache_entry[0] == cache_mtime:
            return cache_entry[1]
        resolved = self._resolve_job_config(job_key, job_cfg)
        result = {
            "whitelist": set(resolved.get("include_user_ids", [])),
            "exclude_ids": resolved.get("exclude_author_ids", set()),
            "name_bl": resolved.get("blacklist_world_name_substrings", []),
        }
        self._job_display_cache[job_key] = (cache_mtime, result)
        return result

    def _load_world_properties(self) -> dict[str, Any]:
        payload = self._read_json(self.world_properties_path, default={})
        return payload if isinstance(payload, dict) else {}

    def _save_world_properties(self, payload: dict[str, Any]) -> None:
        self._write_json(self.world_properties_path, payload)

    def _update_world_properties(self, world_id: str, updates: dict[str, Any]) -> None:
        cleaned_world_id = self._clean_optional_text(world_id)
        if not cleaned_world_id:
            return
        payload = self._load_world_properties()
        current = payload.get(cleaned_world_id)
        if not isinstance(current, dict):
            current = {}
        next_value = dict(current)
        for key, value in updates.items():
            if key == "portal_links":
                cleaned_links = self._normalise_portal_links(value)
                if cleaned_links:
                    next_value[key] = cleaned_links
                else:
                    next_value.pop(key, None)
            elif value not in (None, "", []):
                next_value[key] = value
            else:
                next_value.pop(key, None)
        if next_value:
            payload[cleaned_world_id] = next_value
        else:
            payload.pop(cleaned_world_id, None)
        self._save_world_properties(payload)

    def _apply_world_properties(self, world: dict[str, Any], *, properties: dict[str, Any] | None = None) -> dict[str, Any]:
        world_id = self._clean_optional_text(world.get("id"))
        if not world_id:
            return world
        payload = (properties or self._load_world_properties()).get(world_id)
        if not isinstance(payload, dict) or not payload:
            return world
        merged = self._clone_world(world)
        if "portal_links" in payload:
            merged["portal_links"] = self._normalise_portal_links(payload.get("portal_links"))
        return merged

    def _normalise_portal_links(self, value: Any) -> list[str]:
        if value is None:
            return []
        if isinstance(value, str):
            items = re.split(r"[\n,]", value)
        elif isinstance(value, list):
            items = [str(item) for item in value]
        else:
            items = [str(value)]
        cleaned: list[str] = []
        seen: set[str] = set()
        for item in items:
            text = str(item).strip()
            if not text:
                continue
            match = re.search(r"(wrld_[A-Za-z0-9\-]+)", text)
            normalised = match.group(1) if match else text
            if normalised not in seen:
                seen.add(normalised)
                cleaned.append(normalised)
        return cleaned

    def load_pending_worlds(self, job_key: str) -> list[dict[str, Any]]:
        source_key = f"job:{job_key}"
        resolved = self._resolve_job_for_display(job_key)
        whitelist = resolved["whitelist"]
        exclude_ids = resolved["exclude_ids"]
        name_bl = resolved["name_bl"]
        if not whitelist:
            return []
        worlds = self.storage.load_latest_worlds(source_key)
        worlds = [self._normalise_db_world(w) for w in worlds]
        pending = []
        for w in worlds:
            author_id = (w.get("author_id") or "").strip()
            if author_id in whitelist:
                continue
            if author_id in exclude_ids:
                continue
            if self._should_exclude_world(
                w,
                exclude_author_ids=exclude_ids,
                whitelist_author_ids=whitelist,
                name_blacklist=name_bl,
            ):
                continue
            pending.append(w)
        pending.sort(key=lambda w: (w.get("author_name") or "").casefold())
        return pending

    def list_job_creator_whitelist(self, job_key: str) -> dict[str, Any]:
        resolved = self._resolve_job_config_for_list_file(
            job_key,
            "include_user_ids_file",
            "has no include_user_ids_file configured",
        )
        return {
            "job_key": job_key,
            "path": self._display_path(resolved["list_path"]),
            "items": sorted(self._load_blacklist(resolved["list_path"])),
        }

    def add_job_creator_whitelist_entry(self, *, job_key: str, user_id: str) -> dict[str, Any]:
        resolved = self._resolve_job_config_for_list_file(
            job_key,
            "include_user_ids_file",
            "has no include_user_ids_file configured",
        )
        cleaned = self._clean_optional_text(user_id)
        if not cleaned:
            raise ValueError("user_id is required")
        if not cleaned.startswith("usr_"):
            raise ValueError("user_id must look like usr_...")
        existing = set(self._load_blacklist(resolved["list_path"]))
        existing.add(cleaned)
        self._write_blacklist(resolved["list_path"], existing)
        return {
            "status": "added",
            "job_key": job_key,
            "user_id": cleaned,
            "items": sorted(existing),
            "path": self._display_path(resolved["list_path"]),
        }

    def remove_job_creator_whitelist_entry(self, *, job_key: str, user_id: str) -> dict[str, Any]:
        resolved = self._resolve_job_config_for_list_file(
            job_key,
            "include_user_ids_file",
            "has no include_user_ids_file configured",
        )
        cleaned = self._clean_optional_text(user_id)
        if not cleaned:
            raise ValueError("user_id is required")
        existing = set(self._load_blacklist(resolved["list_path"]))
        existing.discard(cleaned)
        self._write_blacklist(resolved["list_path"], existing)
        return {
            "status": "removed",
            "job_key": job_key,
            "user_id": cleaned,
            "items": sorted(existing),
            "path": self._display_path(resolved["list_path"]),
        }

    def list_job_creator_blacklist(self, job_key: str) -> dict[str, Any]:
        resolved = self._resolve_job_config_for_list_file(
            job_key,
            "exclude_author_ids_file",
            "has no exclude_author_ids_file configured",
        )
        return {
            "job_key": job_key,
            "path": self._display_path(resolved["list_path"]),
            "items": sorted(self._load_blacklist(resolved["list_path"])),
        }

    def add_job_creator_blacklist_entry(self, *, job_key: str, user_id: str) -> dict[str, Any]:
        resolved = self._resolve_job_config_for_list_file(
            job_key,
            "exclude_author_ids_file",
            "has no exclude_author_ids_file configured",
        )
        cleaned = self._clean_optional_text(user_id)
        if not cleaned:
            raise ValueError("user_id is required")
        if not cleaned.startswith("usr_"):
            raise ValueError("user_id must look like usr_...")
        existing = set(self._load_blacklist(resolved["list_path"]))
        existing.add(cleaned)
        self._write_blacklist(resolved["list_path"], existing)
        return {
            "status": "added",
            "job_key": job_key,
            "user_id": cleaned,
            "items": sorted(existing),
            "path": self._display_path(resolved["list_path"]),
        }

    def remove_job_creator_blacklist_entry(self, *, job_key: str, user_id: str) -> dict[str, Any]:
        resolved = self._resolve_job_config_for_list_file(
            job_key,
            "exclude_author_ids_file",
            "has no exclude_author_ids_file configured",
        )
        cleaned = self._clean_optional_text(user_id)
        if not cleaned:
            raise ValueError("user_id is required")
        existing = set(self._load_blacklist(resolved["list_path"]))
        existing.discard(cleaned)
        self._write_blacklist(resolved["list_path"], existing)
        return {
            "status": "removed",
            "job_key": job_key,
            "user_id": cleaned,
            "items": sorted(existing),
            "path": self._display_path(resolved["list_path"]),
        }

    def run_self_check(self) -> dict[str, Any]:
        warnings: list[str] = []
        source_summaries: list[dict[str, Any]] = []

        headers_path = self.legacy_scraper_dir / "headers.json"
        headers_payload = self._read_json(headers_path, default={})
        if not headers_path.exists():
            warnings.append(
                "Missing local auth headers file. This is fine if you use browser-only auth in the page, "
                "but self-check cannot verify browser localStorage sessions."
            )
        elif not isinstance(headers_payload, dict) or not (
            str(headers_payload.get("Cookie", "")).strip() or str(headers_payload.get("Authorization", "")).strip()
        ):
            warnings.append("Auth headers file exists but has no Cookie or Authorization value.")

        available_sources = 0
        for item in self.list_sources():
            if item["available"]:
                available_sources += 1
            raw_worlds = self.load_worlds(item["key"], dedupe=False)
            deduped_worlds = self._dedupe_worlds(raw_worlds)
            ids = [world.get("id") for world in raw_worlds if world.get("id")]
            duplicates = sum(count - 1 for count in Counter(ids).values() if count > 1)
            missing_ids = sum(1 for world in raw_worlds if not world.get("id"))
            missing_names = sum(1 for world in raw_worlds if not world.get("name"))
            suspicious_metrics = sum(
                1
                for world in deduped_worlds
                if (world.get("visits") in (None, 0)) and self._to_int(world.get("favorites")) > 0
            )

            if item["origin"] == "legacy" and not item["available"]:
                warnings.append(f"Missing legacy source: {item['path']}")
            if duplicates and item["key"] != "db:all":
                warnings.append(f"{item['key']} contains {duplicates} duplicate world IDs.")
            if duplicates and item["key"] == "db:all":
                warnings.append(
                    f"db:all aggregates multiple DB sources; {duplicates} repeated world IDs appear across sources."
                )
            if missing_ids:
                warnings.append(f"{item['key']} contains {missing_ids} rows without world IDs.")
            if missing_names:
                warnings.append(f"{item['key']} contains {missing_names} rows without names.")
            if suspicious_metrics:
                warnings.append(
                    f"{item['key']} contains {suspicious_metrics} suspicious rows where visits are missing/zero but favorites are present."
                )

            source_summaries.append(
                {
                    "source": item["key"],
                    "count": len(deduped_worlds),
                    "duplicates": duplicates,
                    "missing_ids": missing_ids,
                    "missing_names": missing_names,
                    "suspicious_metrics": suspicious_metrics,
                }
            )

        for job in self.list_jobs():
            if not job["ready"]:
                warnings.append(f"Job {job['job_key']} is not ready: {job['reason']}")

        for topic in self.list_topics():
            if not topic["rules"]:
                warnings.append(f"Topic {topic['topic_key']} has no active rules.")

        history = self.load_history()
        history_points = 0
        for wid, entries in history.items():
            history_points += len(entries)
            timestamps = [entry.get("timestamp") for entry in entries if entry.get("timestamp") is not None]
            if timestamps != sorted(timestamps):
                warnings.append(f"History for {wid} is not sorted by timestamp.")
            if any(value is None for value in timestamps):
                warnings.append(f"History for {wid} contains missing timestamps.")

        if available_sources == 0:
            warnings.append("No legacy or database sources are available.")
        if history_points == 0:
            warnings.append("No history records were found in legacy or database data.")

        return {
            "status": "warning" if warnings else "ok",
            "checked_at": dt.datetime.now(dt.timezone.utc).isoformat(),
            "warnings": warnings,
            "sources": source_summaries,
            "jobs": self.list_jobs(),
            "history_worlds": len(history),
            "history_records": history_points,
        }

    def build_world_graph(
        self,
        *,
        source: str = "db:all",
        edge_types: list[str] | None = None,
        min_shared_tags: int = 2,
        exclude_system_tags: bool = True,
        max_nodes: int = 300,
    ) -> dict[str, Any]:
        """Return nodes + edges for a force-directed world network graph.

        Nodes are worlds sorted by visits desc (capped at *max_nodes*).
        Edges are built from two relationship types:
        - "author" : worlds sharing the same author_id
        - "tag"    : worlds sharing >= min_shared_tags non-system tags
        - "portal" : world.portal_links pointing at another loaded world
        """
        edge_types = edge_types or ["author", "tag", "portal"]
        all_worlds = self.load_worlds(source, sort="visits", direction="desc")
        worlds = all_worlds[:max_nodes]
        portal_expanded_nodes = 0
        if "portal" in edge_types and worlds:
            world_map = {
                self._clean_optional_text(world.get("id")): world
                for world in all_worlds
                if self._clean_optional_text(world.get("id"))
            }
            selected_ids = {
                self._clean_optional_text(world.get("id"))
                for world in worlds
                if self._clean_optional_text(world.get("id"))
            }
            portal_target_ids: list[str] = []
            seen_targets: set[str] = set()
            for world in worlds:
                for target_id in self._normalise_portal_links(world.get("portal_links")):
                    if target_id in selected_ids or target_id in seen_targets or target_id not in world_map:
                        continue
                    seen_targets.add(target_id)
                    portal_target_ids.append(target_id)
            portal_expansion_cap = min(max(max_nodes // 3, 12), 80)
            for target_id in portal_target_ids[:portal_expansion_cap]:
                target_world = world_map.get(target_id)
                if not target_world:
                    continue
                worlds.append(target_world)
                selected_ids.add(target_id)
                portal_expanded_nodes += 1
        topic_membership_map: dict[str, list[str]] = {}
        for topic in self.storage.list_topics():
            for membership in self.storage.list_topic_memberships(topic["topic_key"]):
                topic_membership_map.setdefault(membership["world_id"], []).append(topic["topic_key"])

        nodes: list[dict[str, Any]] = []
        for w in worlds:
            nodes.append(
                {
                    "id": w["id"],
                    "name": w.get("name") or w["id"],
                    "author_id": w.get("author_id"),
                    "author_name": w.get("author_name"),
                    "visits": self._to_int(w.get("visits")),
                    "favorites": self._to_int(w.get("favorites")),
                    "tags": w.get("tags") or [],
                    "topic_keys": sorted(topic_membership_map.get(w["id"], [])),
                    "source": w.get("source", source),
                    "release_status": w.get("release_status"),
                    "world_url": w.get("world_url"),
                    "portal_links": w.get("portal_links") or [],
                    "publication_date": w.get("publication_date"),
                    "updated_at": w.get("updated_at"),
                    "fetched_at": w.get("fetched_at"),
                }
            )

        edges: list[dict[str, Any]] = []

        if "author" in edge_types:
            author_map: dict[str, list[str]] = {}
            for node in nodes:
                aid = node.get("author_id")
                if aid:
                    author_map.setdefault(aid, []).append(node["id"])
            for aid, wids in author_map.items():
                if len(wids) < 2:
                    continue
                for i in range(len(wids)):
                    for j in range(i + 1, len(wids)):
                        edges.append(
                            {
                                "source": wids[i],
                                "target": wids[j],
                                "type": "same_author",
                                "weight": 3,
                                "shared": [],
                            }
                        )

        if "tag" in edge_types:
            def _meaningful_tags(tags: list) -> set[str]:
                result = set()
                for t in tags:
                    if not isinstance(t, str):
                        continue
                    if exclude_system_tags and t.startswith("system_"):
                        continue
                    result.add(t)
                return result

            tag_sets = {node["id"]: _meaningful_tags(node["tags"]) for node in nodes}
            tag_edges: list[dict[str, Any]] = []
            for i in range(len(nodes)):
                for j in range(i + 1, len(nodes)):
                    id_a = nodes[i]["id"]
                    id_b = nodes[j]["id"]
                    shared = tag_sets[id_a] & tag_sets[id_b]
                    if len(shared) >= min_shared_tags:
                        tag_edges.append(
                            {
                                "source": id_a,
                                "target": id_b,
                                "type": "shared_tag",
                                "weight": len(shared),
                                "shared": sorted(shared)[:8],
                            }
                        )
            # Limit tag edges to top 1500 by weight to keep JSON payload sane
            tag_edges.sort(key=lambda e: e["weight"], reverse=True)
            edges.extend(tag_edges[:1500])

        if "portal" in edge_types:
            node_ids = {node["id"] for node in nodes if node.get("id")}
            seen_portal_edges: set[tuple[str, str]] = set()
            for node in nodes:
                source_id = node.get("id")
                if not source_id:
                    continue
                for target_id in self._normalise_portal_links(node.get("portal_links")):
                    if target_id == source_id or target_id not in node_ids:
                        continue
                    edge_key = tuple(sorted((source_id, target_id)))
                    if edge_key in seen_portal_edges:
                        continue
                    seen_portal_edges.add(edge_key)
                    edges.append(
                        {
                            "source": source_id,
                            "target": target_id,
                            "type": "portal_link",
                            "weight": 2,
                            "shared": ["portal_link"],
                        }
                    )

        return {
            "node_count": len(nodes),
            "edge_count": len(edges),
            "base_node_count": min(len(all_worlds), max_nodes),
            "portal_expanded_nodes": portal_expanded_nodes,
            "source": source,
            "nodes": nodes,
            "edges": edges,
        }

    def collect_tags(self, worlds: list[dict[str, Any]]) -> list[str]:
        tags = {tag for world in worlds for tag in world.get("tags", []) if tag}
        return sorted(tags)

    def _store_sync_result(
        self,
        *,
        source_key: str,
        job_key: str | None,
        trigger_type: str,
        query_label: str | None,
        worlds: list[dict[str, Any]],
        warnings: list[str] | None = None,
        meta: dict[str, Any] | None = None,
        query_batches: list[dict[str, Any]] | None = None,
    ) -> dict[str, Any]:
        started_at = dt.datetime.now(dt.timezone.utc).isoformat()
        run_id = self.storage.create_run(
            source_key=source_key,
            job_key=job_key,
            trigger_type=trigger_type,
            query_label=query_label,
            started_at=started_at,
        )

        try:
            normalised = [self._normalise_api_world(world, self._public_db_source_key(source_key)) for world in worlds]
            normalised = self._dedupe_worlds(normalised)
            normalised_world_ids = {
                str(world.get("id") or "").strip()
                for world in normalised
                if str(world.get("id") or "").strip()
            }
            existing_world_ids = self.storage.get_existing_world_ids(normalised_world_ids)
            self.storage.insert_world_snapshots(
                run_id=run_id,
                source_key=source_key,
                fetched_at=started_at,
                worlds=normalised,
            )
            if query_batches:
                self.storage.insert_run_queries(
                    run_id=run_id,
                    queries=self._build_run_query_rows(
                        query_batches=query_batches,
                        kept_world_ids=normalised_world_ids,
                        existing_world_ids=existing_world_ids,
                    ),
                )
            self.storage.upsert_daily_stats(
                source_key=source_key,
                date=dt.datetime.now(dt.timezone.utc).strftime("%Y/%m/%d"),
                total_worlds=len(normalised),
                new_worlds_today=self._calculate_new_worlds_today(normalised),
            )
            self.storage.finish_run(
                run_id,
                status="completed",
                finished_at=dt.datetime.now(dt.timezone.utc).isoformat(),
                world_count=len(normalised),
            )
        except Exception as exc:
            self.storage.finish_run(
                run_id,
                status="failed",
                finished_at=dt.datetime.now(dt.timezone.utc).isoformat(),
                error_text=str(exc),
            )
            raise

        public_source = self._public_db_source_key(source_key)
        self._refresh_topic_memberships()
        for cache_source in (public_source, "db:all"):
            try:
                self.refresh_analysis_cache(cache_source, source_run_id=run_id)
            except Exception as exc:
                logger.warning("Analysis cache refresh skipped for %s: %s", cache_source, exc)
        return {
            "run_id": run_id,
            "source": public_source,
            "query": query_label,
            "count": len(normalised),
            "items": self.load_worlds(public_source),
            "warnings": warnings or [],
            "meta": meta or {},
        }

    def _build_run_query_rows(
        self,
        *,
        query_batches: list[dict[str, Any]],
        kept_world_ids: set[str],
        existing_world_ids: set[str],
    ) -> list[dict[str, Any]]:
        rows: list[dict[str, Any]] = []
        for index, batch in enumerate(query_batches):
            result_world_ids: list[str] = []
            result_seen: set[str] = set()
            kept_hits: list[dict[str, Any]] = []
            kept_seen: set[str] = set()
            for world in batch.get("worlds", []):
                world_id = str(world.get("id") or world.get("worldId") or "").strip()
                if not world_id:
                    continue
                if world_id not in result_seen:
                    result_seen.add(world_id)
                    result_world_ids.append(world_id)
                if world_id not in kept_world_ids or world_id in kept_seen:
                    continue
                kept_seen.add(world_id)
                kept_hits.append(
                    {
                        "world_id": world_id,
                        "world_name": world.get("name"),
                        "author_id": world.get("authorId") or world.get("author_id"),
                        "rank_index": len(kept_hits) + 1,
                        "is_new_global": world_id not in existing_world_ids,
                    }
                )
            rows.append(
                {
                    "query_index": index,
                    "query_kind": batch.get("kind", "keyword"),
                    "query_value": str(batch.get("value", "")).strip(),
                    "query_label": batch.get("label"),
                    "query_payload": batch.get("payload", {}),
                    "result_count": len(result_world_ids),
                    "kept_count": len(kept_hits),
                    "new_world_count": sum(1 for hit in kept_hits if hit.get("is_new_global")),
                    "hits": kept_hits,
                }
            )
        return [row for row in rows if row.get("query_value")]

    def _make_query_batch(
        self,
        *,
        kind: str,
        value: str,
        label: str,
        worlds: list[dict[str, Any]],
        payload: dict[str, Any] | None = None,
    ) -> dict[str, Any]:
        return {
            "kind": kind,
            "value": value,
            "label": label,
            "worlds": [dict(world) for world in worlds if isinstance(world, dict)],
            "payload": dict(payload or {}),
        }

    def _prepare_sync_worlds(
        self,
        worlds: list[dict[str, Any]],
        *,
        headers: dict[str, Any] | None = None,
    ) -> tuple[list[dict[str, Any]], list[str], dict[str, Any]]:
        warnings: list[str] = []
        deduped_worlds, duplicate_count = self._dedupe_raw_world_payloads(worlds)
        missing_before = sum(1 for world in deduped_worlds if world.get("visits") is None)
        enriched_worlds = deduped_worlds
        if missing_before:
            try:
                enriched_worlds = enrich_visits(deduped_worlds, headers or None, delay=0.0)
            except Exception as exc:
                logger.warning("Visit enrichment failed: %s", exc)
                warnings.append(f"visit enrichment failed: {exc}")
        missing_after = sum(1 for world in enriched_worlds if world.get("visits") is None)
        if missing_after:
            warnings.append(
                f"{missing_after} world(s) still have no visits after API fetch; use a valid VRChat Cookie for full counters."
            )
        return enriched_worlds, warnings, {
            "duplicates_merged_before_enrich": duplicate_count,
            "missing_visits_before_enrich": missing_before,
            "missing_visits_after_enrich": missing_after,
        }

    def _import_world_batch(
        self,
        *,
        source_key: str,
        query_label: str,
        worlds: list[dict[str, Any]],
    ) -> dict[str, Any]:
        self.storage.purge_source(source_key)
        started_at = dt.datetime.now(dt.timezone.utc).isoformat()
        run_id = self.storage.create_run(
            source_key=source_key,
            job_key=None,
            trigger_type="import",
            query_label=query_label,
            started_at=started_at,
        )

        normalised = self._dedupe_worlds([dict(world) for world in worlds if world.get("id")])
        try:
            self.storage.insert_world_snapshots(
                run_id=run_id,
                source_key=source_key,
                fetched_at=started_at,
                worlds=normalised,
            )
            self.storage.finish_run(
                run_id,
                status="completed",
                finished_at=dt.datetime.now(dt.timezone.utc).isoformat(),
                world_count=len(normalised),
            )
        except Exception as exc:
            self.storage.finish_run(
                run_id,
                status="failed",
                finished_at=dt.datetime.now(dt.timezone.utc).isoformat(),
                error_text=str(exc),
            )
            raise

        return {
            "run_id": run_id,
            "source": self._public_db_source_key(source_key),
            "count": len(normalised),
            "latest_fetched_at": max(
                (world.get("fetched_at") for world in normalised if world.get("fetched_at")),
                default=None,
            ),
        }

    def _import_history_batch(self, world_index: dict[str, dict[str, Any]]) -> dict[str, Any] | None:
        history_payload = self._read_json(self.legacy_scraper_dir / "history.json", default={})
        if not isinstance(history_payload, dict):
            self.storage.purge_source("history:legacy")
            return None

        worlds: list[dict[str, Any]] = []
        for world_id, entries in history_payload.items():
            if not isinstance(entries, list):
                continue
            template = dict(world_index.get(world_id) or {})
            for entry in entries:
                if not isinstance(entry, dict):
                    continue
                timestamp = entry.get("timestamp")
                if not isinstance(timestamp, (int, float, str)) or not str(timestamp).isdigit():
                    continue
                worlds.append(self._normalise_history_snapshot(world_id, int(timestamp), entry, template))

        if not worlds:
            self.storage.purge_source("history:legacy")
            return None
        return self._import_world_batch(
            source_key="history:legacy",
            query_label="Legacy history import",
            worlds=worlds,
        )

    def _import_legacy_daily_stats(self) -> int:
        imported_rows = 0
        seen_sources: set[str] = set()
        for path in sorted(self.legacy_analytics_dir.glob("daily_stats_*.xlsx")):
            source_key = self._map_legacy_daily_stats_source_key(path)
            if source_key not in seen_sources:
                self.storage.purge_daily_stats(source_key)
                seen_sources.add(source_key)
            for row in self._read_daily_stats_xlsx(path):
                date_value = self._normalise_daily_stats_date(row.get("date"))
                if not date_value:
                    continue
                self.storage.upsert_daily_stats(
                    source_key=source_key,
                    date=date_value,
                    total_worlds=self._to_int(row.get("total_worlds")),
                    new_worlds_today=self._to_int(row.get("new_worlds_today")),
                )
                imported_rows += 1
        return imported_rows

    def _build_legacy_world_index(self) -> dict[str, dict[str, Any]]:
        index: dict[str, dict[str, Any]] = {}
        for source in ("legacy-raw", "legacy-user", "legacy-taiwan", "legacy-starriver"):
            for world in self.load_worlds(source):
                world_id = world.get("id")
                if not world_id:
                    continue
                current = index.get(world_id)
                if current is None:
                    index[world_id] = self._clone_world(world)
                    continue
                index[world_id] = self._merge_world_records(current, world)
        return index

    def _load_legacy_json_batch(self, source: str) -> list[dict[str, Any]]:
        config = self.legacy_sources[source]
        path = Path(config["path"])
        fetched_at = self._file_timestamp_iso(path)
        worlds = []
        for item in self._read_json(path, default=[]):
            if not isinstance(item, dict):
                continue
            world = self._normalise_api_world(item, self._public_db_source_key(source))
            world["fetched_at"] = fetched_at
            worlds.append(world)
        return worlds

    def _load_legacy_workbook_batch(self, source: str) -> list[dict[str, Any]]:
        config = self.legacy_sources[source]
        path = Path(config["path"])
        worlds = self._load_workbook_worlds(path, source)
        fallback_fetched_at = self._file_timestamp_iso(path)
        defaults = self._legacy_source_defaults(source)
        for world in worlds:
            world["fetched_at"] = self._normalise_snapshot_time(world.get("fetched_at")) or fallback_fetched_at
            for key, value in defaults.items():
                if value and not world.get(key):
                    world[key] = value
        return worlds

    def _normalise_history_snapshot(
        self,
        world_id: str,
        timestamp: int,
        entry: dict[str, Any],
        template: dict[str, Any],
    ) -> dict[str, Any]:
        fetched_at = dt.datetime.fromtimestamp(timestamp, dt.timezone.utc).isoformat()
        base = {
            "source": self._public_db_source_key("history:legacy"),
            "id": world_id,
            "name": self._clean_optional_text(entry.get("name")) or template.get("name"),
            "author_id": template.get("author_id"),
            "author_name": template.get("author_name"),
            "capacity": self._to_int(template.get("capacity")),
            "visits": self._to_int(entry.get("visits")),
            "favorites": self._to_int(entry.get("favorites")),
            "heat": self._to_int(entry.get("heat")),
            "popularity": self._to_int(entry.get("popularity")),
            "created_at": self._clean_optional_text(entry.get("created_at")) or template.get("created_at"),
            "updated_at": self._clean_optional_text(entry.get("updated_at")) or template.get("updated_at"),
            "publication_date": self._clean_optional_text(entry.get("publicationDate")) or template.get("publication_date"),
            "labs_publication_date": self._clean_optional_text(entry.get("labsPublicationDate")) or template.get("labs_publication_date"),
            "release_status": template.get("release_status"),
            "image_url": template.get("image_url"),
            "thumbnail_url": template.get("thumbnail_url"),
            "tags": list(template.get("tags") or []),
            "world_url": template.get("world_url") or (f"https://vrchat.com/home/world/{world_id}" if world_id else None),
            "fetched_at": fetched_at,
        }
        publication_dt = _parse_date(base.get("publication_date"))
        labs_dt = _parse_date(base.get("labs_publication_date"))
        updated_dt = _parse_date(base.get("updated_at"))
        fetched_dt = _parse_date(fetched_at)

        favorite_rate = None
        if base["visits"]:
            favorite_rate = round((base["favorites"] / base["visits"]) * 100, 2)

        visits_per_day = None
        if publication_dt and fetched_dt:
            published_days = max((fetched_dt - publication_dt).days, 1)
            visits_per_day = round(base["visits"] / published_days, 2)

        labs_to_publication_days = None
        if publication_dt and labs_dt:
            labs_to_publication_days = (publication_dt - labs_dt).days

        days_since_update = None
        if updated_dt and fetched_dt:
            days_since_update = (fetched_dt - updated_dt).days

        base["metrics"] = {
            "favorite_rate": favorite_rate,
            "labs_to_publication_days": labs_to_publication_days,
            "days_since_update": days_since_update,
            "visits_per_day": visits_per_day,
        }
        return base

    def _map_legacy_daily_stats_source_key(self, path: Path) -> str:
        suffix = path.stem.removeprefix("daily_stats_")
        lowered = suffix.casefold()
        if lowered == "taiwan":
            return "import:legacy-taiwan"
        if lowered in {"starriver", "starriverarts"}:
            return "import:legacy-starriver"
        return f"import:legacy-daily:{suffix}"

    def _legacy_source_defaults(self, source: str) -> dict[str, Any]:
        if source != "legacy-starriver":
            return {}
        config = self._load_job_configs().get("starriver")
        if not config:
            return {}
        resolved = self._resolve_job_config("starriver", config)
        user_id = resolved.get("user_id")
        if not user_id:
            return {}
        creator_name = str(config.get("author_name", "")).strip() or "StarRiver Arts"
        return {
            "author_id": user_id,
            "author_name": creator_name,
        }

    def _file_timestamp_iso(self, path: Path) -> str:
        if not path.exists():
            return dt.datetime.now(dt.timezone.utc).isoformat()
        return dt.datetime.fromtimestamp(path.stat().st_mtime, dt.timezone.utc).isoformat()

    def _normalise_snapshot_time(self, value: Any) -> str | None:
        parsed = _parse_date(value) if value else None
        return parsed.isoformat() if parsed else None

    def _normalise_daily_stats_date(self, value: Any) -> str | None:
        if value in (None, ""):
            return None
        if isinstance(value, dt.datetime):
            return value.strftime("%Y/%m/%d")
        text = str(value).strip()
        parsed = _parse_date(text)
        if parsed:
            return parsed.strftime("%Y/%m/%d")
        match = re.match(r"^(\d{4})[/-](\d{1,2})[/-](\d{1,2})$", text)
        if not match:
            return None
        return f"{int(match.group(1)):04d}/{int(match.group(2)):02d}/{int(match.group(3)):02d}"

    def _sync_topic_catalog(self, *, refresh: bool = False) -> None:
        topics = []
        configs = self._load_topic_configs()
        for topic_key, config in sorted(configs.items(), key=lambda item: item[1].get("sort_order", 0)):
            resolved = self._resolve_topic_config(topic_key, config)
            topics.append(
                {
                    "topic_key": topic_key,
                    "label": resolved["label"],
                    "description": resolved.get("description"),
                    "color": resolved.get("color"),
                    "sort_order": resolved.get("sort_order", 0),
                    "is_active": resolved.get("is_active", True),
                }
            )
        self.storage.upsert_topics(topics)
        for topic_key, config in configs.items():
            resolved = self._resolve_topic_config(topic_key, config)
            self.storage.replace_topic_rules(topic_key, resolved["rules"])
        stale_topic_keys = {
            topic["topic_key"]
            for topic in self.storage.list_topics()
            if topic["topic_key"] not in configs
        }
        for topic_key in stale_topic_keys:
            self.storage.delete_topic(topic_key)
        if refresh and self.storage.has_data():
            self._refresh_topic_memberships()

    def _load_topic_configs(self) -> dict[str, dict[str, Any]]:
        data = self._read_json(self.topics_path, default={})
        if isinstance(data, dict):
            return {str(key): value for key, value in data.items() if isinstance(value, dict)}
        return {}

    def _resolve_topic_config(self, topic_key: str, config: dict[str, Any]) -> dict[str, Any]:
        rules = []
        for index, rule in enumerate(config.get("rules", [])):
            if not isinstance(rule, dict):
                continue
            rule_type = str(rule.get("type", "")).strip()
            rule_value = str(rule.get("value", "")).strip()
            env_name = str(rule.get("value_env", "")).strip()
            if not rule_value and env_name:
                rule_value = os.getenv(env_name, "").strip()
            rules.append(
                {
                    "rule_type": rule_type,
                    "rule_value": rule_value,
                    "sort_order": index,
                    "is_active": bool(rule.get("is_active", True)),
                }
            )
        topic_type = str(config.get("topic_type", "job")).strip().casefold() or "job"
        if topic_type not in {"job", "view"}:
            topic_type = "job"
        return {
            "label": str(config.get("label", topic_key)).strip() or topic_key,
            "description": str(config.get("description", "")).strip() or None,
            "color": str(config.get("color", "")).strip() or None,
            "topic_type": topic_type,
            "sort_order": self._to_int(config.get("sort_order")),
            "is_active": bool(config.get("is_active", True)),
            "rules": [rule for rule in rules if rule["rule_type"] and rule["rule_value"]],
        }

    def _refresh_topic_memberships(self, topic_keys: set[str] | None = None) -> None:
        topics = [
            topic
            for topic in self.storage.list_topics()
            if topic_keys is None or topic["topic_key"] in topic_keys
        ]
        rules_by_topic = {
            topic["topic_key"]: [
                rule for rule in self.storage.list_topic_rules(topic["topic_key"]) if rule.get("is_active", 1)
            ]
            for topic in topics
        }
        needs_all_worlds = any(
            (not any(rule.get("rule_type") == "source" for rule in rules))
            and any(rule.get("rule_type") != "source" for rule in rules)
            for rules in rules_by_topic.values()
        )
        worlds = self.load_worlds("db:all") if needs_all_worlds else []
        for topic in topics:
            rules = rules_by_topic.get(topic["topic_key"], [])
            existing = self.storage.get_existing_topic_memberships(topic["topic_key"])
            memberships_by_world: dict[str, dict[str, Any]] = {}

            source_rules = [rule for rule in rules if rule.get("rule_type") == "source"]
            other_rules = [rule for rule in rules if rule.get("rule_type") != "source"]

            source_worlds_by_id: dict[str, tuple[dict[str, Any], str]] = {}
            for rule in source_rules:
                for world in self._load_topic_source_worlds(rule.get("rule_value") or ""):
                    world_id = world.get("id")
                    if not world_id:
                        continue
                    source_worlds_by_id[world_id] = (world, f"source:{rule.get('rule_value')}")

            if source_rules and other_rules:
                for world_id, (world, source_match) in source_worlds_by_id.items():
                    matched_by = self._match_topic(world, other_rules)
                    if not matched_by:
                        continue
                    current = existing.get(world_id)
                    memberships_by_world[world_id] = {
                        "world_id": world_id,
                        "first_seen_at": (current or {}).get("first_seen_at")
                        or world.get("fetched_at")
                        or dt.datetime.now(dt.timezone.utc).isoformat(),
                        "last_seen_at": world.get("fetched_at") or dt.datetime.now(dt.timezone.utc).isoformat(),
                        "matched_by": f"{source_match} & {matched_by}",
                    }
            else:
                for world_id, (world, source_match) in source_worlds_by_id.items():
                    current = existing.get(world_id)
                    memberships_by_world[world_id] = {
                        "world_id": world_id,
                        "first_seen_at": (current or {}).get("first_seen_at")
                        or world.get("fetched_at")
                        or dt.datetime.now(dt.timezone.utc).isoformat(),
                        "last_seen_at": world.get("fetched_at") or dt.datetime.now(dt.timezone.utc).isoformat(),
                        "matched_by": source_match,
                    }

            if not source_rules:
                for world in worlds:
                    matched_by = self._match_topic(world, other_rules)
                    if not matched_by:
                        continue
                    world_id = world.get("id")
                    if not world_id:
                        continue
                    current = existing.get(world_id)
                    memberships_by_world.setdefault(
                        world_id,
                        {
                            "world_id": world_id,
                            "first_seen_at": (current or {}).get("first_seen_at")
                            or world.get("fetched_at")
                            or dt.datetime.now(dt.timezone.utc).isoformat(),
                            "last_seen_at": world.get("fetched_at") or dt.datetime.now(dt.timezone.utc).isoformat(),
                            "matched_by": matched_by,
                        },
                    )
            self.storage.replace_topic_memberships(topic["topic_key"], list(memberships_by_world.values()))

    def _match_topic(self, world: dict[str, Any], rules: list[dict[str, Any]]) -> str | None:
        for rule in rules:
            matched = self._world_matches_rule(world, rule["rule_type"], rule["rule_value"])
            if matched:
                return f"{rule['rule_type']}:{rule['rule_value']}"
        return None

    def _world_matches_rule(self, world: dict[str, Any], rule_type: str, rule_value: str) -> bool:
        if not rule_value:
            return False
        if rule_type == "source":
            source_value = world.get("source") or ""
            return source_value == rule_value or source_value == self._public_db_source_key(rule_value)
        if rule_type == "author_id":
            return (world.get("author_id") or "") == rule_value
        if rule_type == "world_id":
            return (world.get("id") or "") == rule_value
        if rule_type == "tag":
            return rule_value in (world.get("tags") or [])
        if rule_type == "keyword":
            haystacks = [
                world.get("name") or "",
                world.get("author_name") or "",
                world.get("author_id") or "",
                " ".join(world.get("tags") or []),
            ]
            needle = rule_value.casefold()
            return any(needle in text.casefold() for text in haystacks if text)
        if rule_type == "published_within_days":
            days = self._to_int(rule_value)
            if days <= 0:
                return False
            return self._within_days(self._new_world_event_date(world), days, dt.datetime.now(dt.timezone.utc))
        if rule_type == "visits_min":
            return self._to_int(world.get("visits")) >= self._to_int(rule_value)
        if rule_type == "visits_max":
            return self._to_int(world.get("visits")) <= self._to_int(rule_value)
        if rule_type == "favorites_min":
            return self._to_int(world.get("favorites")) >= self._to_int(rule_value)
        if rule_type == "favorites_max":
            return self._to_int(world.get("favorites")) <= self._to_int(rule_value)
        if rule_type == "heat_min":
            return self._to_int(world.get("heat")) >= self._to_int(rule_value)
        if rule_type == "heat_max":
            return self._to_int(world.get("heat")) <= self._to_int(rule_value)
        if rule_type == "popularity_min":
            return self._to_int(world.get("popularity")) >= self._to_int(rule_value)
        if rule_type == "popularity_max":
            return self._to_int(world.get("popularity")) <= self._to_int(rule_value)
        if rule_type == "favorite_rate_min":
            metrics = world.get("metrics", {}) if isinstance(world.get("metrics"), dict) else {}
            favorite_rate = self._to_float(metrics.get("favorite_rate"))
            if favorite_rate is None:
                visits = self._to_int(world.get("visits"))
                favorites = self._to_int(world.get("favorites"))
                favorite_rate = (favorites / visits) * 100 if visits > 0 else None
            return (favorite_rate or 0.0) >= float(rule_value)
        if rule_type == "favorite_rate_max":
            metrics = world.get("metrics", {}) if isinstance(world.get("metrics"), dict) else {}
            favorite_rate = self._to_float(metrics.get("favorite_rate"))
            if favorite_rate is None:
                visits = self._to_int(world.get("visits"))
                favorites = self._to_int(world.get("favorites"))
                favorite_rate = (favorites / visits) * 100 if visits > 0 else None
            return (favorite_rate or 0.0) <= float(rule_value)
        if rule_type == "updated_within_days":
            days = self._to_int(rule_value)
            if days <= 0:
                return False
            return self._within_days(world.get("updated_at"), days, dt.datetime.now(dt.timezone.utc))
        return False

    def _load_job_configs(self) -> dict[str, dict[str, Any]]:
        data = self._read_json(self.jobs_path, default={})
        if isinstance(data, dict):
            return {str(key): value for key, value in data.items() if isinstance(value, dict)}
        return {}

    def _resolve_job_config(self, job_key: str, config: dict[str, Any]) -> dict[str, Any]:
        job_type = str(config.get("type", "")).strip()
        label = str(config.get("label", job_key)).strip() or job_key
        source_key = str(config.get("source_key", f"job:{job_key}")).strip() or f"job:{job_key}"

        resolved: dict[str, Any] = {
            "job_key": job_key,
            "label": label,
            "type": job_type,
            "source_key": source_key,
            "ready": True,
            "reason": None,
        }

        if job_type == "keywords":
            keywords = [str(item).strip() for item in config.get("keywords", []) if str(item).strip()]
            if not keywords:
                resolved["ready"] = False
                resolved["reason"] = "keywords job requires at least one keyword"
            resolved["keywords"] = keywords
            resolved["blacklist_file"] = config.get("blacklist_file")
            include_user_ids = {
                str(item).strip()
                for item in config.get("include_user_ids", [])
                if str(item).strip().startswith("usr_")
            }
            include_user_ids_file = config.get("include_user_ids_file")
            if include_user_ids_file:
                include_user_ids.update(
                    item for item in self._load_blacklist(include_user_ids_file) if item.startswith("usr_")
                )
            resolved["include_user_ids_file"] = include_user_ids_file
            resolved["include_user_ids"] = sorted(include_user_ids)
            exclude_author_ids_file = config.get("exclude_author_ids_file")
            resolved["exclude_author_ids_file"] = exclude_author_ids_file
            resolved["exclude_author_ids"] = self._load_blacklist(exclude_author_ids_file)
            resolved["creator_review_enabled"] = bool(config.get("creator_review_enabled"))
            resolved["blacklist_world_name_substrings"] = [
                s.lower() for s in config.get("blacklist_world_name_substrings", []) if s
            ]
            resolved["limit_per_keyword"] = max(self._to_int(config.get("limit_per_keyword")), 1) or 50
            return resolved

        if job_type == "user":
            user_id = str(config.get("user_id", "")).strip()
            user_id_env = str(config.get("user_id_env", "")).strip()
            if not user_id and user_id_env:
                user_id = os.getenv(user_id_env, "").strip()
            if not user_id:
                resolved["ready"] = False
                resolved["reason"] = "user job requires user_id or user_id_env"
            elif not user_id.startswith("usr_"):
                resolved["ready"] = False
                resolved["reason"] = "user job requires a VRChat user ID like usr_..."
            resolved["user_id"] = user_id
            resolved["limit"] = max(self._to_int(config.get("limit")), 1) or 50
            resolved["creator_review_enabled"] = bool(config.get("creator_review_enabled"))
            return resolved

        if job_type in {"world_search", "worlds"}:
            tags = self._csv_items(config.get("tags"))
            notags = self._csv_items(config.get("notags"))
            sort = str(config.get("sort", "popularity")).strip() or "popularity"
            order = str(config.get("order", "descending")).strip() or "descending"
            include_user_ids = {
                str(item).strip()
                for item in config.get("include_user_ids", [])
                if str(item).strip().startswith("usr_")
            }
            include_user_ids_file = config.get("include_user_ids_file")
            if include_user_ids_file:
                include_user_ids.update(
                    item for item in self._load_blacklist(include_user_ids_file) if item.startswith("usr_")
                )
            resolved["type"] = "world_search"
            resolved["search"] = str(config.get("search", "")).strip()
            resolved["tags"] = tags
            resolved["notags"] = notags
            resolved["include_user_ids_file"] = include_user_ids_file
            resolved["include_user_ids"] = sorted(include_user_ids)
            resolved["sort"] = sort
            resolved["order"] = order
            resolved["featured"] = self._optional_bool(config.get("featured"))
            resolved["active"] = bool(self._optional_bool(config.get("active")) or False)
            resolved["release_status"] = self._clean_optional_text(config.get("release_status"))
            resolved["platform"] = self._clean_optional_text(config.get("platform"))
            resolved["limit"] = max(self._to_int(config.get("limit")), 1) or 50
            resolved["creator_review_enabled"] = bool(config.get("creator_review_enabled"))
            return resolved

        resolved["ready"] = False
        resolved["reason"] = f"unsupported job type: {job_type or '(missing)'}"
        return resolved

    # Simplified-Chinese-only characters (never appear in Traditional Chinese)
    _SIMPLIFIED_CHARS: frozenset[str] = frozenset(
        "们动时电话问见这对说东国传边义习爱总为来实发读长带该产给还进让热换变联联应运战种转组农气钱亲请让认样阴语园"
    )

    def _should_exclude_world(
        self,
        world: dict[str, Any],
        *,
        exclude_author_ids: set[str],
        whitelist_author_ids: set[str],
        name_blacklist: list[str] | None = None,
    ) -> bool:
        author_id = str(world.get("author_id") or world.get("authorId") or "").strip()
        if author_id and author_id in whitelist_author_ids:
            return False
        if author_id and author_id in exclude_author_ids:
            return True
        name = str(world.get("name") or "")
        name_lower = name.lower()
        if name_blacklist and any(sub in name_lower for sub in name_blacklist):
            return True
        if any(ch in self._SIMPLIFIED_CHARS for ch in name):
            return True
        has_kana = any("぀" <= ch <= "ヿ" for ch in name)
        has_cjk = any("一" <= ch <= "鿿" for ch in name)
        if has_kana and not has_cjk:
            return True
        return False

    def _load_blacklist(self, value: Any) -> set[str]:
        if not value:
            return set()
        path = Path(str(value))
        if not path.is_absolute():
            path = self.repo_root / path
        if not path.exists():
            return set()
        if path.suffix.lower() == ".txt":
            return {
                line.strip()
                for line in path.read_text(encoding="utf-8").splitlines()
                if line.strip() and not line.strip().startswith("#")
            }
        return set()

    def _load_topic_source_worlds(self, source_value: str) -> list[dict[str, Any]]:
        cleaned = self._clean_optional_text(source_value)
        if not cleaned:
            return []
        candidates = [cleaned]
        if not cleaned.startswith("db:"):
            candidates.append(self._public_db_source_key(cleaned))
        for candidate in candidates:
            try:
                return self.load_worlds(candidate)
            except KeyError:
                continue
        return []

    def _write_blacklist(self, path: Path, entries: set[str]) -> None:
        path.parent.mkdir(parents=True, exist_ok=True)
        payload = "\n".join(sorted(entry for entry in entries if entry)) + ("\n" if entries else "")
        path.write_text(payload, encoding="utf-8")

    def _resolve_job_config_for_blacklist(self, job_key: str) -> dict[str, Any]:
        return self._resolve_job_config_for_list_file(job_key, "blacklist_file", "has no blacklist_file configured")

    def _resolve_job_config_for_list_file(self, job_key: str, field_name: str, missing_message: str) -> dict[str, Any]:
        configs = self._load_job_configs()
        if job_key not in configs:
            raise KeyError(f"Unknown job: {job_key}")
        resolved = self._resolve_job_config(job_key, configs[job_key])
        raw_path = resolved.get(field_name)
        if not raw_path:
            raise ValueError(f"Job {job_key} {missing_message}")
        list_path = Path(str(raw_path))
        if not list_path.is_absolute():
            list_path = self.repo_root / list_path
        resolved["list_path"] = list_path
        resolved["blacklist_path"] = list_path
        return resolved

    def _normalise_db_world(self, world: dict[str, Any]) -> dict[str, Any]:
        world["source"] = self._public_db_source_key(world.get("_db_source_key") or "")
        return world

    def _world_preview(self, world: dict[str, Any]) -> dict[str, Any]:
        return {
            "id": world.get("id"),
            "name": world.get("name") or world.get("id"),
            "author_id": world.get("author_id"),
            "author_name": world.get("author_name"),
            "visits": self._to_int(world.get("visits")),
            "favorites": self._to_int(world.get("favorites")),
            "updated_at": world.get("updated_at"),
            "world_url": world.get("world_url"),
        }

    def _event_world_preview(self, world: dict[str, Any]) -> dict[str, Any]:
        return {
            "world_id": world.get("id"),
            "name": world.get("name") or world.get("id"),
            "author_id": world.get("author_id"),
            "author_name": world.get("author_name"),
            "visits": self._to_int(world.get("visits")),
            "favorites": self._to_int(world.get("favorites")),
            "heat": self._to_int(world.get("heat")),
            "popularity": self._to_int(world.get("popularity")),
            "updated_at": world.get("updated_at"),
            "publication_date": world.get("publication_date"),
            "world_url": world.get("world_url"),
        }

    def _compute_diff_spike(self, item: dict[str, Any]) -> dict[str, Any]:
        latest_visits = self._to_int((item.get("latest") or {}).get("visits"))
        previous_visits = self._to_int((item.get("previous") or {}).get("visits"))
        latest_favorites = self._to_int((item.get("latest") or {}).get("favorites"))
        previous_favorites = self._to_int((item.get("previous") or {}).get("favorites"))
        visits_delta = self._to_int(item.get("visits_delta"))
        favorites_delta = self._to_int(item.get("favorites_delta"))
        visits_growth = (visits_delta / previous_visits) if previous_visits > 0 else None
        favorites_growth = (favorites_delta / previous_favorites) if previous_favorites > 0 else None
        score = (
            (visits_growth * 100 if visits_growth is not None else 25 if visits_delta > 0 else 0)
            + max(0, favorites_delta) * 6
            + max(0, self._to_int(item.get("heat_delta"))) * 4
            + max(0, self._to_int(item.get("popularity_delta"))) * 4
        )
        return {
            "latest_visits": latest_visits,
            "latest_favorites": latest_favorites,
            "visits_delta": visits_delta,
            "favorites_delta": favorites_delta,
            "visits_growth": visits_growth,
            "favorites_growth": favorites_growth,
            "score": score,
        }

    def _is_notable_diff_spike(self, item: dict[str, Any], spike: dict[str, Any]) -> bool:
        visits_delta = self._to_int(spike.get("visits_delta"))
        favorites_delta = self._to_int(spike.get("favorites_delta"))
        visits_growth = spike.get("visits_growth")
        favorites_growth = spike.get("favorites_growth")
        return (
            (visits_delta >= 25 and visits_growth is not None and visits_growth >= 0.45)
            or visits_delta >= 60
            or (favorites_delta >= 6 and favorites_growth is not None and favorites_growth >= 0.35)
            or (self._to_int(item.get("heat_delta")) >= 3 and visits_delta >= 15)
            or (self._to_int(item.get("popularity_delta")) >= 3 and visits_delta >= 15)
        )

    def _signed_number(self, value: Any) -> str:
        number = self._to_int(value)
        return f"{number:+,}"

    def _trend_score(self, trend: dict[str, Any]) -> int:
        momentum_score = self._to_float(trend.get("momentum_score"))
        if momentum_score is not None:
            return int(round(momentum_score))
        return (
            self._to_int(trend.get("visits_delta_7d"))
            + self._to_int(trend.get("visits_delta_30d")) // 2
            + self._to_int(trend.get("favorites_delta_7d")) * 6
        )

    def _freshness_multiplier(self, days_since_publication: int | None) -> float:
        if days_since_publication is None:
            return 0.9
        if days_since_publication <= 3:
            return 2.4
        if days_since_publication <= 7:
            return 2.1
        if days_since_publication <= 14:
            return 1.75
        if days_since_publication <= 30:
            return 1.35
        if days_since_publication <= 60:
            return 1.0
        return 0.7

    def _momentum_score(self, trend: dict[str, Any]) -> float:
        visits_delta_7d = max(0, self._to_int(trend.get("visits_delta_7d")))
        visits_delta_30d = max(0, self._to_int(trend.get("visits_delta_30d")))
        favorites_delta_7d = max(0, self._to_int(trend.get("favorites_delta_7d")))
        visits_growth_7d = max(0.0, self._to_float(trend.get("visits_growth_7d")) or 0.0)
        acceleration = max(0, visits_delta_7d - max(0, self._to_int(trend.get("visits_delta_prev_7d"))))
        return round(
            visits_delta_7d
            + visits_delta_30d * 0.35
            + favorites_delta_7d * 7
            + min(visits_growth_7d, 4.0) * 42
            + acceleration * 0.25,
            2,
        )

    def _breakout_score(self, trend: dict[str, Any]) -> float:
        days_since_publication = trend.get("days_since_publication")
        visits_delta_1d = max(0, self._to_int(trend.get("visits_delta_1d")))
        visits_delta_7d = max(0, self._to_int(trend.get("visits_delta_7d")))
        favorites_delta_7d = max(0, self._to_int(trend.get("favorites_delta_7d")))
        publication_visits_per_day = max(0.0, self._to_float(trend.get("publication_visits_per_day")) or 0.0)
        visits_growth_1d = max(0.0, self._to_float(trend.get("visits_growth_1d")) or 0.0)
        visits_growth_7d = max(0.0, self._to_float(trend.get("visits_growth_7d")) or 0.0)
        favorite_rate = max(0.0, self._to_float(trend.get("favorite_rate")) or 0.0)
        acceleration = max(0, visits_delta_7d - max(0, self._to_int(trend.get("visits_delta_prev_7d"))))
        freshness = self._freshness_multiplier(days_since_publication if isinstance(days_since_publication, int) else None)
        freshness_bonus = 0.0
        if isinstance(days_since_publication, int):
            freshness_bonus = max(0, 14 - min(days_since_publication, 14)) * 4.0
        return round(
            (
                visits_delta_1d * 2.5
                + visits_delta_7d * 1.1
                + favorites_delta_7d * 11
                + publication_visits_per_day * 3.5
                + min(max(visits_growth_1d, visits_growth_7d), 5.0) * 45
                + acceleration * 0.35
                + favorite_rate * 3.0
            )
            * freshness
            + freshness_bonus,
            2,
        )

    def _worth_watching_score(self, trend: dict[str, Any]) -> float:
        favorite_rate = max(0.0, self._to_float(trend.get("favorite_rate")) or 0.0)
        favorites_delta_7d = max(0, self._to_int(trend.get("favorites_delta_7d")))
        visits_delta_7d = max(0, self._to_int(trend.get("visits_delta_7d")))
        since_update_visits_per_day = max(0.0, self._to_float(trend.get("since_update_visits_per_day")) or 0.0)
        heat = max(0, self._to_int(trend.get("heat")))
        popularity = max(0, self._to_int(trend.get("popularity")))
        days_since_publication = trend.get("days_since_publication")
        freshness_bonus = 0.0
        if isinstance(days_since_publication, int) and days_since_publication <= 30:
            freshness_bonus = (31 - max(days_since_publication, 0)) * 1.8
        return round(
            favorite_rate * 16
            + favorites_delta_7d * 9
            + min(visits_delta_7d, 300) * 0.55
            + since_update_visits_per_day * 2.5
            + heat * 4
            + popularity * 3
            + freshness_bonus,
            2,
        )

    def _new_hot_score(self, trend: dict[str, Any]) -> float:
        days_since_publication = trend.get("days_since_publication")
        freshness = self._freshness_multiplier(days_since_publication if isinstance(days_since_publication, int) else None)
        return round(
            (self._breakout_score(trend) * 0.82 + self._worth_watching_score(trend) * 0.18)
            * max(freshness, 1.0),
            2,
        )

    def _rising_now_score(self, trend: dict[str, Any]) -> float:
        visits_delta_1d = max(0, self._to_int(trend.get("visits_delta_1d")))
        visits_delta_7d = max(0, self._to_int(trend.get("visits_delta_7d")))
        favorites_delta_1d = max(0, self._to_int(trend.get("favorites_delta_1d")))
        visits_growth_1d = max(0.0, self._to_float(trend.get("visits_growth_1d")) or 0.0)
        visits_delta_prev_7d = max(0, self._to_int(trend.get("visits_delta_prev_7d")))
        acceleration = max(0, visits_delta_7d - visits_delta_prev_7d)
        days_since_publication = trend.get("days_since_publication")
        freshness_bonus = 0.0
        if isinstance(days_since_publication, int) and days_since_publication <= 21:
            freshness_bonus = (22 - max(days_since_publication, 0)) * 2.5
        return round(
            visits_delta_1d * 3.0
            + visits_delta_7d * 0.6
            + favorites_delta_1d * 14
            + min(visits_growth_1d, 5.0) * 55
            + acceleration * 0.35
            + freshness_bonus,
            2,
        )

    def _dormant_revival_score(self, trend: dict[str, Any]) -> float:
        days_since_publication = trend.get("days_since_publication")
        update_gap_days = self._to_int(trend.get("update_gap_days"))
        visits_delta_7d = max(0, self._to_int(trend.get("visits_delta_7d")))
        visits_delta_1d = max(0, self._to_int(trend.get("visits_delta_1d")))
        since_update_visits_delta = max(0, self._to_int(trend.get("since_update_visits_delta")))
        visits_delta_prev_7d = max(0, self._to_int(trend.get("visits_delta_prev_7d")))
        favorite_rate = max(0.0, self._to_float(trend.get("favorite_rate")) or 0.0)
        is_old_world = isinstance(days_since_publication, int) and days_since_publication >= 45
        is_long_gap_update = update_gap_days >= 120
        has_revival_signal = (
            visits_delta_7d >= 40
            or visits_delta_1d >= 15
            or since_update_visits_delta >= 80
            or trend.get("update_effect_tag") == "ACTIVE"
        )
        if not has_revival_signal or not (is_old_world or is_long_gap_update):
            return 0.0
        age_bonus = min(max(self._to_int(days_since_publication) - 45, 0), 365) * 0.25 if is_old_world else 0.0
        gap_bonus = min(update_gap_days, 720) * 0.08 if is_long_gap_update else 0.0
        acceleration = max(0, visits_delta_7d - visits_delta_prev_7d)
        return round(
            visits_delta_7d * 1.15
            + visits_delta_1d * 2.2
            + since_update_visits_delta * 0.55
            + acceleration * 0.25
            + favorite_rate * 4.0
            + age_bonus
            + gap_bonus,
            2,
        )

    def _creator_momentum_score(self, bucket: dict[str, Any]) -> float:
        recent_7d = max(0, self._to_int(bucket.get("recent_visits_delta_7d")))
        recent_30d = max(0, self._to_int(bucket.get("recent_visits_delta_30d")))
        active_worlds = max(0, self._to_int(bucket.get("active_worlds_30d")))
        breakout_worlds = max(0, self._to_int(bucket.get("breakout_worlds")))
        rising_worlds = max(0, self._to_int(bucket.get("rising_worlds")))
        worth_watching_worlds = max(0, self._to_int(bucket.get("worth_watching_worlds")))
        average_favorite_rate = max(0.0, (bucket.get("favorite_rate_sum") or 0.0) / max(bucket.get("favorite_rate_count") or 1, 1))
        top_share = 0.0
        total_visits = self._to_int(bucket.get("total_visits"))
        top_world_visits = self._to_int(bucket.get("top_world_visits"))
        if total_visits > 0 and top_world_visits > 0:
            top_share = top_world_visits / total_visits
        concentration_penalty = max(0.0, top_share - 0.75) * 180
        return round(
            recent_7d * 1.4
            + recent_30d * 0.45
            + active_worlds * 28
            + breakout_worlds * 120
            + rising_worlds * 90
            + worth_watching_worlds * 75
            + average_favorite_rate * 12
            - concentration_penalty,
            2,
        )

    def _trend_sort_metric(self, world: dict[str, Any], trend: dict[str, Any], field: str) -> float:
        if field == "breakout":
            return self._to_float(trend.get("breakout_score")) or 0.0
        if field == "new_hot":
            return self._to_float(trend.get("new_hot_score")) or 0.0
        if field == "momentum":
            return self._to_float(trend.get("momentum_score")) or 0.0
        if field == "worth_watching":
            return self._to_float(trend.get("worth_watching_score")) or 0.0
        if field == "recent_update":
            return self._update_effectiveness_score(trend)
        if field == "publication_velocity":
            return self._to_float(trend.get("publication_visits_per_day")) or 0.0
        return float(self._sort_value(world, field) or 0.0)

    def _discovery_reason(self, trend: dict[str, Any], *, mode: str) -> str:
        days_since_publication = trend.get("days_since_publication")
        favorite_rate = self._to_float(trend.get("favorite_rate"))
        visits_delta_1d = self._to_int(trend.get("visits_delta_1d"))
        visits_delta_7d = self._to_int(trend.get("visits_delta_7d"))
        publication_visits_per_day = self._to_float(trend.get("publication_visits_per_day"))
        since_update_visits_per_day = self._to_float(trend.get("since_update_visits_per_day"))

        parts: list[str] = []
        if isinstance(days_since_publication, int):
            parts.append(f"{days_since_publication}d old")
        if mode in {"new_hot", "breakout"}:
            if visits_delta_1d > 0:
                parts.append(f"1d {self._signed_number(visits_delta_1d)}")
            if publication_visits_per_day is not None and publication_visits_per_day > 0:
                parts.append(f"{round(publication_visits_per_day, 1)}/day since publish")
        elif mode == "rising_now":
            if visits_delta_1d > 0:
                parts.append(f"1d {self._signed_number(visits_delta_1d)}")
            growth_1d = self._to_float(trend.get("visits_growth_1d"))
            if growth_1d is not None and growth_1d > 0:
                parts.append(f"{round(growth_1d * 100, 1)}% 1d growth")
            if visits_delta_7d > 0:
                parts.append(f"7d {self._signed_number(visits_delta_7d)}")
        elif mode == "momentum":
            if visits_delta_7d > 0:
                parts.append(f"7d {self._signed_number(visits_delta_7d)}")
            growth_7d = self._to_float(trend.get("visits_growth_7d"))
            if growth_7d is not None and growth_7d > 0:
                parts.append(f"{round(growth_7d * 100, 1)}% 7d growth")
        elif mode == "worth_watching":
            if favorite_rate is not None and favorite_rate > 0:
                parts.append(f"{round(favorite_rate, 2)}% fav rate")
            if visits_delta_7d > 0:
                parts.append(f"7d {self._signed_number(visits_delta_7d)}")
            if since_update_visits_per_day is not None and since_update_visits_per_day > 0:
                parts.append(f"{round(since_update_visits_per_day, 1)}/day after update")
        elif mode == "revival":
            update_gap_days = self._to_int(trend.get("update_gap_days"))
            since_update_visits_delta = self._to_int(trend.get("since_update_visits_delta"))
            if update_gap_days > 0:
                parts.append(f"{update_gap_days}d update gap")
            if visits_delta_7d > 0:
                parts.append(f"7d {self._signed_number(visits_delta_7d)}")
            if since_update_visits_delta > 0:
                parts.append(f"post-update {self._signed_number(since_update_visits_delta)}")
        return " / ".join(parts[:3]) or "watch list candidate"

    def _update_effectiveness_score(self, trend: dict[str, Any]) -> float:
        visits_since_update = max(0, self._to_int(trend.get("since_update_visits_delta")))
        favorites_since_update = max(0, self._to_int(trend.get("since_update_favorites_delta")))
        visits_delta_1d = max(0, self._to_int(trend.get("visits_delta_1d")))
        days_since_update = trend.get("days_since_update")
        velocity_divisor = 1
        if isinstance(days_since_update, int) and days_since_update > 0:
            velocity_divisor = min(days_since_update, 30)
        visits_velocity = visits_since_update / velocity_divisor
        burst_bonus = visits_delta_1d * 0.8
        return round(visits_velocity + favorites_since_update * 5 + burst_bonus, 2)

    def _anomaly_ratio(self, current: int | None, previous: int | None, floor: int = 10) -> float | None:
        if current is None:
            return None
        prev_value = self._to_int(previous)
        baseline = max(prev_value, floor)
        if baseline <= 0:
            return None
        return round(self._to_int(current) / baseline, 3)

    def _anomaly_score(self, trend: dict[str, Any]) -> float:
        visits_delta_7d = self._to_int(trend.get("visits_delta_7d"))
        visits_delta_prev_7d = self._to_int(trend.get("visits_delta_prev_7d"))
        visits_delta_1d = self._to_int(trend.get("visits_delta_1d"))
        favorites_delta_7d = self._to_int(trend.get("favorites_delta_7d"))
        ratio = self._anomaly_ratio(visits_delta_7d, visits_delta_prev_7d) or 0.0
        acceleration = max(0, visits_delta_7d - max(0, visits_delta_prev_7d))
        return round(
            acceleration
            + visits_delta_1d * 1.25
            + favorites_delta_7d * 8
            + max(0.0, ratio - 1.0) * 40,
            2,
        )

    def _is_notable_anomaly(self, trend: dict[str, Any]) -> bool:
        visits_delta_7d = self._to_int(trend.get("visits_delta_7d"))
        visits_delta_prev_7d = self._to_int(trend.get("visits_delta_prev_7d"))
        visits_delta_1d = self._to_int(trend.get("visits_delta_1d"))
        favorites_delta_7d = self._to_int(trend.get("favorites_delta_7d"))
        ratio = self._anomaly_ratio(visits_delta_7d, visits_delta_prev_7d) or 0.0
        return (
            (visits_delta_7d >= 80 and ratio >= 1.6)
            or visits_delta_7d >= 160
            or (visits_delta_1d >= 30 and visits_delta_7d >= 40)
            or (favorites_delta_7d >= 10 and visits_delta_7d >= 40)
        )

    def _is_starriver_scope(
        self,
        worlds: list[dict[str, Any]],
        *,
        topic_key: str | None = None,
        source: str | None = None,
    ) -> bool:
        if topic_key == "personal":
            return True
        if source == "db:job:starriver":
            return True
        author_ids = {str(world.get("author_id") or "").strip() for world in worlds if world.get("author_id")}
        return author_ids == {"usr_0673194d-712d-4b5d-8167-1f03ed3233cb"}

    def _build_world_trend_metrics(self, world: dict[str, Any], history_entries: list[dict[str, Any]]) -> dict[str, Any]:
        entries = sorted(
            [entry for entry in history_entries if entry.get("timestamp") is not None],
            key=lambda item: item.get("timestamp") or 0,
        )
        latest_visits = self._to_int(world.get("visits"))
        latest_favorites = self._to_int(world.get("favorites"))
        favorite_rate = None
        if latest_visits > 0:
            favorite_rate = round((latest_favorites / latest_visits) * 100, 2)

        updated_at = _parse_date(world.get("updated_at"))
        publication_date = _parse_date(world.get("publication_date"))
        now = dt.datetime.now(dt.timezone.utc)
        days_since_update = (now - updated_at).days if updated_at else None
        days_since_publication = (now - publication_date).days if publication_date else None

        def metric_at_or_before(target_ts: int, field: str) -> int | None:
            candidate = None
            for entry in entries:
                timestamp = entry.get("timestamp") or 0
                if timestamp <= target_ts:
                    candidate = entry
                else:
                    break
            if candidate is None:
                return None
            return self._to_optional_int(candidate.get(field))

        latest_ts = entries[-1]["timestamp"] if entries else None
        visits_delta_7d = None
        visits_delta_prev_7d = None
        visits_delta_30d = None
        visits_delta_1d = None
        visits_growth_1d = None
        visits_growth_7d = None
        favorites_delta_7d = None
        favorites_delta_1d = None
        favorites_growth_7d = None
        if latest_ts is not None:
            current_visits = self._to_int(entries[-1].get("visits"))
            current_favorites = self._to_int(entries[-1].get("favorites"))
            visits_1d = metric_at_or_before(latest_ts - 1 * 86400, "visits")
            visits_7d = metric_at_or_before(latest_ts - 7 * 86400, "visits")
            visits_14d = metric_at_or_before(latest_ts - 14 * 86400, "visits")
            visits_30d = metric_at_or_before(latest_ts - 30 * 86400, "visits")
            favorites_1d = metric_at_or_before(latest_ts - 1 * 86400, "favorites")
            favorites_7d = metric_at_or_before(latest_ts - 7 * 86400, "favorites")
            if visits_1d is not None:
                visits_delta_1d = current_visits - visits_1d
                visits_growth_1d = ((current_visits - visits_1d) / visits_1d) if visits_1d > 0 else None
            if visits_7d is not None:
                visits_delta_7d = current_visits - visits_7d
                visits_growth_7d = ((current_visits - visits_7d) / visits_7d) if visits_7d > 0 else None
            if visits_7d is not None and visits_14d is not None:
                visits_delta_prev_7d = visits_7d - visits_14d
            if visits_30d is not None:
                visits_delta_30d = current_visits - visits_30d
            if favorites_1d is not None:
                favorites_delta_1d = current_favorites - favorites_1d
            if favorites_7d is not None:
                favorites_delta_7d = current_favorites - favorites_7d
                favorites_growth_7d = ((current_favorites - favorites_7d) / favorites_7d) if favorites_7d > 0 else None

        update_gap_days = None
        distinct_updates = []
        for entry in entries:
            parsed = _parse_date(entry.get("updated_at"))
            if not parsed:
                continue
            if not distinct_updates or distinct_updates[-1] != parsed:
                distinct_updates.append(parsed)
        if updated_at and distinct_updates:
            latest_known = distinct_updates[-1]
            if abs((latest_known - updated_at).total_seconds()) < 86400 and len(distinct_updates) > 1:
                update_gap_days = (distinct_updates[-1] - distinct_updates[-2]).days

        since_update_visits_delta = None
        since_update_favorites_delta = None
        since_update_visits_per_day = None
        since_update_favorites_per_day = None
        publication_visits_per_day = None
        if updated_at and latest_ts is not None:
            baseline_visits = metric_at_or_before(int(updated_at.timestamp()), "visits")
            baseline_favorites = metric_at_or_before(int(updated_at.timestamp()), "favorites")
            if baseline_visits is not None:
                since_update_visits_delta = self._to_int(entries[-1].get("visits")) - baseline_visits
            if baseline_favorites is not None:
                since_update_favorites_delta = self._to_int(entries[-1].get("favorites")) - baseline_favorites
            elapsed_days = max((entries[-1]["timestamp"] - int(updated_at.timestamp())) / 86400, 0)
            if elapsed_days > 0 and since_update_visits_delta is not None:
                since_update_visits_per_day = round(since_update_visits_delta / elapsed_days, 2)
            if elapsed_days > 0 and since_update_favorites_delta is not None:
                since_update_favorites_per_day = round(since_update_favorites_delta / elapsed_days, 2)
        if publication_date:
            age_days = max((now - publication_date).total_seconds() / 86400, 0)
            if latest_visits > 0:
                publication_visits_per_day = round(latest_visits / max(age_days, 1.0), 2)

        tags: list[str] = []
        if favorite_rate is not None and favorite_rate >= 8 and latest_favorites >= 25:
            tags.append("LOVED WORLD")
        if visits_delta_7d is not None and visits_delta_prev_7d is not None:
            if visits_delta_prev_7d >= 100 and visits_delta_7d <= int(visits_delta_prev_7d * 0.35):
                tags.append("INACTIVE")
            elif visits_delta_7d >= max(200, int(visits_delta_prev_7d * 2)):
                tags.append("ACTIVE")
            elif (
                visits_delta_30d is not None
                and visits_delta_30d >= 200
                and visits_delta_7d >= 40
                and visits_delta_prev_7d > 0
                and 0.6 <= (visits_delta_7d / visits_delta_prev_7d) <= 1.4
            ):
                tags.append("STEADY FLOW")
        if days_since_update is not None and days_since_update <= 60 and since_update_visits_delta is not None and since_update_visits_delta <= 60:
            tags.append("SILENCE UPDATE")
        if days_since_update is not None and days_since_update <= 30 and update_gap_days is not None and update_gap_days > 365:
            tags.append("REVIVE")

        update_effect_tag = None
        if days_since_update is not None and days_since_update <= 60:
            if since_update_visits_delta is not None and since_update_visits_delta <= 60:
                update_effect_tag = "SILENCE UPDATE"
            elif since_update_visits_delta is not None and since_update_visits_delta >= 200:
                update_effect_tag = "ACTIVE"
            elif since_update_visits_delta is not None and since_update_visits_delta >= 80:
                update_effect_tag = "STEADY FLOW"

        trend_payload = {
            "favorite_rate": favorite_rate,
            "visits_delta_1d": visits_delta_1d,
            "days_since_update": days_since_update,
            "days_since_publication": days_since_publication,
            "visits_delta_7d": visits_delta_7d,
            "visits_delta_prev_7d": visits_delta_prev_7d,
            "visits_delta_30d": visits_delta_30d,
            "visits_growth_1d": round(visits_growth_1d, 4) if visits_growth_1d is not None else None,
            "visits_growth_7d": round(visits_growth_7d, 4) if visits_growth_7d is not None else None,
            "favorites_delta_1d": favorites_delta_1d,
            "favorites_delta_7d": favorites_delta_7d,
            "favorites_growth_7d": round(favorites_growth_7d, 4) if favorites_growth_7d is not None else None,
            "since_update_visits_delta": since_update_visits_delta,
            "since_update_favorites_delta": since_update_favorites_delta,
            "since_update_visits_per_day": since_update_visits_per_day,
            "since_update_favorites_per_day": since_update_favorites_per_day,
            "publication_visits_per_day": publication_visits_per_day,
            "update_gap_days": update_gap_days,
            "update_effect_tag": update_effect_tag,
            "tags": tags[:3],
            "heat": self._to_optional_int(world.get("heat")),
            "popularity": self._to_optional_int(world.get("popularity")),
        }
        trend_payload["momentum_score"] = self._momentum_score(trend_payload)
        trend_payload["breakout_score"] = self._breakout_score(trend_payload)
        trend_payload["worth_watching_score"] = self._worth_watching_score(trend_payload)
        trend_payload["new_hot_score"] = self._new_hot_score(trend_payload)
        trend_payload["update_effectiveness_score"] = self._update_effectiveness_score(trend_payload)

        return trend_payload

    def _build_signal_analysis(self, rows: list[dict[str, Any]], *, limit: int = 12) -> dict[str, Any]:
        usable_rows = [row for row in rows if row.get("id")]
        heat_rows = [row for row in usable_rows if row.get("heat") is not None]
        popularity_rows = [row for row in usable_rows if row.get("popularity") is not None]
        favorite_rate_rows = [row for row in usable_rows if row.get("favorite_rate") is not None]

        correlations = [
            self._build_signal_correlation(usable_rows, "heat", "visits", "Heat vs Visits"),
            self._build_signal_correlation(usable_rows, "popularity", "visits", "Popularity vs Visits"),
            self._build_signal_correlation(usable_rows, "heat", "favorites", "Heat vs Favorites"),
            self._build_signal_correlation(usable_rows, "popularity", "favorites", "Popularity vs Favorites"),
            self._build_signal_correlation(usable_rows, "heat", "favorite_rate", "Heat vs Favorite Rate"),
            self._build_signal_correlation(usable_rows, "popularity", "favorite_rate", "Popularity vs Favorite Rate"),
            self._build_signal_correlation(usable_rows, "heat", "popularity", "Heat vs Popularity"),
        ]

        charts = [
            self._build_signal_chart(usable_rows, x_metric="heat", y_metric="visits", title="Heat vs Visits"),
            self._build_signal_chart(usable_rows, x_metric="popularity", y_metric="visits", title="Popularity vs Visits"),
            self._build_signal_chart(usable_rows, x_metric="heat", y_metric="favorite_rate", title="Heat vs Favorite Rate"),
            self._build_signal_chart(usable_rows, x_metric="popularity", y_metric="favorite_rate", title="Popularity vs Favorite Rate"),
        ]

        heat_leaders = sorted(
            heat_rows,
            key=lambda item: (
                self._to_int(item.get("heat")),
                self._to_int(item.get("visits")),
            ),
            reverse=True,
        )[:limit]
        popularity_leaders = sorted(
            popularity_rows,
            key=lambda item: (
                self._to_int(item.get("popularity")),
                self._to_int(item.get("visits")),
            ),
            reverse=True,
        )[:limit]
        bucketed_rows: dict[str, list[dict[str, Any]]] = {}
        for row in usable_rows:
            visits = self._to_int(row.get("visits"))
            if visits <= 0:
                continue
            bucketed_rows.setdefault(self._signal_visit_bucket(visits), []).append(row)
        efficiency_rows = []
        for bucket, bucket_rows in bucketed_rows.items():
            heat_values = [self._to_int(item.get("heat")) for item in bucket_rows if item.get("heat") is not None]
            popularity_values = [self._to_int(item.get("popularity")) for item in bucket_rows if item.get("popularity") is not None]
            for row in bucket_rows:
                visits = self._to_int(row.get("visits"))
                heat = self._to_optional_int(row.get("heat"))
                popularity = self._to_optional_int(row.get("popularity"))
                heat_percentile = self._percentile_rank(heat_values, heat) if heat is not None else None
                popularity_percentile = self._percentile_rank(popularity_values, popularity) if popularity is not None else None
                confidence_weight = round(visits / (visits + 500), 4)
                weighted_score = self._weighted_signal_efficiency_score(
                    heat_percentile=heat_percentile,
                    popularity_percentile=popularity_percentile,
                    confidence_weight=confidence_weight,
                )
                if weighted_score is None:
                    continue
                efficiency_rows.append(
                    {
                        **row,
                        "visit_bucket": bucket,
                        "heat_percentile": round(heat_percentile * 100, 1) if heat_percentile is not None else None,
                        "popularity_percentile": round(popularity_percentile * 100, 1) if popularity_percentile is not None else None,
                        "confidence_weight": round(confidence_weight * 100, 1),
                        "signal_efficiency_score": round(weighted_score, 2),
                    }
                )
        signal_efficiency = sorted(
            efficiency_rows,
            key=lambda item: (
                float(item.get("signal_efficiency_score") or 0.0),
                self._to_int(item.get("favorites")),
                self._to_int(item.get("visits")),
            ),
            reverse=True,
        )[:limit]

        return {
            "summary": {
                "world_count": len(usable_rows),
                "heat_count": len(heat_rows),
                "popularity_count": len(popularity_rows),
                "favorite_rate_count": len(favorite_rate_rows),
                "avg_heat": self._average([row.get("heat") for row in heat_rows]),
                "avg_popularity": self._average([row.get("popularity") for row in popularity_rows]),
                "avg_favorite_rate": self._average([row.get("favorite_rate") for row in favorite_rate_rows]),
                "median_heat": self._median([row.get("heat") for row in heat_rows]),
                "median_popularity": self._median([row.get("popularity") for row in popularity_rows]),
            },
            "correlations": [item for item in correlations if item["sample_size"] >= 3],
            "charts": charts,
            "leaderboards": {
                "heat_leaders": heat_leaders,
                "popularity_leaders": popularity_leaders,
                "signal_efficiency": signal_efficiency,
            },
        }

    def _signal_visit_bucket(self, visits: int) -> str:
        if visits < 100:
            return "0-99"
        if visits < 300:
            return "100-299"
        if visits < 1000:
            return "300-999"
        if visits < 3000:
            return "1k-2.9k"
        if visits < 10000:
            return "3k-9.9k"
        return "10k+"

    def _percentile_rank(self, values: list[int], target: int | None) -> float | None:
        if target is None:
            return None
        numbers = sorted(int(value) for value in values if value is not None)
        if not numbers:
            return None
        less_or_equal = sum(1 for value in numbers if value <= target)
        return less_or_equal / len(numbers)

    def _weighted_signal_efficiency_score(
        self,
        *,
        heat_percentile: float | None,
        popularity_percentile: float | None,
        confidence_weight: float,
    ) -> float | None:
        weighted_sum = 0.0
        total_weight = 0.0
        if heat_percentile is not None:
            weighted_sum += heat_percentile * 0.55
            total_weight += 0.55
        if popularity_percentile is not None:
            weighted_sum += popularity_percentile * 0.45
            total_weight += 0.45
        if total_weight <= 0:
            return None
        return (weighted_sum / total_weight) * confidence_weight * 100

    def _build_signal_chart(
        self,
        rows: list[dict[str, Any]],
        *,
        x_metric: str,
        y_metric: str,
        title: str,
    ) -> dict[str, Any]:
        points = []
        for row in rows:
            x_value = row.get(x_metric)
            y_value = row.get(y_metric)
            if x_value is None or y_value is None:
                continue
            points.append(
                {
                    "id": row.get("id"),
                    "name": row.get("name"),
                    "author_name": row.get("author_name"),
                    "x": float(x_value),
                    "y": float(y_value),
                    "visits": row.get("visits"),
                    "favorites": row.get("favorites"),
                    "heat": row.get("heat"),
                    "popularity": row.get("popularity"),
                    "favorite_rate": row.get("favorite_rate"),
                    "world_url": row.get("world_url"),
                }
            )
        points.sort(key=lambda item: (item["y"], item["x"]), reverse=True)
        return {
            "key": f"{x_metric}_vs_{y_metric}",
            "title": title,
            "x_metric": x_metric,
            "y_metric": y_metric,
            "x_label": x_metric.replace("_", " ").title(),
            "y_label": y_metric.replace("_", " ").title(),
            "points": points[:120],
            "sample_size": len(points),
        }

    def _build_signal_correlation(
        self,
        rows: list[dict[str, Any]],
        left_key: str,
        right_key: str,
        label: str,
    ) -> dict[str, Any]:
        pairs: list[tuple[float, float]] = []
        for row in rows:
            left = row.get(left_key)
            right = row.get(right_key)
            if left is None or right is None:
                continue
            pairs.append((float(left), float(right)))
        coefficient = self._pearson_correlation(pairs)
        return {
            "key": f"{left_key}_vs_{right_key}",
            "label": label,
            "left_key": left_key,
            "right_key": right_key,
            "sample_size": len(pairs),
            "coefficient": coefficient,
            "strength": self._describe_correlation(coefficient),
        }

    def _average(self, values: list[Any]) -> float | None:
        numbers = [float(value) for value in values if value is not None]
        if not numbers:
            return None
        return round(sum(numbers) / len(numbers), 2)

    def _median(self, values: list[Any]) -> float | None:
        numbers = sorted(float(value) for value in values if value is not None)
        if not numbers:
            return None
        middle = len(numbers) // 2
        if len(numbers) % 2:
            return round(numbers[middle], 2)
        return round((numbers[middle - 1] + numbers[middle]) / 2, 2)

    def _pearson_correlation(self, pairs: list[tuple[float, float]]) -> float | None:
        if len(pairs) < 3:
            return None
        xs = [pair[0] for pair in pairs]
        ys = [pair[1] for pair in pairs]
        mean_x = sum(xs) / len(xs)
        mean_y = sum(ys) / len(ys)
        numerator = sum((x - mean_x) * (y - mean_y) for x, y in pairs)
        denom_x = sum((x - mean_x) ** 2 for x in xs) ** 0.5
        denom_y = sum((y - mean_y) ** 2 for y in ys) ** 0.5
        if denom_x <= 0 or denom_y <= 0:
            return None
        return round(numerator / (denom_x * denom_y), 3)

    def _describe_correlation(self, coefficient: float | None) -> str:
        if coefficient is None:
            return "insufficient"
        magnitude = abs(coefficient)
        if magnitude < 0.2:
            return "weak"
        if magnitude < 0.45:
            return "light"
        if magnitude < 0.7:
            return "moderate"
        return "strong"

    def _load_workbook_worlds(self, path: Path, source: str) -> list[dict[str, Any]]:
        if not path.exists():
            return []
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        worlds = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or not any(value not in (None, "") for value in row):
                continue
            worlds.append(self._normalise_workbook_row(row, source))
        return worlds

    def _normalise_api_world(self, world: dict[str, Any], source: str) -> dict[str, Any]:
        publication_date = self._clean_optional_text(world.get("publicationDate"))
        updated_at = self._clean_optional_text(world.get("updated_at"))
        created_at = self._clean_optional_text(world.get("created_at"))
        labs_publication_date = self._clean_optional_text(world.get("labsPublicationDate"))
        visits = self._to_optional_int(world.get("visits"))
        favorites = self._to_optional_int(world.get("favorites"))
        capacity = self._to_optional_int(world.get("capacity"))

        world_id = self._clean_optional_text(world.get("id")) or self._clean_optional_text(world.get("worldId"))
        tags = [tag for tag in world.get("tags", []) if isinstance(tag, str)]
        portal_links = self._normalise_portal_links(world.get("portal_links"))

        return {
            "source": source,
            "id": world_id,
            "name": self._clean_optional_text(world.get("name")),
            "description": self._clean_optional_text(world.get("description")),
            "author_id": self._clean_optional_text(world.get("authorId")),
            "author_name": self._clean_optional_text(world.get("authorName")),
            "capacity": capacity,
            "visits": visits,
            "favorites": favorites,
            "heat": self._to_optional_int(world.get("heat")),
            "popularity": self._to_optional_int(world.get("popularity")),
            "created_at": created_at,
            "updated_at": updated_at,
            "publication_date": publication_date,
            "labs_publication_date": labs_publication_date,
            "release_status": self._clean_optional_text(world.get("releaseStatus")),
            "image_url": self._clean_optional_text(world.get("imageUrl")),
            "thumbnail_url": self._clean_optional_text(world.get("thumbnailImageUrl")),
            "tags": tags,
            "portal_links": portal_links,
            "world_url": self._clean_optional_text(world.get("worldUrl"))
            or (f"https://vrchat.com/home/world/{world_id}" if world_id else None),
            "metrics": self._calculate_metrics_for_world(
                {
                    "visits": visits,
                    "favorites": favorites,
                    "updated_at": updated_at,
                    "publication_date": publication_date,
                    "labs_publication_date": labs_publication_date,
                }
            ),
        }

    def _normalise_workbook_row(self, row: tuple[Any, ...], source: str) -> dict[str, Any]:
        values = list(row[:15])
        while len(values) < 15:
            values.append(None)
        world_id = self._clean_optional_text(values[2])
        release_or_days = self._clean_optional_text(values[13])
        release_status = None
        days_since_publication = None
        if release_or_days:
            lowered = release_or_days.casefold()
            if lowered in {"public", "private", "hidden", "labs"}:
                release_status = release_or_days
            else:
                days_since_publication = self._to_int(release_or_days)
        return {
            "source": source,
            "id": world_id,
            "name": self._clean_optional_text(values[1]),
            "author_id": None,
            "author_name": None,
            "capacity": self._to_optional_int(values[6]),
            "visits": self._to_optional_int(values[5]),
            "favorites": self._to_optional_int(values[7]),
            "heat": self._to_optional_int(values[8]),
            "popularity": self._to_optional_int(values[9]),
            "created_at": None,
            "updated_at": self._clean_optional_text(values[4]),
            "publication_date": self._clean_optional_text(values[3]),
            "labs_publication_date": None,
            "release_status": release_status,
            "image_url": None,
            "thumbnail_url": None,
            "tags": [],
            "fetched_at": self._clean_optional_text(values[0]),
            "world_url": f"https://vrchat.com/home/world/{world_id}" if world_id else None,
            "metrics": {
                "favorite_rate": self._parse_ratio(values[11]),
                "labs_to_publication_days": self._to_int(values[10]),
                "days_since_update": self._to_int(values[12]),
                "days_since_publication": days_since_publication,
                "visits_per_day": self._to_float(values[14]),
            },
        }

    def _normalise_history_entry(self, world_id: str, entry: dict[str, Any], origin: str) -> dict[str, Any]:
        timestamp = entry.get("timestamp")
        timestamp_int = int(timestamp) if isinstance(timestamp, (int, float, str)) and str(timestamp).isdigit() else None
        return {
            "world_id": world_id,
            "origin": origin,
            "timestamp": timestamp_int,
            "iso_time": dt.datetime.fromtimestamp(timestamp_int, dt.timezone.utc).isoformat() if timestamp_int else None,
            "name": self._clean_optional_text(entry.get("name")),
            "created_at": self._clean_optional_text(entry.get("created_at")),
            "updated_at": self._clean_optional_text(entry.get("updated_at")),
            "publication_date": self._clean_optional_text(entry.get("publicationDate")),
            "labs_publication_date": self._clean_optional_text(entry.get("labsPublicationDate")),
            "visits": self._to_optional_int(entry.get("visits")),
            "favorites": self._to_optional_int(entry.get("favorites")),
            "heat": self._to_optional_int(entry.get("heat")),
            "popularity": self._to_optional_int(entry.get("popularity")),
        }

    def _normalise_db_history_entry(self, world_id: str, entry: dict[str, Any]) -> dict[str, Any]:
        fetched_at = entry.get("fetched_at")
        fetched_dt = _parse_date(fetched_at)
        timestamp_int = int(fetched_dt.timestamp()) if fetched_dt else None
        return {
            "world_id": world_id,
            "origin": "db",
            "timestamp": timestamp_int,
            "iso_time": fetched_at,
            "name": self._clean_optional_text(entry.get("name")),
            "created_at": self._clean_optional_text(entry.get("created_at")),
            "updated_at": self._clean_optional_text(entry.get("updated_at")),
            "publication_date": self._clean_optional_text(entry.get("publication_date")),
            "labs_publication_date": self._clean_optional_text(entry.get("labs_publication_date")),
            "visits": self._to_optional_int(entry.get("visits")),
            "favorites": self._to_optional_int(entry.get("favorites")),
            "heat": self._to_optional_int(entry.get("heat")),
            "popularity": self._to_optional_int(entry.get("popularity")),
        }

    def _dedupe_worlds(self, worlds: list[dict[str, Any]]) -> list[dict[str, Any]]:
        unique: dict[str, dict[str, Any]] = {}
        passthrough: list[dict[str, Any]] = []
        for world in worlds:
            world_id = world.get("id")
            if not world_id:
                passthrough.append(world)
                continue
            current = unique.get(world_id)
            if current is None:
                unique[world_id] = self._clone_world(world)
                continue
            unique[world_id] = self._merge_world_records(current, world)
        return list(unique.values()) + passthrough

    def _dedupe_raw_world_payloads(self, worlds: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], int]:
        unique: dict[str, dict[str, Any]] = {}
        passthrough: list[dict[str, Any]] = []
        merged_duplicates = 0
        for world in worlds:
            world_id = self._clean_optional_text(world.get("id")) or self._clean_optional_text(world.get("worldId"))
            if not world_id:
                passthrough.append(dict(world))
                continue
            candidate = dict(world)
            candidate["id"] = candidate.get("id") or world_id
            current = unique.get(world_id)
            if current is None:
                unique[world_id] = candidate
                continue
            unique[world_id] = self._merge_raw_world_payloads(current, candidate)
            merged_duplicates += 1
        return list(unique.values()) + passthrough, merged_duplicates

    def _merge_raw_world_payloads(self, left: dict[str, Any], right: dict[str, Any]) -> dict[str, Any]:
        merged = dict(left)

        for field in ("visits", "favorites", "heat", "popularity", "capacity"):
            merged[field] = self._prefer_higher_number(left.get(field), right.get(field))

        merged["id"] = self._clean_optional_text(left.get("id")) or self._clean_optional_text(right.get("id"))
        merged["worldId"] = (
            self._clean_optional_text(left.get("worldId"))
            or self._clean_optional_text(right.get("worldId"))
            or merged.get("id")
        )
        merged["updated_at"] = self._pick_latest_date(left.get("updated_at"), right.get("updated_at"))
        merged["created_at"] = self._pick_earliest_date(left.get("created_at"), right.get("created_at"))
        merged["publicationDate"] = self._pick_earliest_date(
            left.get("publicationDate"),
            right.get("publicationDate"),
        )
        merged["labsPublicationDate"] = self._pick_earliest_date(
            left.get("labsPublicationDate"),
            right.get("labsPublicationDate"),
        )
        merged["releaseStatus"] = self._pick_release_status(
            left.get("releaseStatus"),
            right.get("releaseStatus"),
        )

        for field in ("name", "authorId", "authorName", "imageUrl", "thumbnailImageUrl", "worldUrl"):
            merged[field] = self._pick_preferred_text(left.get(field), right.get(field))

        merged["tags"] = self._merge_tags(left.get("tags"), right.get("tags"))
        return merged

    def _clone_world(self, world: dict[str, Any]) -> dict[str, Any]:
        cloned = dict(world)
        cloned["metrics"] = dict(world.get("metrics", {}))
        cloned["tags"] = list(world.get("tags", []))
        return cloned

    def _merge_world_records(self, left: dict[str, Any], right: dict[str, Any]) -> dict[str, Any]:
        primary, secondary = self._pick_primary_world(left, right)
        merged = self._clone_world(primary)
        secondary_metrics = secondary.get("metrics", {})
        primary_metrics = primary.get("metrics", {})

        for field in ("visits", "favorites", "heat", "popularity", "capacity"):
            merged[field] = self._prefer_higher_number(primary.get(field), secondary.get(field))

        merged["fetched_at"] = self._pick_latest_date(primary.get("fetched_at"), secondary.get("fetched_at"))
        merged["updated_at"] = self._pick_latest_date(primary.get("updated_at"), secondary.get("updated_at"))
        merged["created_at"] = self._pick_earliest_date(primary.get("created_at"), secondary.get("created_at"))
        merged["publication_date"] = self._pick_earliest_date(
            primary.get("publication_date"),
            secondary.get("publication_date"),
        )
        merged["labs_publication_date"] = self._pick_earliest_date(
            primary.get("labs_publication_date"),
            secondary.get("labs_publication_date"),
        )
        merged["release_status"] = self._pick_release_status(
            primary.get("release_status"),
            secondary.get("release_status"),
        )

        for field in ("name", "author_id", "author_name", "image_url", "thumbnail_url", "world_url"):
            merged[field] = self._pick_preferred_text(primary.get(field), secondary.get(field))

        merged["tags"] = self._merge_tags(primary.get("tags"), secondary.get("tags"))
        merged["metrics"] = self._build_merged_metrics(
            merged,
            primary_metrics if isinstance(primary_metrics, dict) else {},
            secondary_metrics if isinstance(secondary_metrics, dict) else {},
        )
        return merged

    def _pick_primary_world(self, left: dict[str, Any], right: dict[str, Any]) -> tuple[dict[str, Any], dict[str, Any]]:
        if self._dedupe_score(right) > self._dedupe_score(left):
            return self._clone_world(right), self._clone_world(left)
        return self._clone_world(left), self._clone_world(right)

    def _dedupe_score(self, world: dict[str, Any]) -> tuple[int, int, int, int, int, int, int]:
        completeness = sum(
            1
            for field in (
                "name",
                "author_id",
                "author_name",
                "created_at",
                "updated_at",
                "publication_date",
                "labs_publication_date",
                "release_status",
                "image_url",
                "thumbnail_url",
                "world_url",
            )
            if self._clean_optional_text(world.get(field))
        )
        completeness += 1 if world.get("tags") else 0
        metrics = world.get("metrics", {}) if isinstance(world.get("metrics"), dict) else {}
        completeness += sum(
            1
            for key in ("favorite_rate", "labs_to_publication_days", "days_since_update", "days_since_publication", "visits_per_day")
            if metrics.get(key) not in (None, "")
        )
        signal_count = sum(
            1
            for field in ("visits", "favorites", "heat", "popularity", "capacity")
            if self._to_int(world.get(field)) > 0
        )
        fetched = self._date_score(world.get("fetched_at"))
        updated = self._date_score(world.get("updated_at"))
        publication = self._date_score(world.get("publication_date"))
        visits = self._to_int(world.get("visits"))
        return (completeness, signal_count, visits, updated, publication, fetched, self._to_int(world.get("favorites")))

    def _prefer_higher_number(self, first: Any, second: Any) -> int | None:
        first_value = self._to_optional_int(first)
        second_value = self._to_optional_int(second)
        if first_value is None:
            return second_value
        if second_value is None:
            return first_value
        return max(first_value, second_value)

    def _pick_latest_date(self, first: Any, second: Any) -> str | None:
        first_text = self._clean_optional_text(first)
        second_text = self._clean_optional_text(second)
        if not first_text:
            return second_text
        if not second_text:
            return first_text
        if self._date_score(second_text) > self._date_score(first_text):
            return second_text
        return first_text

    def _pick_earliest_date(self, first: Any, second: Any) -> str | None:
        first_text = self._clean_optional_text(first)
        second_text = self._clean_optional_text(second)
        if not first_text:
            return second_text
        if not second_text:
            return first_text
        first_score = self._date_score(first_text)
        second_score = self._date_score(second_text)
        if first_score == 0:
            return second_text
        if second_score == 0:
            return first_text
        if second_score < first_score:
            return second_text
        return first_text

    def _pick_release_status(self, first: Any, second: Any) -> str | None:
        valid = {"public", "private", "hidden", "labs"}
        first_text = self._clean_optional_text(first)
        if first_text and first_text.casefold() in valid:
            return first_text
        second_text = self._clean_optional_text(second)
        if second_text and second_text.casefold() in valid:
            return second_text
        return first_text or second_text

    def _pick_preferred_text(self, first: Any, second: Any) -> str | None:
        first_text = self._clean_optional_text(first)
        second_text = self._clean_optional_text(second)
        if not first_text:
            return second_text
        return first_text

    def _merge_tags(self, first: Any, second: Any) -> list[str]:
        merged: list[str] = []
        for value in (first or []):
            if isinstance(value, str) and value not in merged:
                merged.append(value)
        for value in (second or []):
            if isinstance(value, str) and value not in merged:
                merged.append(value)
        return merged

    def _build_merged_metrics(
        self,
        world: dict[str, Any],
        first: dict[str, Any],
        second: dict[str, Any],
    ) -> dict[str, Any]:
        metrics = self._calculate_metrics_for_world(world)
        days_since_publication = metrics.get("days_since_publication")
        labs_to_publication_days = metrics.get("labs_to_publication_days")
        days_since_update = metrics.get("days_since_update")
        visits_per_day = metrics.get("visits_per_day")
        favorite_rate = metrics.get("favorite_rate")

        explicit_days_since_publication = self._pick_numeric_metric(
            first.get("days_since_publication"),
            second.get("days_since_publication"),
        )
        if explicit_days_since_publication is not None:
            days_since_publication = explicit_days_since_publication

        explicit_labs_days = self._pick_numeric_metric(
            first.get("labs_to_publication_days"),
            second.get("labs_to_publication_days"),
        )
        if explicit_labs_days is not None and labs_to_publication_days is None:
            labs_to_publication_days = explicit_labs_days

        explicit_days_since_update = self._pick_numeric_metric(
            first.get("days_since_update"),
            second.get("days_since_update"),
        )
        if explicit_days_since_update is not None and days_since_update is None:
            days_since_update = explicit_days_since_update

        explicit_visits_per_day = self._pick_float_metric(
            first.get("visits_per_day"),
            second.get("visits_per_day"),
        )
        if explicit_visits_per_day is not None and visits_per_day is None:
            visits_per_day = explicit_visits_per_day

        return {
            "favorite_rate": favorite_rate,
            "labs_to_publication_days": labs_to_publication_days,
            "days_since_update": days_since_update,
            "days_since_publication": days_since_publication,
            "visits_per_day": visits_per_day,
        }

    def _pick_numeric_metric(self, first: Any, second: Any) -> int | None:
        values = [self._to_int(value) for value in (first, second) if value not in (None, "")]
        return max(values) if values else None

    def _pick_float_metric(self, first: Any, second: Any) -> float | None:
        values = [self._to_float(value) for value in (first, second)]
        values = [value for value in values if value is not None]
        return max(values) if values else None

    def _sort_value(self, world: dict[str, Any], field: str) -> Any:
        if field == "name":
            return (world.get("name") or "").casefold()
        if field in {"visits", "favorites", "heat", "popularity", "capacity"}:
            return self._to_int(world.get(field))
        if field == "updated":
            return self._date_score(world.get("updated_at"))
        if field == "publication":
            return self._date_score(world.get("publication_date"))
        if field == "fetched":
            return self._date_score(world.get("fetched_at"))
        return self._to_int(world.get("visits"))

    def _sort_worlds(
        self,
        worlds: list[dict[str, Any]],
        *,
        sort: str,
        direction: str,
        history: dict[str, list[dict[str, Any]]] | None = None,
    ) -> list[dict[str, Any]]:
        reverse = direction != "asc"
        trend_sort_fields = {"breakout", "new_hot", "momentum", "worth_watching", "recent_update", "publication_velocity"}
        if sort not in trend_sort_fields:
            items = list(worlds)
            items.sort(key=lambda item: self._sort_value(item, sort), reverse=reverse)
            return items

        history = history if history is not None else self.load_history()
        scored_worlds = []
        for world in worlds:
            world_id = world.get("id")
            trend = self._build_world_trend_metrics(world, history.get(world_id, [])) if world_id else {}
            score = self._trend_sort_metric(world, trend, sort)
            scored_worlds.append(
                (
                    score,
                    self._date_score(world.get("publication_date") or world.get("updated_at") or world.get("fetched_at")),
                    self._to_int(world.get("visits")),
                    world,
                )
            )
        scored_worlds.sort(key=lambda item: (item[0], item[1], item[2]), reverse=reverse)
        return [item[3] for item in scored_worlds]

    def _read_daily_stats_xlsx(self, path: Path) -> list[dict[str, Any]]:
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        items = []
        for row in rows[1:]:
            if not row or not any(value not in (None, "") for value in row):
                continue
            values = list(row[:3])
            while len(values) < 3:
                values.append(None)
            items.append(
                {
                    "date": values[0],
                    "total_worlds": self._to_int(values[1]),
                    "new_worlds_today": self._to_int(values[2]),
                }
            )
        return items

    def _read_json(self, path: Path | str, default: Any) -> Any:
        file_path = Path(path)
        if not file_path.exists():
            return default
        try:
            return json.loads(file_path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError) as exc:
            logger.warning("Failed to read %s: %s", file_path, exc)
            return default

    def _write_json(self, path: Path | str, payload: Any) -> None:
        file_path = Path(path)
        file_path.parent.mkdir(parents=True, exist_ok=True)
        file_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    def _display_path(self, path: Path) -> str:
        try:
            return str(path.relative_to(self.repo_root)).replace("\\", "/")
        except ValueError:
            return str(path)

    def _topic_color(self, key: str, index: int) -> str:
        palette = ["#155e75", "#b45309", "#0f766e", "#7c3aed", "#2563eb", "#c2410c", "#1d4ed8", "#0f766e"]
        if key == "starriver" or key == "personal":
            return "#155e75"
        if key == "racing":
            return "#b45309"
        if key == "taiwan":
            return "#0f766e"
        return palette[index % len(palette)]

    def _date_score(self, value: Any) -> int:
        parsed = _parse_date(value) if value else None
        return int(parsed.timestamp()) if parsed else 0

    def _slugify(self, value: str) -> str:
        slug = re.sub(r"[^A-Za-z0-9_-]+", "_", value.strip())
        return slug.strip("_") or "default"

    def _csv_items(self, value: Any) -> list[str]:
        if value in (None, ""):
            return []
        if isinstance(value, str):
            raw_items = value.split(",")
        elif isinstance(value, (list, tuple, set)):
            raw_items = list(value)
        else:
            raw_items = [value]
        items: list[str] = []
        for item in raw_items:
            cleaned = str(item).strip()
            if cleaned and cleaned not in items:
                items.append(cleaned)
        return items

    def _load_world_lookup_for_group_memberships(
        self,
        memberships: list[dict[str, Any]],
    ) -> dict[str, dict[str, Any]]:
        world_ids = {
            self._clean_optional_text(item.get("world_id"))
            for item in memberships
            if self._clean_optional_text(item.get("world_id"))
        }
        if not world_ids:
            return {}
        lookup: dict[str, dict[str, Any]] = {}
        for world in self.load_worlds("db:all", sort="visits", direction="desc"):
            world_id = self._clean_optional_text(world.get("id"))
            if world_id and world_id in world_ids:
                lookup[world_id] = world
        return lookup

    def _enrich_group_world_membership(
        self,
        membership: dict[str, Any],
        world_lookup: dict[str, dict[str, Any]],
    ) -> dict[str, Any]:
        world_id = self._clean_optional_text(membership.get("world_id"))
        world = world_lookup.get(world_id or "", {})
        metrics = world.get("metrics") or self._calculate_metrics_for_world(world) if world else {}
        return {
            "group_id": membership.get("group_id"),
            "group_name": membership.get("group_name") or membership.get("group_id"),
            "group_region": membership.get("group_region"),
            "group_category": membership.get("group_category"),
            "world_id": world_id,
            "membership_role": membership.get("membership_role"),
            "linked_at": membership.get("linked_at"),
            "source_key": membership.get("source_key"),
            "world_name": world.get("name") or world_id,
            "author_id": world.get("author_id"),
            "author_name": world.get("author_name"),
            "visits": self._to_int(world.get("visits")),
            "favorites": self._to_int(world.get("favorites")),
            "heat": self._to_int(world.get("heat")),
            "popularity": self._to_int(world.get("popularity")),
            "updated_at": world.get("updated_at"),
            "publication_date": world.get("publication_date"),
            "thumbnail_url": world.get("thumbnail_url"),
            "world_url": world.get("world_url"),
            "tags": world.get("tags") or [],
            "favorite_rate": metrics.get("favorite_rate"),
            "visits_delta_7d": metrics.get("visits_delta_7d"),
            "visits_delta_1d": metrics.get("visits_delta_1d"),
        }

    def _optional_bool(self, value: Any) -> bool | None:
        if value in (None, ""):
            return None
        if isinstance(value, bool):
            return value
        text = str(value).strip().casefold()
        if text in {"1", "true", "yes", "y", "on"}:
            return True
        if text in {"0", "false", "no", "n", "off"}:
            return False
        return None

    def _world_search_label(
        self,
        *,
        search: str | None,
        tags: list[str],
        sort: str,
        active: bool,
        featured: bool | None,
    ) -> str:
        parts = []
        cleaned_search = (search or "").strip()
        if cleaned_search:
            parts.append(cleaned_search)
        if tags:
            parts.append("tag:" + ",".join(tags))
        if featured is True:
            parts.append("featured")
        if active:
            parts.append("active")
        parts.append(f"sort:{sort or 'popularity'}")
        return " | ".join(parts)

    def _normalise_2fa_method(self, value: Any) -> str:
        raw = str(value or "").strip()
        lowered = raw.casefold()
        aliases = {
            "emailotp": "emailotp",
            "email_otp": "emailotp",
            "email-otp": "emailotp",
            "emailOtp".casefold(): "emailotp",
            "otp": "otp",
            "totp": "totp",
        }
        return aliases.get(lowered, lowered)

    def _editable_source_key(self, source: str) -> str:
        if not source.startswith("db:"):
            raise ValueError("Only database sources can be edited.")
        if source == "db:all":
            raise ValueError("Choose a specific database source before editing or deleting records.")
        return source.removeprefix("db:")

    def _clean_optional_text(self, value: Any) -> str | None:
        if value is None:
            return None
        text = str(value).strip()
        if not text or text.lower() in {"none", "null", "nan"}:
            return None
        return text

    def _to_optional_int(self, value: Any) -> int | None:
        if value in (None, ""):
            return None
        if isinstance(value, bool):
            return int(value)
        try:
            cleaned = str(value).replace("%", "").replace(",", "")
            match = re.search(r"-?\d+(?:\.\d+)?", cleaned)
            if match:
                cleaned = match.group(0)
            return int(float(cleaned))
        except (TypeError, ValueError):
            return None

    def _to_int(self, value: Any) -> int:
        if value in (None, ""):
            return 0
        if isinstance(value, bool):
            return int(value)
        try:
            cleaned = str(value).replace("%", "").replace(",", "")
            match = re.search(r"-?\d+(?:\.\d+)?", cleaned)
            if match:
                cleaned = match.group(0)
            return int(float(cleaned))
        except (TypeError, ValueError):
            return 0

    def _to_float(self, value: Any) -> float | None:
        if value in (None, ""):
            return None
        try:
            return float(str(value).replace("%", "").replace(",", ""))
        except (TypeError, ValueError):
            return None

    def _parse_ratio(self, value: Any) -> float | None:
        ratio = self._to_float(value)
        if ratio is None:
            return None
        return round(ratio, 2)

    def _calculate_metrics_for_world(self, world: dict[str, Any]) -> dict[str, Any]:
        publication_dt = _parse_date(world.get("publication_date"))
        labs_dt = _parse_date(world.get("labs_publication_date"))
        updated_dt = _parse_date(world.get("updated_at"))
        fetched_dt = _parse_date(world.get("fetched_at")) or dt.datetime.now(dt.timezone.utc)

        visits = self._to_optional_int(world.get("visits"))
        favorites = self._to_optional_int(world.get("favorites"))

        favorite_rate = None
        if visits not in (None, 0) and favorites is not None:
            favorite_rate = round((favorites / visits) * 100, 2)

        visits_per_day = None
        days_since_publication = None
        if publication_dt and fetched_dt:
            age_days = max((fetched_dt - publication_dt).days, 0)
            days_since_publication = age_days
            if age_days > 0 and visits is not None:
                visits_per_day = round(visits / age_days, 2)

        labs_to_publication_days = None
        if publication_dt and labs_dt:
            labs_to_publication_days = (publication_dt - labs_dt).days

        days_since_update = None
        if updated_dt and fetched_dt:
            days_since_update = max((fetched_dt - updated_dt).days, 0)

        explicit_metrics = world.get("metrics", {}) if isinstance(world.get("metrics"), dict) else {}
        if explicit_metrics.get("favorite_rate") is not None and favorite_rate is None:
            favorite_rate = self._to_float(explicit_metrics.get("favorite_rate"))
        if explicit_metrics.get("labs_to_publication_days") is not None and labs_to_publication_days is None:
            labs_to_publication_days = self._to_optional_int(explicit_metrics.get("labs_to_publication_days"))
        if explicit_metrics.get("days_since_update") is not None and days_since_update is None:
            days_since_update = self._to_optional_int(explicit_metrics.get("days_since_update"))
        if explicit_metrics.get("days_since_publication") is not None and days_since_publication is None:
            days_since_publication = self._to_optional_int(explicit_metrics.get("days_since_publication"))
        if explicit_metrics.get("visits_per_day") is not None and visits_per_day is None:
            visits_per_day = self._to_float(explicit_metrics.get("visits_per_day"))

        return {
            "favorite_rate": favorite_rate,
            "labs_to_publication_days": labs_to_publication_days,
            "days_since_update": days_since_update,
            "days_since_publication": days_since_publication,
            "visits_per_day": visits_per_day,
        }

    def _calculate_new_worlds_today(self, worlds: list[dict[str, Any]]) -> int:
        today = dt.datetime.now(dt.timezone.utc).date()
        count = 0
        for world in worlds:
            publication = _parse_date(self._new_world_event_date(world))
            if publication and publication.astimezone(dt.timezone.utc).date() == today:
                count += 1
        return count

    def _new_world_event_date(self, world: dict[str, Any] | None) -> str | None:
        if not world:
            return None
        publication_date = self._clean_optional_text(world.get("publication_date"))
        if publication_date:
            return publication_date
        return self._clean_optional_text(world.get("labs_publication_date"))

    def _within_days(self, value: Any, days: int, now: dt.datetime) -> bool:
        parsed = _parse_date(value) if value else None
        if not parsed:
            return False
        return (now - parsed.astimezone(dt.timezone.utc)).days <= days

    def _date_bucket(self, value: Any) -> str | None:
        parsed = _parse_date(value) if value else None
        if not parsed:
            return None
        return parsed.astimezone(dt.timezone.utc).strftime("%Y-%m-%d")

    def _public_db_source_key(self, source_key: str) -> str:
        return f"db:{source_key}"

    def _label_for_db_source(self, source_key: str) -> str:
        if source_key in IMPORT_SOURCE_LABELS:
            return IMPORT_SOURCE_LABELS[source_key]
        for job_key, config in self._load_job_configs().items():
            resolved = self._resolve_job_config(job_key, config)
            if resolved["source_key"] == source_key:
                return resolved["label"]
        if source_key.startswith("manual:keyword:"):
            return f"Manual keyword: {source_key.removeprefix('manual:keyword:')}"
        if source_key.startswith("manual:user:"):
            return f"Manual creator: {source_key.removeprefix('manual:user:')}"
        if source_key.startswith("manual:fixed:"):
            return f"Manual fixed: {source_key.removeprefix('manual:fixed:')}"
        return source_key

    def _calculate_rate_limit_cooldown(
        self,
        *,
        retry_after_seconds: int,
        now: dt.datetime | None = None,
    ) -> int:
        now = now or dt.datetime.now(dt.timezone.utc)
        recent_count = self.storage.count_rate_limit_events_since((now - dt.timedelta(hours=6)).isoformat())
        base = max(retry_after_seconds * 6, 1800)
        penalty_steps = min(max(recent_count, 0), 4)
        return min(base + (penalty_steps * 900), 86400)

    def _build_rate_limit_strategy_hint(self, *, count_24h: int, remaining_seconds: int) -> str:
        if remaining_seconds > 0:
            wait_minutes = max(round(remaining_seconds / 60), 1)
            return f"Cooldown active. Wait about {wait_minutes} minutes before the next VRChat crawl, and avoid using two accounts from the same IP."
        if count_24h >= 5:
            return "Frequent 429s in the last 24h. Reduce high-volume jobs to 12h or 1d, and prefer creator follow-ups over broad keyword sweeps."
        if count_24h >= 1:
            return "Recent 429s detected. Space out manual searches and lower per-keyword limits before the next run."
        return "No recent 429 events recorded."

    def _decorate_run(self, run: dict[str, Any] | None) -> dict[str, Any] | None:
        if run is None:
            return None
        item = dict(run)
        item["source"] = self._public_db_source_key(item["source_key"])
        item["label"] = self._label_for_db_source(item["source_key"])
        return item
