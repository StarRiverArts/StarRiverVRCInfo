import sys
from pathlib import Path


def test_save_worlds_headers(tmp_path):
    base = Path(__file__).resolve().parent.parent / "world_info"
    sys.path.append(str(base))
    import actions as actions_module
    from openpyxl import load_workbook

    file = tmp_path / "out.xlsx"
    actions_module.save_worlds([], file)
    wb = load_workbook(file)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    assert headers == ["爬取日期"] + actions_module.METRIC_COLS
